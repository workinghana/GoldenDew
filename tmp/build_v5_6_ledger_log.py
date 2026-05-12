"""V5-6 LoyaltyLedger 흐름 정리 — 회원 TEST02007.

시나리오 단계별로 LoyaltyLedger 변경을 시간순 정렬하고 누적 잔액 변화 표시.
「판매 삭제 (DELETE SAL04043)」 시점에 ledger 변경이 0건인 사실을 명시.
"""

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V5-6 Ledger 분석.xlsx"

# 시간순으로 정렬된 LoyaltyLedger + 시나리오 단계 매핑
# (시각, TJ Id 약식, EventType, 포인트 종류, 포인트, 시나리오 단계, 부가설명)
LEDGER = [
    # (KST_time, ledger_id, tj_id, event, type_label, points, stage, note)
    ("12:44:09", "05wJO0000004LgzYAE", "0lVJO000000CAtF2AW", "Credit", "이벤트포인트 (1nThpYAE)", 80000, "(직전 시도 잔재)", "SAL04041/SAL04042 시도 과정의 적립 기록"),
    ("12:44:58", "05wJO0000004LkDYAU", "0lVJO000000CAy52AG", "Debit",  "(CD 04 / Order)", 28622, "1) 주문 SOR04010", "포인트 10만 사용 분개 #1"),
    ("12:44:58", "05wJO0000004LkEYAU", "0lVJO000000CAy52AG", "Debit",  "(CD 04 / Order)", 71378, "1) 주문 SOR04010", "포인트 10만 사용 분개 #2 (합 100,000)"),
    ("12:50:21", "05wJO0000004LnRYAU", "0lVJO000000CB1K2AW", "Credit", "구매포인트 (BAZlYAO × 2.00)", 37400, "2) 판매 SAL04043", "구매포인트 적립"),
    ("12:50:21", "05wJO0000004LnSYAU", "0lVJO000000CB1K2AW", "Credit", "사은포인트 (BAY9YAO × 3.00)", 56100, "2) 판매 SAL04043", "사은포인트 적립 (합 93,500)"),
    ("12:54:26", "05wJO0000004LkFYAU", "0lVJO000000CAy92AG", "Debit",  "구매포인트 (BAZlYAO)", 20000, "3) 부분반품 SAL04114", "구매포인트 적립 cancel"),
    ("12:54:26", "05wJO0000004LkGYAU", "0lVJO000000CAy92AG", "Debit",  "사은포인트 (BAY9YAO)", 30000, "3) 부분반품 SAL04114", "사은포인트 적립 cancel (합 50,000)"),
    ("12:55:53", "05wJO0000004LqgYAE", "0lVJO000000CB7l2AG", "Credit", "사은포인트 (BAY9YAO)", 30000, "4) 부분반품 삭제", "cancel 분 복원 #1"),
    ("12:55:53", "05wJO0000004LqfYAE", "0lVJO000000CB7l2AG", "Credit", "구매포인트 (BAZlYAO)", 20000, "4) 부분반품 삭제", "cancel 분 복원 #2 (합 50,000)"),
    ("13:06:18", "05wJO0000004KxrYAE", "0lVJO000000C9va2AC", "Credit", "(BAUvYAO / Order)", 28622, "5) 완전반품 SAL04115", "Order 포인트 사용 복원 #1"),
    ("13:06:18", "05wJO0000004KxqYAE", "0lVJO000000C9va2AC", "Credit", "이벤트포인트 (1nThpYAE)", 71378, "5) 완전반품 SAL04115", "Order 포인트 사용 복원 #2 (합 100,000)"),
    ("13:06:19", "05wJO0000004KxsYAE", "0lVJO000000C9ve2AC", "Debit",  "구매포인트 (BAZlYAO)", 37400, "5) 완전반품 SAL04115", "SAL04043 구매 적립 cancel"),
    ("13:06:19", "05wJO0000004KxtYAE", "0lVJO000000C9ve2AC", "Debit",  "사은포인트 (BAY9YAO)", 56100, "5) 완전반품 SAL04115", "SAL04043 사은 적립 cancel (합 93,500)"),
    ("13:10:19", "05wJO0000004IxgYAE", "0lVJO000000C7oz2AC", "Credit", "구매포인트 (BAZlYAO × 2.00)", 33660, "6) 완전반품 삭제", "적립 복원 분할 1 (구매)"),
    ("13:10:19", "05wJO0000004IxhYAE", "0lVJO000000C7oz2AC", "Credit", "사은포인트 (BAY9YAO × 3.00)", 50490, "6) 완전반품 삭제", "적립 복원 분할 1 (사은) (소계 84,150)"),
    ("13:10:34", "05wJO0000004LvVYAU", "0lVJO000000CBCb2AO", "Debit",  "(BAUvYAO / Order)", 28622, "6) 완전반품 삭제", "Order 포인트 재차감 #1"),
    ("13:10:34", "05wJO0000004LvWYAU", "0lVJO000000CBCb2AO", "Debit",  "이벤트포인트 (1nThpYAE)", 71378, "6) 완전반품 삭제", "Order 포인트 재차감 #2 (합 100,000)"),
    ("13:10:50", "05wJO0000004Lx7YAE", "0lVJO000000CBED2A4", "Credit", "구매포인트 (BAZlYAO × 2.00)", 3740,  "6) 완전반품 삭제", "적립 복원 분할 2 (구매)"),
    ("13:10:50", "05wJO0000004Lx8YAE", "0lVJO000000CBED2A4", "Credit", "사은포인트 (BAY9YAO × 3.00)", 5610,  "6) 완전반품 삭제", "적립 복원 분할 2 (사은) (소계 9,350 / 분할 1+2 합 93,500)"),
    # 7) 판매 삭제 SAL04043 — Ledger 변경 0건 (가상 행)
    ("13:18:57", "(없음)", "(audit a12JO000000MXdIYAW)", "—", "—", 0, "7) 판매 삭제 SAL04043 ⚠️", "DELETE Sale 호출 / Ledger 변경 0건 → 포인트 꼬임 발생"),
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger 흐름"

    # ===== 상단 노트 =====
    note = (
        "※ V5-6 「판매 삭제 잘못해서 포인트 꼬임」 시나리오 — 회원 TEST02007 / 2026-05-09\n"
        "   주문 SOR04010 → 판매 SAL04043 → 부분반품 SAL04114 → 부분반품 삭제\n"
        "   → 완전반품 SAL04115 → 완전반품 삭제 → 판매 삭제 SAL04043\n"
        "   ⚠ 마지막 7) 판매 삭제 시점에 LoyaltyLedger 변경이 0건 발생 → 회원 잔액 −6,500P 로 영구 비정상 상태"
    )
    ws.cell(row=1, column=1, value=note)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=1, column=1).font = Font(color="9C5700", italic=True, bold=True)
    ws.row_dimensions[1].height = 70

    # ===== 헤더 =====
    headers = [
        "#",
        "시각 (KST)",
        "시나리오 단계",
        "EventType",
        "포인트 종류",
        "포인트",
        "Credit 누적",
        "Debit 누적",
        "회원 잔액 변화 (V5-6 한정)",
        "비고 / TJ / Ledger Id",
    ]
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 32

    # ===== 본문 =====
    body_align = Alignment(vertical="top", wrap_text=True)
    num_align = Alignment(vertical="top", horizontal="right")

    credit_sum = 0
    debit_sum = 0

    # 직전 시도 잔재 (12:44:09 80K event credit) 는 V5-6 본 시나리오 직전이라 별도 표시
    # 회원 잔액 변화 누적은 12:44:58 부터 (Order 부터)
    balance_start_idx = 1  # index 1 (12:44:58 Debit) 부터 누적 시작 / 12:44:09 prior 는 0번째

    thin = Side(border_style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    warn_fill = PatternFill("solid", fgColor="FCE4E4")
    stage_colors = {
        "1)": "DDEBF7",
        "2)": "E2EFDA",
        "3)": "FFF2CC",
        "4)": "FFE699",
        "5)": "DEEBF6",
        "6)": "FCE4D6",
        "7)": "FCE4E4",
        "(": "F2F2F2",
    }

    for i, row in enumerate(LEDGER, start=1):
        kst, ledger_id, tj_id, event, type_label, points, stage, note_text = row
        r = i + 2  # row index (1=note, 2=header, 3+ data)

        # 누적 계산
        if i >= balance_start_idx + 1:  # skip the prior "12:44:09" row (i=1)
            if event == "Credit":
                credit_sum += points
            elif event == "Debit":
                debit_sum += points
        balance = credit_sum - debit_sum if i >= balance_start_idx + 1 else None

        # row fill by stage
        prefix = stage[:2] if len(stage) >= 2 else "("
        fill_color = stage_colors.get(prefix, "FFFFFF")
        if "⚠" in stage or "삭제 SAL04043" in stage:
            fill_color = "FCE4E4"
        row_fill = PatternFill("solid", fgColor=fill_color)

        cells = [
            (1, i),
            (2, kst),
            (3, stage),
            (4, event),
            (5, type_label),
            (6, points if points else ""),
            (7, credit_sum if i >= balance_start_idx + 1 else ""),
            (8, debit_sum if i >= balance_start_idx + 1 else ""),
            (9, balance if balance is not None else ""),
            (10, f"TJ {tj_id}\nLedger {ledger_id}\n— {note_text}"),
        ]
        for col, val in cells:
            c = ws.cell(row=r, column=col, value=val)
            c.fill = row_fill
            c.alignment = num_align if col in (6, 7, 8, 9) else body_align
            c.border = border
            if "⚠" in stage and col == 3:
                c.font = Font(bold=True, color="C00000")
            if col == 6 and event == "Debit":
                c.font = Font(color="C00000")
            elif col == 6 and event == "Credit":
                c.font = Font(color="2E7D32")

        ws.row_dimensions[r].height = 36

    # ===== 요약 행 =====
    summary_r = len(LEDGER) + 4
    ws.cell(row=summary_r, column=1, value="합계 (V5-6 한정)")
    ws.merge_cells(start_row=summary_r, start_column=1, end_row=summary_r, end_column=5)
    ws.cell(row=summary_r, column=1).font = Font(bold=True)
    ws.cell(row=summary_r, column=1).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=summary_r, column=6, value="").alignment = num_align
    ws.cell(row=summary_r, column=7, value=credit_sum).font = Font(bold=True, color="2E7D32")
    ws.cell(row=summary_r, column=7).alignment = num_align
    ws.cell(row=summary_r, column=8, value=debit_sum).font = Font(bold=True, color="C00000")
    ws.cell(row=summary_r, column=8).alignment = num_align
    ws.cell(row=summary_r, column=9, value=credit_sum - debit_sum).font = Font(bold=True, color="C00000")
    ws.cell(row=summary_r, column=9).alignment = num_align
    ws.cell(row=summary_r, column=10, value="기대값: 0  /  실측: −6,500  →  꼬임 6,500P")
    ws.cell(row=summary_r, column=10).font = Font(bold=True, color="C00000")

    # ===== 결론 노트 =====
    conclusion_r = summary_r + 2
    conclusion = (
        "[진단]\n"
        " - 1)~6) 단계는 Credit/Debit 가 시점마다 정합적으로 대칭됨 (부분반품 ↔ 부분반품 삭제, 완전반품 ↔ 완전반품 삭제 모두 net 0)\n"
        " - 시나리오 종료 후 기대값: net 변화 0 (즉 Order 시점 대비 회원 잔액 동일해야 함)\n"
        " - 그러나 7) 판매 삭제 (DELETE SAL04043) 시점에 LoyaltyLedger 변경이 0건 → −6,500P 잔여\n"
        " - 원인 후보: Sale DELETE 분기에서 ledger cancel 호출이 누락 (SaleDeposit DELETE 가 ledger 정리를 담당하나 V5-6 흐름에는 SaleDeposit DELETE 가 없음)"
    )
    ws.cell(row=conclusion_r, column=1, value=conclusion)
    ws.merge_cells(start_row=conclusion_r, start_column=1, end_row=conclusion_r, end_column=10)
    ws.cell(row=conclusion_r, column=1).fill = PatternFill("solid", fgColor="F2F2F2")
    ws.cell(row=conclusion_r, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.cell(row=conclusion_r, column=1).font = Font(color="333333")
    ws.row_dimensions[conclusion_r].height = 100

    # ===== 컬럼 폭 =====
    widths = [5, 12, 26, 10, 30, 12, 14, 14, 22, 64]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "C3"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] {len(LEDGER)} rows / Credit {credit_sum:,} / Debit {debit_sum:,} / net {credit_sum-debit_sum:+,}")


if __name__ == "__main__":
    main()
