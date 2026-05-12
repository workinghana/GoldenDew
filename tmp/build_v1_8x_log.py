"""V1 (8-x) Log.xlsx — 회원 TEST03006 / 5/10 18:00 ~ 5/11 02:00 KST 전체 활동.

V1 8-1 ~ 8-6 시나리오 (쿠폰 9908EY4T + 포인트 50,000 + 부분반품/삭제/재진행/통반품) 의
정확한 매칭 로그가 캐시된 시간 범위에 없어, 회원 TEST03006 의 같은 시간대 전체 활동을
부분 매핑으로 정리. (V5-1 시나리오 흐름과 일부 중복)
"""
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V1 (8-x) Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v1_8x_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V1"
MEMBER_NO = "TEST03006"
SHEET_NAME = "V1 (8-x)"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower().replace("_", "")
    if "memberhelpinfo" in ac:
        return "회원 도움창"
    if "memberapi" in ac:
        return "회원"
    if "voucherhelp" in ac:
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac:
        return "포인트 (적용 조회)"
    if "saledeposit" in ac:
        return "판매 입금"
    if "returndeposit" in ac:
        return "반품 입금"
    if "orderapi" in ac:
        return "주문"
    if "saleapi" in ac:
        return "판매"
    if "returnapi" in ac:
        return "반품"
    if "pointcredit" in ac:
        return "포인트 (지급)"
    if "pointdebit" in ac:
        return "포인트 (사용/차감)"
    if "erpasync" in ac or "asyncjob" in ac:
        return "ERP 콜아웃"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    "a12JO000000MdfRYAS": (
        "사전",
        "ERP 포인트 적립 콜아웃 — PointSave 30,000P (cd_type 07)",
        "ERP 콜아웃 PointSave 30,000P (시간 17:57:50 KST, TEST03006)",
    ),
    "a12JO000000Mdh3YAC": (
        "사전",
        "ERP 쿠폰 동기화 콜아웃 — COP2026MY0004 (mil 00218246)",
        "ERP 콜아웃 CouponDetail COP2026MY0004 (시간 17:58:02 KST)",
    ),
    "a12JO000000Me24YAC": (
        "8-1 사전",
        "주문 등록 — SOR04011 (1,000,000원 / 포인트 + 쿠폰 + 카드)",
        (
            "주문 번호 : SOR04011  (시간 : 2026-05-10 18:02:52 KST)\n"
            "- 회원 : TEST03006\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 프로모션 : PRO202605041079\n"
            "- 응답 : 「주문 등록 완료」 — PT_USE 10,000\n"
            "- 시나리오 매핑 : 8-1 사전 — 쿠폰/포인트 적용 주문 (V5-1 시나리오와 동일 흐름)"
        ),
    ),
    "a12JO000000Me3kYAC": (
        "8-1",
        "판매 등록 — SAL04044 (PT_TARGET 9,600P)",
        (
            "판매 번호 : SAL04044  (시간 : 2026-05-10 18:07:25 KST)\n"
            "- 원거래 주문 번호 : SOR04011\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 9,600P (구매포인트 2배)\n"
            "- 시나리오 매핑 : 8-1 쿠폰/포인트/에코포인트 적용 판매\n"
            "- ※ 시나리오 메모상 적립 9,200 + 27,600 + 6,000 = 42,800P 와 차이 있음"
        ),
    ),
    "a12JO000000McmcYAC": (
        "8-1 에코",
        "에코포인트 6,000P 지급 — SAL04044 대상",
        (
            "포인트 지급 — 이벤트포인트 6,000P  (시간 : 2026-05-10 18:09:23 KST)\n"
            "- 회원 : TEST03006  /  대상 판매 : SAL04044\n"
            "- 응답 : 「포인트 지급 성공」 — PT_TARGET 6,000\n"
            "- 시나리오 매핑 : 8-1 에코포인트 6,000P 적립"
        ),
    ),
    "a12JO000000MdR2YAK": (
        "8-2",
        "부분반품 — SAL04116 (PT_USE 4,950 복원)",
        (
            "반품 번호 : SAL04116  /  원거래 판매 번호 : SAL04044\n"
            "- 시간 : 2026-05-10 18:14:13 KST\n"
            "- 실결제 금액 : 500,000원 (1개 부분반품)\n"
            "- 응답 : 「반품 등록 완료」 — PT_USE 4,950 복원\n"
            "- 시나리오 매핑 : 8-2 부분반품 진행 (포인트 부분 환불)\n"
            "  ※ 시나리오 메모상 부분반품 480,000원과 다름"
        ),
    ),
    "a12JO000000Mdh8YAC": (
        "8-6",
        "통반품 — SAL04117 (PT_USE 4,650 복원)",
        (
            "반품 번호 : SAL04117  /  원거래 판매 번호 : SAL04044\n"
            "- 시간 : 2026-05-10 18:18:45 KST\n"
            "- 실결제 금액 : 500,000원 (1개 통반품)\n"
            "- 응답 : 「반품 등록 완료」 — PT_USE 4,650 복원\n"
            "- 시나리오 매핑 : 8-6 통반품\n"
            "  ※ 시나리오 메모상 통반품 920,000원과 다름 — 8-3/8-4/8-5 (부분반품 삭제/재진행) 단계 누락"
        ),
    ),
    "a12JO000000Me0TYAS": (
        "부수",
        "ERP 쿠폰 동기화 콜아웃 — COP2026MY0003 (mil 00218249)",
        "ERP 콜아웃 CouponDetail COP2026MY0003 (시간 18:21:01 KST)",
    ),
    "a12JO000000MeLQYA0": (
        "부수",
        "다른 주문 등록 — SOR04012 (500,000원 / PRO202604161041)",
        (
            "주문 번호 : SOR04012  (시간 : 2026-05-10 18:26:01 KST)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 프로모션 : PRO202604161041 (다른 프로모션)\n"
            "- 응답 : 「주문 등록 완료」 — PT_USE 10,000\n"
            "- 시나리오 매핑 : 부수 — 8-x 시나리오와 별개 흐름"
        ),
    ),
    "a12JO000000MeQGYA0": (
        "부수",
        "다른 판매 등록 — SAL04046 (500,000원, PT_TARGET 4,850)",
        (
            "판매 번호 : SAL04046  (시간 : 2026-05-10 18:30:51 KST)\n"
            "- 원거래 주문 : SOR04012\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 4,850P\n"
            "- 시나리오 매핑 : 부수"
        ),
    ),
    "a12JO000000MeYIYA0": (
        "부수",
        "판매 수정 — SAL04046 (PT_TARGET 150P)",
        "SAL04046 수정 (시간 18:36:25 KST, PT_TARGET 150)",
    ),
    "a12JO000000MegLYAS": (
        "부수",
        "판매 입금 삭제 — DELETE DEP04093 (SAL04046)",
        (
            "Apex 클래스 : GdSaleDepositApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-10 18:43:49 KST\n"
            "- 대상 입금 번호 : DEP04093  /  DeposiSeqNo : 4\n"
            "- 응답 : 「판매 입금 삭제 삭제 완료」 — PT_USE 4,365 복원 (SAL04046)\n"
            "- 시나리오 매핑 : 부수"
        ),
    ),
    "a12JO000000MkDtYAK": (
        "부수",
        "에코포인트 6,000P 지급 — SAL04066 (5/11 새벽)",
        "PointCredit 6,000P SAL04066 (시간 5/11 01:11:22 KST)",
    ),
    "a12JO000000MhrIYAS": (
        "부수",
        "판매 전체 삭제 — DELETE SAL04066 (PT_USE 6,000 복원)",
        (
            "Apex 클래스 : GdSaleApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-11 01:17:58 KST\n"
            "- 대상 판매 번호 : SAL04066\n"
            "- 응답 : 「판매 전체 삭제 완료」 — PT_USE 6,000 복원\n"
            "- 시나리오 매핑 : 부수"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 50, 95, 16, 24, 42, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
