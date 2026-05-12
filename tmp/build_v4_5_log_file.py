"""V4 Log (5-1~5-5).xlsx — 회원 173000400470 VVIP, 시나리오 V4 / 5-1 ~ 5-5.

「5. 예술의전당 프로모션 이용」 — [TEST] 통합시나리오V4_계열별_특별_행사 / 10% 할인 쿠폰 사용
- 5-1 판매저장 — 구매 21,376P / 사은 32,064P 적립 예상
- 5-2 에코포인트 지급 — 9,000P
- 5-3 부분반품 시도 — 정률쿠폰 사용 → 부분 반품 X
- 5-4 통반품 시도 — 0P 으로 복원, 사용 포인트 복원
- 5-5 재판매처리 — 구매 7,126 + 사은 10,689 = 총 17,815P

※ audit 시점 : 2026-05-07 17:29 ~ 17:40 KST (UTC 08:29 ~ 08:40)
※ 회원 173000400470 의 실제 audit 기록은 PromotionNo PRO202604161041 사용 — 시나리오 기재 값과 적립 포인트가 다름 (참고 노트 참고)
"""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V4 Log (5-1~5-5).xlsx"
MATCHED_CSV = ROOT / "tmp" / "v4_5_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V4"
MEMBER_NO = "173000400470"
SHEET_NAME = "V4 (5-1~5-5)"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac.replace("_", ""):
        return "포인트 (적용 조회)"
    if "saledeposit" in ac.replace("_", ""):
        return "판매 입금"
    if "returndeposit" in ac.replace("_", ""):
        return "반품 입금"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    return "POST"


SCENARIO = {
    # ───── 5-1 : 판매저장 (예술의전당 5% + VVIP 5% 쿠폰) ─────
    "a12JO000000LwM4YAK": (
        "5-1",
        "판매저장 — 예술의전당 5% 할인 + VVIP 5% 쿠폰 사용 후 적립 확인",
        (
            "판매 번호 : SAL202605070225  (시간 : 2026-05-07 17:29:18 KST)\n"
            "- 매장 코드 : 99998\n"
            "- 실결제 금액 : 1,000,000원  (정상 1,200,000 → 5% 할인 → VVIP 5% 쿠폰)\n"
            "- 프로모션 번호 : PRO202604161041\n"
            "- 판매 항목 (3개 품목 → audit 에는 2건만 기록) :\n"
            "    1) 상품 코드 2A2600001  /  수량 1  /  순매가 단가 500,000\n"
            "    2) 상품 코드 2A2600001  /  수량 1  /  순매가 단가 500,000\n"
            "- 거래 원장 (입금 번호 DEP202605070282) :\n"
            "    1) 결제 수단 93 (정률쿠폰?)  /  입금 구분 02  /  거래 금액 30,000원\n"
            "    2) 결제 수단 01 (현금)  /  입금 구분 02  /  거래 금액 920,000원\n"
            "    3) 결제 수단 81 (포인트)  /  입금 구분 02  /  거래 금액 50,000원\n"
            "- 응답 적립 포인트 :\n"
            "    · 구매포인트 × 2.00 → 18,400P  (CD 01)\n"
            "    · 사은포인트 × 3.00 → 27,600P  (CD 01)\n"
            "- 시나리오 기준 : 「구매 포인트 21,376P (VVIP 2%) + 사은 포인트 32,064P」\n"
            "- ※ audit 적립 합계 (46,000) 와 시나리오 기재 (53,440) 차이 — 실제 결제 금액 / 할인 적용 방식 차이로 추정. 실제 ERP 적립 내역 화면과 비교 필요"
        ),
    ),
    # ───── 5-2 : 에코포인트 지급 ─────
    "a12JO000000LwsDYAS": (
        "5-2",
        "에코포인트 지급 (cd_type_point 05)",
        (
            "Apex 클래스 : GdPointCreditApiControllerV1 — /gd/v1/point/credit\n"
            "- 시간 : 2026-05-07 17:29:28 KST  (5-1 판매 직후 10초)\n"
            "- 회원 : 173000400470\n"
            "- 대상 판매 번호 : SAL202605070225\n"
            "- PointsList : [ point 6,000P  /  cd_type_point 05 (에코포인트) ]\n"
            "- 응답 메시지 : 「포인트 지급 성공」 / 적립 6,000P (TX_REMARK 「이벤트포인트」)\n"
            "- 시나리오 기준 : 「에코 포인트 지급 : 9,000P」\n"
            "- ※ audit 값 6,000P, 시나리오 기재 9,000P 차이 — 별도 확인"
        ),
    ),
    # ───── 5-3 : 부분반품 시도 (정률쿠폰 사용 → 부분반품 X) ─────
    "a12JO000000LvOLYA0": (
        "5-3",
        "부분반품 시도 — 정률쿠폰 사용 건 부분 반품 X (1차 시도)",
        (
            "반품 번호 : SAL202605070226  /  원거래 판매 번호 : SAL202605070225\n"
            "- 시간 : 2026-05-07 17:31:42 KST\n"
            "- 실결제 금액 : 500,000원 (1개 품목 부분반품)\n"
            "- 판매 항목 : 상품 코드 2A2600001 / 수량 -1\n"
            "- 거래 원장 (입금 번호 DEP202605070283) :\n"
            "    1) 결제 수단 81 (포인트) / 입금 구분 03 / 40,000원\n"
            "    2) 결제 수단 93 (정률쿠폰?) / 입금 구분 03 / 30,000원\n"
            "    3) 결제 수단 01 (현금) / 입금 구분 03 / 430,000원\n"
            "- 응답 메시지 : 「반품 등록 완료」 (실제 처리됨 — 응답 success: true)\n"
            "- 응답 적립 포인트 소멸 : CD 01 / 8,600P + CD 06 / 12,900P (총 21,500P 소멸)\n"
            "- 시나리오 기준 : 「정률쿠폰 사용되어 부분 반품 X」 (= 시스템이 차단해야 함)\n"
            "- ※ audit 상으로는 반품이 정상 등록되어 있어 시나리오 의도(차단)와 결과 불일치 — 정률쿠폰 차단 로직 누락 가능성"
        ),
    ),
    # ───── 5-4 : 통반품 시도 (전체 반품) ─────
    "a12JO000000LrHWYA0": (
        "5-4",
        "통반품 시도 — 0P 으로 복원, 사용 포인트 복원",
        (
            "통반품 번호 : SAL202605070232  /  원거래 판매 번호 : SAL202605070225\n"
            "- 시간 : 2026-05-07 17:37:46 KST\n"
            "- 실결제 금액 : 1,000,000원 (전체 반품)\n"
            "- 판매 항목 : 상품 코드 2A2600001 × 2 (각 -1)\n"
            "- 거래 원장 (입금 번호 DEP202605070288) :\n"
            "    1) 결제 수단 81 (포인트) / 입금 구분 03 / 50,000원  ← 포인트 사용 복원\n"
            "    2) 결제 수단 93 (정률쿠폰) / 입금 구분 03 / 30,000원  ← 쿠폰 환불\n"
            "    3) 결제 수단 01 (현금) / 입금 구분 03 / 920,000원  ← 현금 환불\n"
            "- 응답 메시지 : 「반품 등록 완료」\n"
            "- 응답 적립 포인트 소멸 : CD 01 / 18,400P + CD 06 / 27,600P (= 5-1 적립 전체 46,000P 소멸)\n"
            "- 시나리오 검증 : 「다시 0P 으로 복원, 사용 포인트 복원」 ✓\n"
            "- ※ 통반품 시 에코포인트 9,000원 모두 소멸 처리도 시나리오 기준이나 audit 의 CancelPointsCreditedList 에 CD 05 (에코) 항목 미확인 — 별도 확인"
        ),
    ),
    # ───── 5-5 : 재판매처리 ─────
    "a12JO000000LwyoYAC": (
        "5-5",
        "재판매처리 — 실구매 제품 1개 (구매 + 사은 적립)",
        (
            "판매 번호 : SAL202604300266  (시간 : 2026-05-07 17:40:31 KST)\n"
            "- 실결제 금액 : 500,000원 (1개 품목)\n"
            "- 프로모션 번호 : PRO202604161041\n"
            "- 판매 항목 : 상품 코드 2A2600001 / 수량 1 / 순매가 단가 500,000\n"
            "- 거래 원장 (입금 번호 DEP202605070290) :\n"
            "    · 결제 수단 01 (현금) / 입금 구분 02 / 500,000원 (단순 현금 결제)\n"
            "- 응답 적립 포인트 :\n"
            "    · 구매포인트 × 2.00 → 10,000P\n"
            "    · 사은포인트 × 3.00 → 15,000P\n"
            "- 시나리오 기준 : 「구매 7,126P + 사은 10,689P = 총 17,815P 적립」\n"
            "- ※ audit 적립 합계 (25,000) 와 시나리오 기재 (17,815) 차이 — 시나리오 가정한 기준 결제액과 실 결제액 차이로 추정"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    def num_sort_key(item):
        no = item[1][0]
        parts = no.split("-")
        try:
            return tuple(int(p) for p in parts)
        except ValueError:
            return (99, 99, 99)

    sorted_items = sorted(SCENARIO.items(), key=num_sort_key)

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 10, 40, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
