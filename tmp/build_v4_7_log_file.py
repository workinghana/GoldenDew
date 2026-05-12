"""V4-7 Log.xlsx — 회원 123002600463 SVIP, 시나리오 V4 / 4-7 (아울렛 판매 → 삭제 → 재판매 → 통반품).

흐름:
- 4-7-a 아울렛 주문 등록 (SOR202605080147)
- 4-7-b 아울렛 판매 (SAL202605080164)
- 4-7-c 아울렛 판매 삭제 (SaleNo SAL164 DELETE)
- 4-7-d 아울렛 재판매 (SAL202605080167)
- 4-7-e 아울렛 통반품 (SAL202605080170)
"""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V4-7 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "_v4_4_refresh.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V4"
MEMBER_NO = "123002600463"
SHEET_NAME = "V4-7"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac.replace("_", ""):
        return "포인트 (적용 조회)"
    if "saledeposit" in ac.replace("_", ""):
        return "판매 입금"
    if "returndeposit" in ac.replace("_", ""):
        return "반품 입금"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    # SaleNo only URL (no /gd/v1/...) = DELETE
    if isinstance(url, str) and url.startswith('{"SaleNo'):
        return "DELETE"
    return "POST"


SCENARIO = {
    # 4-7-a 아울렛 주문 등록
    "a12JO000000MEPRYA4": (
        "4-7-a",
        "아울렛 주문 등록 — SOR202605080147",
        (
            "주문 번호 : SOR202605080147  (시간 : 2026-05-08 14:43:12 KST)\n"
            "- 매장 코드 : 99997 (아울렛)  /  유형 : 01 (계약)\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 프로모션 번호 : (없음)\n"
            "- 주문 항목 :\n"
            "    1) 상품 코드 2A2600001  /  수량 1  /  순매가 단가 500,000  /  할인율 0%\n"
            "    2) 상품 코드 3A2600006  /  수량 1  /  순매가 단가 500,000  /  할인율 0%\n"
            "- 거래 원장 (입금 번호 DEP202605080172) :\n"
            "    1) 결제 수단 01 (현금) / 입금 구분 01 / 거래 금액 100,000원\n"
            "    2) 결제 수단 90 (쿠폰) / 입금 구분 01 / 거래 금액 100,000원 (쿠폰 COP2026JA0028;-;25)\n"
            "- 응답 메시지 : 「주문 등록 완료」  /  isGoldenBar: False  /  storeType: 20 (아울렛)\n"
            "- 검증 : 아울렛 주문 정상 등록"
        ),
    ),
    # 4-7-b 아울렛 판매
    "a12JO000000MEsOYAW": (
        "4-7-b",
        "아울렛 판매 — SAL202605080164",
        (
            "판매 번호 : SAL202605080164  /  원거래 주문 번호 : SOR202605080147\n"
            "- 시간 : 2026-05-08 14:47:38 KST\n"
            "- 매장 코드 : 99997 (아울렛)\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 판매 항목 :\n"
            "    1) 상품 코드 2A2600001  /  수량 1  /  순매가 단가 500,000  /  할인율 0%\n"
            "    2) 상품 코드 3A2600006  /  수량 1  /  순매가 단가 500,000  /  할인율 0%  /  제외 여부 true\n"
            "- 거래 원장 :\n"
            "    · DEP202605080172-1 : 현금 01 / 계약금 01 / 100,000원\n"
            "    · DEP202605080177-1 : 쿠폰 90 / 계약금 01 / 100,000원 (COP2026JA0028)\n"
            "    · DEP202605080179-1 : 현금 01 / 잔금 02 / 800,000원\n"
            "- 응답 메시지 : 「판매 등록 완료」  /  storeType: 20 (아울렛)\n"
            "- 응답 적립 포인트 목록 : (없음) — 아울렛은 적립 대상 아님\n"
            "- 검증 : 아울렛 판매 정상 등록 / 적립 포인트 0"
        ),
    ),
    # 4-7-c 아울렛 판매 삭제
    "a12JO000000MBPtYAO": (
        "4-7-c",
        "아울렛 판매 삭제 — SAL202605080164 DELETE",
        (
            "Apex 클래스 : GdSaleApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-08 14:48:49 KST  (4-7-b 판매 후 약 71초)\n"
            "- 대상 판매 번호 : SAL202605080164\n"
            "- URL : {\"SaleNo__c\":\"SAL202605080164\"} (Request Body 없음 — DELETE 패턴)\n"
            "- 응답 처리 : 판매 건 삭제 완료\n"
            "- 검증 : 4-7-b 판매 건 정상 삭제 → 4-7-d 재판매 가능 상태로 복구"
        ),
    ),
    # 4-7-d 아울렛 재판매
    "a12JO000000MEyqYAG": (
        "4-7-d",
        "아울렛 재판매 진행 — SAL202605080167",
        (
            "판매 번호 : SAL202605080167  /  원거래 주문 번호 : SOR202605080147\n"
            "- 시간 : 2026-05-08 14:49:31 KST  (4-7-c 판매 삭제 후 42초)\n"
            "- 매장 코드 : 99997 (아울렛)\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 판매 항목 : 4-7-b 와 동일 (2A2600001 + 3A2600006, 각 500,000)\n"
            "- 거래 원장 :\n"
            "    · DEP202605080172-1 : 현금 01 / 계약금 01 / 100,000원\n"
            "    · DEP202605080177-1 : 쿠폰 90 / 계약금 01 / 100,000원 (COP2026JA0028)\n"
            "    · DEP202605080182-1 : 현금 01 / 잔금 02 / 800,000원  ← 새 입금 번호로 재발행\n"
            "- 응답 메시지 : 「판매 등록 완료」\n"
            "- 검증 : 동일 주문(SOR147)에 대해 새 SAL167 로 재판매 정상 등록 (이전 SAL164 삭제 분과 별개)"
        ),
    ),
    # 4-7-e 아울렛 통반품
    "a12JO000000MDDFYA4": (
        "4-7-e",
        "아울렛 통반품 — SAL202605080170 (전체 환불)",
        (
            "통반품 번호 : SAL202605080170  /  원거래 판매 번호 : SAL202605080167\n"
            "- 시간 : 2026-05-08 14:51:59 KST\n"
            "- 실결제 금액 : 1,000,000원 (전체 환불)\n"
            "- 판매 항목 :\n"
            "    1) 상품 코드 2A2600001 / 수량 -1 / 순매가 단가 500,000\n"
            "    2) 상품 코드 3A2600006 / 수량 -1 / 순매가 단가 500,000\n"
            "- 거래 원장 (입금 번호 DEP202605080186) :\n"
            "    1) 결제 수단 90 (쿠폰) / 입금 구분 03 (반품) / 100,000원 (쿠폰 COP2026JA0028 환불·복원)\n"
            "    2) 결제 수단 01 (현금) / 입금 구분 03 / 900,000원 (현금 환불)\n"
            "- 응답 메시지 : 「반품 등록 완료」  /  success: true\n"
            "- 검증 : 통반품 시 쿠폰 + 현금 모두 환불 처리 / 아울렛 적립 없으므로 적립 소멸 분개 부재 (정상)"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 10, 40, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
