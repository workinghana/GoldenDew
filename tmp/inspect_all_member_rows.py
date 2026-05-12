import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "tmp" / "apex_audit_trail_export_20260508.xlsx"

KEY_FIELDS = [
    "SaleNo__c", "OrderNo__c", "F_OrderNo__c", "F_SaleNo__c",
    "OriginalOrderId", "OriginalOrderId__c", "originalOrderId", "orderId",
    "ActualPaymentAmount__c", "TotalAmount", "MemberNo__c", "F_MemberNo__c",
    "OrderStatus__c", "SaleStatus__c", "DeleteDateTime__c", "IsDeleted__c",
    "IsReductionOrder__c",
    "code", "success", "message",
    "PT_TARGET", "TX_REMARK", "CD_TYPE_SAVE", "CD_TYPE_POINT",
    "DepositNo__c", "DepositType__c", "TransactionAmount", "PaymentMethod",
    "CouponNo", "couponNo",
]


def short(v, n=200):
    s = json.dumps(v, ensure_ascii=False) if not isinstance(v, str) else v
    s = s.replace("\n", " ")
    return s if len(s) <= n else s[:n] + "..."


def extract_keys(text):
    if not text:
        return {}
    out = {}
    try:
        obj = json.loads(text)
    except Exception:
        return {"_raw_excerpt": short(text, 240)}

    def walk(node):
        if isinstance(node, dict):
            for k, v in node.items():
                if k in KEY_FIELDS and k not in out:
                    out[k] = v
                if isinstance(v, (dict, list)):
                    walk(v)
        elif isinstance(node, list):
            for item in node[:3]:
                walk(item)
    walk(obj)
    return out


def main():
    wb = load_workbook(SRC, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({
            "MatchedNumbers": row[idx["MatchedNumbers"]] or "",
            "Id": row[idx["Id"]],
            "CreatedDate": row[idx["CreatedDate"]],
            "ApexClass__c": row[idx["ApexClass__c"]],
            "RequestUrl__c": row[idx["RequestUrl__c"]],
            "RequestBody__c": row[idx["RequestBody__c"]] or "",
            "ResponseBody__c": row[idx["ResponseBody__c"]] or "",
        })
    rows.sort(key=lambda r: (str(r["CreatedDate"] or ""), str(r["Id"] or "")))

    for i, r in enumerate(rows, start=1):
        print("=" * 100)
        print(f"[{i:>2}] {r['CreatedDate']}  Id={r['Id']}")
        print(f"     ApexClass = {r['ApexClass__c']}")
        print(f"     URL       = {short(r['RequestUrl__c'], 80)}")
        print(f"     Matched   = {r['MatchedNumbers']}")
        rk = extract_keys(r['RequestBody__c'])
        sk = extract_keys(r['ResponseBody__c'])
        print(f"     Req keys  : {rk}")
        print(f"     Res keys  : {sk}")


if __name__ == "__main__":
    main()
