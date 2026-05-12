"""V5-2 Log.xlsx — 회원 TEST02009 / [TEST]_프로모션없음.

판매처리 후 입금수정 단계에서 포인트/쿠폰을 삭제하고
해당 삭제금액만큼 현금/카드로 금액 수정 후 저장 테스트.
"""
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V5-2 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v5_2_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V5-2"
MEMBER_NO = "TEST02009"
SHEET_NAME = "V5-2"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower().replace("_", "")
    if "memberhelpinfo" in ac:
        return "회원 도움창"
    if "memberapi" in ac:
        return "회원"
    if "voucherhelp" in ac:
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac:
        return "포인트 (적용 조회)"
    if "saledeposit" in ac:
        return "판매 입금"
    if "returndeposit" in ac:
        return "반품 입금"
    if "orderapi" in ac:
        return "주문"
    if "saleapi" in ac:
        return "판매"
    if "returnapi" in ac:
        return "반품"
    if "pointcredit" in ac:
        return "포인트 (지급)"
    if "pointdebit" in ac:
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    # 테스트 준비 (1)~(2) : 주문 SOR04017 등록
    "a12JO000000MjWSYA0": (
        "준비 (1)~(2)",
        "주문 등록 — SOR04017 (포인트 10,000 + 정률 쿠폰 + 선수금 현금 10%)",
        (
            "주문 번호 : SOR04017  (시간 : 2026-05-11 02:06:21 KST)\n"
            "- 회원 : TEST02009\n"
            "- 매장 코드 : 99998 (백화점)\n"
            "- 실결제 금액 : 400,000원\n"
            "- 프로모션 : (시나리오상 「없음」 — RequestBody 의 PRO 값은 무시)\n"
            "- 결제 구성 : 포인트 10,000원 + 정률 쿠폰 + 선수금 현금 10%\n"
            "- 응답 : 「주문 등록 완료」 — PT_USE 8,480\n"
            "- 시나리오 단계 : 테스트 준비 (1)~(2)"
        ),
    ),
    # 테스트 준비 (3) : 주문 → 판매 처리 (잔금 카드 완불)
    "a12JO000000MkvUYAS": (
        "준비 (3)",
        "판매 처리 — SAL04071 (주문 끌어와 잔금 카드 완불)",
        (
            "판매 번호 : SAL04071  (시간 : 2026-05-11 02:12:26 KST)\n"
            "- 원거래 주문 번호 : SOR04017\n"
            "- 매장 코드 : 99998\n"
            "- 실결제 금액 : 400,000원\n"
            "- 결제 구성 : 주문 분개(현금 35,000 + 쿠폰 + 포인트) 그대로 승계 + 잔금 카드\n"
            "- 응답 : 「판매 등록 완료」 — 구매포인트 7,000P 적립\n"
            "- 시나리오 단계 : 테스트 준비 (3) — 주문 → 판매 처리"
        ),
    ),
    # 진행 (1) - 포인트 입금 삭제
    "a12JO000000Mkx3YAC": (
        "진행 (1)",
        "판매 입금 삭제 — DELETE DEP040024 (포인트 row)",
        (
            "Apex 클래스 : GdSaleDepositApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-11 02:15:39 KST\n"
            "- 대상 입금 번호 : DEP040024  /  DeposiSeqNo : 1\n"
            "- URL : {\"DeposiSeqNo__c\":\"1\",\"DepositNo__c\":\"DEP040024\"} (RequestBody 없음)\n"
            "- 응답 : 「판매 입금 삭제 삭제 완료」\n"
            "- 시나리오 단계 : 진행 (1) — 주문에서 입력한 포인트 삭제"
        ),
    ),
    # 진행 (1) - 쿠폰 입금 삭제
    "a12JO000000MkyfYAC": (
        "진행 (1)",
        "판매 입금 삭제 — DELETE DEP040026 (쿠폰 row)",
        (
            "Apex 클래스 : GdSaleDepositApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-11 02:15:53 KST  (포인트 삭제 후 14초)\n"
            "- 대상 입금 번호 : DEP040026  /  DeposiSeqNo : 3\n"
            "- URL : {\"DeposiSeqNo__c\":\"3\",\"DepositNo__c\":\"DEP040026\"} (RequestBody 없음)\n"
            "- 응답 : 「판매 입금 삭제 삭제 완료」\n"
            "- 시나리오 단계 : 진행 (1) — 주문에서 입력한 쿠폰 삭제\n"
            "- ※ (2), (3) 시도 후 실패하여 (4) 로 이동 — 별도 로그 없음"
        ),
    ),
    # 진행 (4) - 판매 수정 (재판매처리)
    "a12JO000000Ml0IYAS": (
        "진행 (4)",
        "판매 수정 — SAL04071 (입금 전부 삭제 후 재판매처리, 카드 추가 입금)",
        (
            "판매 번호 : SAL04071  (시간 : 2026-05-11 02:18:56 KST)\n"
            "- 매장 코드 : 99998\n"
            "- 실결제 금액 : 400,000원\n"
            "- 결제 구성 : 카드 365,000원 (= 315,000 잔금 + 50,000 쿠폰/포인트 흡수)\n"
            "  + 현금 35,000원 (주문 분개 그대로 유지)\n"
            "- 응답 : 「판매 수정 완료」 — 구매포인트 1,000P\n"
            "- 시나리오 단계 : 진행 (4) — 입금 전부 삭제 후 재판매처리\n"
            "  (주문에서 사용한 포인트/쿠폰을 다시 불러오는 효과)\n"
            "- 메모 : 정책상 주문에서 사용한 포인트/쿠폰은 판매에서 수정 불가\n"
            "  하지만 로직상 판매에서 수정 가능. 결과적으로 저장은 되지만,\n"
            "  주문에서 사용된 쿠폰/포인트는 그대로 사용 된 것으로 남음. (이게 맞음)"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 12, 50, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
