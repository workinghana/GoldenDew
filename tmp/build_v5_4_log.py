"""V5-4 Log.xlsx — 회원 TEST02009 / [TEST] 통합시나리오V5_비활성화 (PRO202605041081).

주문에서 입력 한 프로모션 번호를 비활성화(삭제) 처리 → 판매처리 테스트.
"""
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V5-4 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v5_4_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V5-4"
MEMBER_NO = "TEST02009"
SHEET_NAME = "V5-4"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower().replace("_", "")
    if "memberhelpinfo" in ac:
        return "회원 도움창"
    if "memberapi" in ac:
        return "회원"
    if "voucherhelp" in ac:
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac:
        return "포인트 (적용 조회)"
    if "saledeposit" in ac:
        return "판매 입금"
    if "returndeposit" in ac:
        return "반품 입금"
    if "orderapi" in ac:
        return "주문"
    if "saleapi" in ac:
        return "판매"
    if "returnapi" in ac:
        return "반품"
    if "pointcredit" in ac:
        return "포인트 (지급)"
    if "pointdebit" in ac:
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    # 테스트 준비 (1) : 주문 등록 (비활성화 대상 프로모션)
    "a12JO000000MfvoYAC": (
        "준비 (1)",
        "주문 등록 — SOR04015 (구매포인트 적립 불가 프로모션 사용)",
        (
            "주문 번호 : SOR04015  (시간 : 2026-05-10 20:11:13 KST)\n"
            "- 회원 : TEST02009\n"
            "- 매장 코드 : 99998 (백화점)\n"
            "- 실결제 금액 : 400,000원\n"
            "- 프로모션 : PRO202605041081 ([TEST] 통합시나리오V5_비활성화)\n"
            "  → 구매포인트 적립 불가 프로모션\n"
            "- 응답 : 「주문 등록 완료」 — DebitPointList 빈 배열 (구매포인트 적립 안 됨)\n"
            "- 시나리오 단계 : 테스트 준비 (1) — 주문에 프로모션 입력 (구매포인트 적립 불가)"
        ),
    ),
    # 진행 (1-2) : 프로모션 비활성화 처리 후 판매 (SAL04049)
    "a12JO000000MdpBYAS": (
        "진행 (1-2)",
        "판매 처리 — SAL04049 (프로모션 비활성화 후 / 구매포인트 적립 안됨 ✓ PASS)",
        (
            "판매 번호 : SAL04049  (시간 : 2026-05-10 20:13:33 KST)\n"
            "- 원거래 주문 번호 : SOR04015\n"
            "- 매장 코드 : 99998\n"
            "- 실결제 금액 : 400,000원\n"
            "- 프로모션 : PRO202605041081 (판매 직전 비활성화 처리됨)\n"
            "- 응답 : 「판매 등록 완료」 — 사은포인트 12,000P 만 적립\n"
            "  → 구매포인트는 CreditPointList 에 등장하지 않음 (적립 안 됨)\n"
            "- 시나리오 단계 : 진행 (1-2) — 비활성화 후 판매처리, 구매포인트 적립 안됨 (PASS)\n"
            "- 검증 : 프로모션 비활성화 후에도 판매 자체는 저장 성공, 구매포인트만 차단"
        ),
    ),
    # 사전 정리 : SAL04049 의 입금 삭제 (DEP040005)
    "a12JO000000Mei9YAC": (
        "(2) 사전",
        "판매 입금 삭제 — DELETE DEP040005 (SAL04049 입금 정리)",
        (
            "Apex 클래스 : GdSaleDepositApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-10 20:20:03 KST\n"
            "- 대상 입금 번호 : DEP040005  /  DeposiSeqNo : 1\n"
            "- URL : {\"DeposiSeqNo__c\":\"1\",\"DepositNo__c\":\"DEP040005\"} (RequestBody 없음)\n"
            "- 응답 : 「판매 입금 삭제 삭제 완료」 — PT_USE 6,000 복원\n"
            "- 시나리오 단계 : 진행 (2) 사전 정리 — 다음 시도 위해 SAL04049 입금 삭제"
        ),
    ),
    # 진행 (2-2) : 프로모션 기간 만료 후 판매 (SAL04050)
    "a12JO000000Mg76YAC": (
        "진행 (2-2)",
        "판매 처리 — SAL04050 (프로모션 기간 만료 후 / 구매포인트 적립 안됨 ✓ PASS)",
        (
            "판매 번호 : SAL04050  (시간 : 2026-05-10 20:23:22 KST)\n"
            "- 원거래 주문 번호 : SOR04015\n"
            "- 매장 코드 : 99998\n"
            "- 실결제 금액 : 400,000원\n"
            "- 프로모션 : PRO202605041081 (프로모션 기간을 과거로 수정 → 만료 상태)\n"
            "- 응답 : 「판매 등록 완료」 — 사은포인트 12,000P 만 적립\n"
            "  → 구매포인트는 CreditPointList 에 등장하지 않음 (적립 안 됨)\n"
            "- 시나리오 단계 : 진행 (2-2) — 프로모션 기간 만료 후 판매처리, 구매포인트 적립 안됨 (PASS)\n"
            "- 검증 : 프로모션 만료 상태에서도 판매 저장 성공, 구매포인트만 차단"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 50, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
