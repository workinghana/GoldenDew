"""V5-6 Log.xlsx (정정) — 회원 TEST02007 / 더럽게 테스트 시나리오.

「7) 판매 삭제 잘못해서 포인트 꼬임」
- [TEST]_프로모션없음
- 주문 SOR04010 → 판매 SAL04043 → 부분반품 SAL04114 → 부분반품 삭제
   → 완전반품 SAL04115 → 완전반품 삭제 → 판매 삭제 SAL04043 (포인트 꼬임)
"""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V5-6 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v5_6_correct_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V5-6"
MEMBER_NO = "TEST02007"
SHEET_NAME = "V5-6"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac.replace("_", ""):
        return "포인트 (적용 조회)"
    if "saledeposit" in ac.replace("_", ""):
        return "판매 입금"
    if "returndeposit" in ac.replace("_", ""):
        return "반품 입금"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    # 1) 주문 SOR04010
    "a12JO000000MV5DYAW": (
        "1)",
        "주문 등록 (정액쿠폰 + 포인트 10만 + 품목 4개 + 선입금 10%)",
        (
            "주문 번호 : SOR04010  (시간 : 2026-05-09 21:44:19 KST)\n"
            "- 회원 : TEST02007\n"
            "- 매장 코드 : 99998 (백화점)\n"
            "- 실결제 금액 : 2,000,000원\n"
            "- 프로모션 번호 : (없음 - [TEST]_프로모션없음)\n"
            "- 주문 항목 4개 / 정액쿠폰 + 포인트 10만 사용 + 선입금 10% 입력\n"
            "- 응답 : 「주문 등록 완료」\n"
            "- ※ 12:44:59 에 동일 SOR04010 으로 한 번 더 호출 (수정/재저장 추정)"
        ),
    ),
    # 2) 판매처리 SAL04043
    "a12JO000000MWB2YAO": (
        "2)",
        "판매 처리 — SAL04043",
        (
            "판매 번호 : SAL04043  (시간 : 2026-05-09 21:50:21 KST)\n"
            "- 원거래 주문 번호 : SOR04010\n"
            "- 매장 코드 : 99998\n"
            "- 실결제 금액 : 2,000,000원\n"
            "- 응답 : 「판매 등록 완료」\n"
            "- ※ 직전에 SAL04041 (12:46:44 / 12:47:09), SAL04042 (12:47:23 / 12:50:08) 등 4건의 시도 후 최종 SAL04043 으로 확정"
        ),
    ),
    # 3) 부분 반품 SAL04114
    "a12JO000000MXAKYA4": (
        "3)",
        "부분 반품 — SAL04114",
        (
            "반품 번호 : SAL04114  /  원거래 판매 번호 : SAL04043\n"
            "- 시간 : 2026-05-09 21:54:27 KST\n"
            "- 실결제 금액 : 2,000,000원 (응답 본문 별도 확인 — 부분반품이므로 일부 품목만 -1)\n"
            "- 응답 : 「반품 등록 완료」\n"
            "- 시나리오 단계 : 3) 부분 반품"
        ),
    ),
    # 4) 부분 반품 삭제
    "a12JO000000MTxvYAG": (
        "4)",
        "부분 반품 삭제 — DELETE SAL04114",
        (
            "Apex 클래스 : GdReturnApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-09 21:56:03 KST  (3) 부분반품 후 약 1분 36초)\n"
            "- 대상 반품 번호 : SAL04114\n"
            "- URL : {\"SaleNo__c\":\"SAL04114\"} (RequestBody 없음)\n"
            "- 응답 : 「[판매반품, 교환반품] 전체 삭제 완료」\n"
            "- 시나리오 단계 : 4) 부분 반품 삭제"
        ),
    ),
    # 5) 완전 반품 SAL04115
    "a12JO000000MVy5YAG": (
        "5)",
        "완전 반품 — SAL04115 (잔여 품목 전부 / 13:06:19 본 등록)",
        (
            "반품 번호 : SAL04115  /  원거래 판매 번호 : SAL04043\n"
            "- 시간 : 2026-05-09 22:06:19 KST\n"
            "- 실결제 금액 : 1,683,000원 (잔여 금액 — 정액쿠폰/포인트 차감 후)\n"
            "- 응답 : 「반품 등록 완료」\n"
            "- ※ 22:05:43 에 1차 시도 (a12JO000000MXQMYA4) 후 22:06:19 에 본 등록 완료\n"
            "- 시나리오 단계 : 5) 완전 반품"
        ),
    ),
    # 6) 완전 반품 삭제
    "a12JO000000MWaoYAG": (
        "6)",
        "완전 반품 삭제 — DELETE SAL04115",
        (
            "Apex 클래스 : GdReturnApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-09 22:11:06 KST  (5) 완전반품 후 약 4분 47초)\n"
            "- 대상 반품 번호 : SAL04115\n"
            "- URL : {\"SaleNo__c\":\"SAL04115\"} (RequestBody 없음)\n"
            "- 응답 : 「[판매반품, 교환반품] 전체 삭제 완료」\n"
            "- 시나리오 단계 : 6) 완전 반품 삭제"
        ),
    ),
    # 7) 판매 삭제 — 포인트 꼬임 발생
    "a12JO000000MXdIYAW": (
        "7)",
        "판매 삭제 — DELETE SAL04043 ⚠️ 「포인트 꼬임」 발생",
        (
            "Apex 클래스 : GdSaleApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-09 22:18:57 KST  (6) 완전반품 삭제 후 약 7분 51초)\n"
            "- 대상 판매 번호 : SAL04043\n"
            "- URL : {\"SaleNo__c\":\"SAL04043\"} (RequestBody 없음)\n"
            "- 시나리오 표기 : 「여기서 판매 삭제 잘못해서 포인트 꼬임」 ⚠️\n"
            "- 검증 포인트 :\n"
            "    · 부분반품(SAL04114) 후 삭제 → 포인트 일부 복원\n"
            "    · 완전반품(SAL04115) 후 삭제 → 포인트 일부 복원\n"
            "    · 마지막에 원판매(SAL04043) 자체를 삭제 → 이미 복원/사용된 포인트와 적립 ledger 가 꼬임\n"
            "- 응답 본문 / 후속 LoyaltyLedger·Traceability 별도 확인 필요"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 8, 40, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
