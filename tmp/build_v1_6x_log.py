"""V1 (6-x) Log.xlsx — 정률쿠폰 사용 시 부분반품 불가 시나리오.

6-1 : 기존 회원의 쿠폰 5%할인_VVIP특별바우처3 사용 (TEST03005 정상 판매 흐름)
6-2 : 정률쿠폰 사용 시 부분 반품 불가 확인
       응답 「정률 쿠폰 사용 시, 부분 반품이 불가합니다. 해당 판매 ID : 801JO00000IBg3CYAT」

※ 시나리오 메모상 6-1 회원 TEST03005 + 6-2 회원 TEST02007 (응답 IBg3CYAT 매칭).
   IAugFYAT 케이스는 본 raw csv 시간 범위 (5/10 18:00 ~ 23:00 KST) 외.
"""
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V1 (6-x) Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v1_6x_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V1"
SHEET_NAME = "V1 (6-x)"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower().replace("_", "")
    if "memberhelpinfo" in ac:
        return "회원 도움창"
    if "memberapi" in ac:
        return "회원"
    if "voucherhelp" in ac:
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac:
        return "포인트 (적용 조회)"
    if "saledeposit" in ac:
        return "판매 입금"
    if "returndeposit" in ac:
        return "반품 입금"
    if "orderapi" in ac:
        return "주문"
    if "saleapi" in ac:
        return "판매"
    if "returnapi" in ac:
        return "반품"
    if "pointcredit" in ac:
        return "포인트 (지급)"
    if "pointdebit" in ac:
        return "포인트 (사용/차감)"
    if "erpasync" in ac or "asyncjob" in ac:
        return "ERP 콜아웃 (쿠폰 동기화)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


# (member_no, no, label, criterion)
SCENARIO = {
    # 6-1 사전 : 정률쿠폰 ERP 동기화 콜아웃
    "a12JO000000MeIAYA0": (
        "TEST03005",
        "6-1 사전",
        "ERP 쿠폰 동기화 콜아웃 — 정률쿠폰 COP2026MY0003 (mil 00218250)",
        (
            "Apex 클래스 : ErpAsyncJob (CALLOUT)\n"
            "- 시간 : 2026-05-10 18:45:44 KST\n"
            "- 회원 : TEST03005\n"
            "- URL : callout:GoldenDewApiNC/Crm/CouponDetail\n"
            "- 요청 : no_man_coupon_mil=00218250, no_man_coupon=COP2026MY0003\n"
            "- 응답 : ❌ \"SQL - 회원이 존재하지 않습니다.\"\n"
            "- 시나리오 단계 : 6-1 정률 쿠폰 (VVIP특별바우처3 / COP2026MY0003) 사용 시\n"
            "  ERP 측 동기화 콜아웃 (마이그 환경 회원 미존재로 실패)"
        ),
    ),
    # 6-1 주문 : SOR04013
    "a12JO000000MdlxYAC": (
        "TEST03005",
        "6-1 주문",
        "주문 등록 — SOR04013 (정률쿠폰 + 포인트 10,000 사용)",
        (
            "주문 번호 : SOR04013  (시간 : 2026-05-10 18:47:58 KST)\n"
            "- 회원 : TEST03005  /  매장 코드 : 99998\n"
            "- 실결제 금액 : 500,000원\n"
            "- 프로모션 : PRO202604161041\n"
            "- 결제 구성 : 포인트 10,000 사용 + 정률쿠폰 적용\n"
            "- 응답 : 「주문 등록 완료」 — PT_USE 10,000\n"
            "- 시나리오 단계 : 6-1 정률쿠폰 사용한 주문 등록"
        ),
    ),
    # 6-1 판매 : SAL04047
    "a12JO000000MeLUYA0": (
        "TEST03005",
        "6-1 판매",
        "판매 등록 — SAL04047 (정상 처리, PT_TARGET 4,400P)",
        (
            "판매 번호 : SAL04047  (시간 : 2026-05-10 18:51:13 KST)\n"
            "- 원거래 주문 번호 : SOR04013\n"
            "- 실결제 금액 : 500,000원\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 4,400P\n"
            "- 시나리오 단계 : 6-1 정률쿠폰 사용 판매 정상 처리"
        ),
    ),
    # 6-1 판매 수정 1차
    "a12JO000000MereYAC": (
        "TEST03005",
        "6-1 판매수정",
        "판매 수정 1차 — SAL04047 (PT_TARGET 160P)",
        (
            "판매 번호 : SAL04047  (시간 : 2026-05-10 18:56:27 KST)\n"
            "- 응답 : 「판매 수정 완료」 — PT_TARGET 160\n"
            "- 시나리오 단계 : 6-1 판매 수정 (1차)"
        ),
    ),
    # 6-1 판매 수정 2차
    "a12JO000000MewUYAS": (
        "TEST03005",
        "6-1 판매수정",
        "판매 수정 2차 — SAL04047 (PT_TARGET 440P)",
        (
            "판매 번호 : SAL04047  (시간 : 2026-05-10 19:02:52 KST)\n"
            "- 응답 : 「판매 수정 완료」 — PT_TARGET 440\n"
            "- 시나리오 단계 : 6-1 판매 수정 (2차)"
        ),
    ),
    # 6-2 사전 : SAL02V017 정률쿠폰 판매 (TEST02007)
    "a12JO000000Mf9QYAS": (
        "TEST02007",
        "6-2 사전",
        "판매 등록 — SAL02V017 (정률쿠폰 사용, PT_TARGET 19,000P / 원판매 IBg3CYAT)",
        (
            "판매 번호 : SAL02V017  (시간 : 2026-05-10 19:28:05 KST)\n"
            "- 회원 : TEST02007  /  매장 코드 : 99998\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 프로모션 : PRO202604161041\n"
            "- 결제 구성 : [TEST]_정률_10% 쿠폰 사용\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 19,000P\n"
            "- 시나리오 단계 : 6-2 사전 — 정률쿠폰 사용 판매 (이후 부분반품 시도 대상)\n"
            "- ※ 이 판매의 OrderId 가 801JO00000IBg3CYAT — 다음 부분반품 시도 응답에 노출"
        ),
    ),
    # 6-1 입금 삭제
    "a12JO000000MfFpYAK": (
        "TEST03005",
        "6-1 입금삭제",
        "판매 입금 삭제 — DELETE DEP040001 (SAL04047 정리)",
        (
            "Apex 클래스 : GdSaleDepositApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-10 19:33:14 KST\n"
            "- 대상 입금 번호 : DEP040001  /  DeposiSeqNo : 4\n"
            "- 응답 : 「판매 입금 삭제 삭제 완료」 — PT_USE 4,400 복원\n"
            "- 시나리오 단계 : 6-1 SAL04047 입금 정리"
        ),
    ),
    # 6-2 ★ 핵심 : 정률쿠폰 부분반품 불가 응답
    "a12JO000000MfKgYAK": (
        "TEST02007",
        "6-2 ★ 핵심",
        "부분반품 시도 — SAL02V018 ❌ 「정률 쿠폰 사용 시 부분 반품 불가」 ✓ PASS",
        (
            "반품 번호 : SAL02V018  /  원거래 판매 번호 : SAL02V017\n"
            "- 시간 : 2026-05-10 19:34:11 KST\n"
            "- 실결제 금액 : 500,000원 (한 품목 부분반품 시도)\n"
            "- 응답 : ❌ code 400\n"
            "    「정률 쿠폰 사용 시, 부분 반품이 불가합니다.\n"
            "     해당 판매 ID : 801JO00000IBg3CYAT」\n"
            "- 시나리오 단계 : 6-2 정률쿠폰 부분반품 불가 메시지 노출 ✓ PASS\n"
            "- 검증 : 정률쿠폰 사용 판매 (SAL02V017 / OrderId IBg3CYAT) 부분반품 차단 정상\n"
            "- ※ IAugFYAT 응답 케이스 (시나리오 메모 표기) 는 본 시간 범위 외 — 미수집"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (member_no, no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            member_no,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 50, 95, 16, 24, 42, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
