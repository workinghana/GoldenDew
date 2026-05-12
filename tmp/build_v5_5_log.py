"""V5-5 Log.xlsx — 회원 TEST03007 / [TEST] V5_6사은적립불가 (PRO202605051083).

이전 판매에서 생성된 구매적립 포인트를 끌어다 판매처리 한 뒤,
이전 판매처리를 취소하는 경우 (B 판매 취소 불가 확인).
"""
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V5-5 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v5_5_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V5-5"
MEMBER_NO = "TEST03007"
SHEET_NAME = "V5-5"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower().replace("_", "")
    if "memberhelpinfo" in ac:
        return "회원 도움창"
    if "memberapi" in ac:
        return "회원"
    if "voucherhelp" in ac:
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac:
        return "포인트 (적용 조회)"
    if "saledeposit" in ac:
        return "판매 입금"
    if "returndeposit" in ac:
        return "반품 입금"
    if "orderapi" in ac:
        return "주문"
    if "saleapi" in ac:
        return "판매"
    if "returnapi" in ac:
        return "반품"
    if "pointcredit" in ac:
        return "포인트 (지급)"
    if "pointdebit" in ac:
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    # 준비 (1) A 판매 — SAL04058
    "a12JO000000MhfsYAC": (
        "준비 (1) A",
        "A 판매 등록 — SAL04058 (적립포인트 10,000원 생성)",
        (
            "판매 번호 : SAL04058  (시간 : 2026-05-10 23:02:23 KST)\n"
            "- 회원 : TEST03007  (시나리오 메모 회원 TEST01009 — 마이그 전 표기)\n"
            "- 매장 코드 : 99998 (백화점)\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 프로모션 : PRO202605051083 ([TEST] V5_6사은적립불가)\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 10,000 (A판매 적립포인트)\n"
            "- 시나리오 단계 : 준비 (1) A 판매 — 적립포인트 생성"
        ),
    ),
    # 준비 (1) B 판매 — SAL04059
    "a12JO000000MhkiYAC": (
        "준비 (1) B",
        "B 판매 등록 — SAL04059 (적립포인트 20,000원 생성)",
        (
            "판매 번호 : SAL04059  (시간 : 2026-05-10 23:03:54 KST)\n"
            "- 회원 : TEST03007\n"
            "- 매장 코드 : 99998\n"
            "- 실결제 금액 : 2,000,000원\n"
            "- 프로모션 : PRO202605051083\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 20,000 (B판매 적립포인트)\n"
            "- 시나리오 단계 : 준비 (1) B 판매 — 적립포인트 생성 (※ 이후 취소 시도 대상)"
        ),
    ),
    # 준비 (1) C 판매 — SAL04060
    "a12JO000000MhnwYAC": (
        "준비 (1) C",
        "C 판매 등록 — SAL04060 (적립포인트 10,000원 생성)",
        (
            "판매 번호 : SAL04060  (시간 : 2026-05-10 23:05:07 KST)\n"
            "- 회원 : TEST03007\n"
            "- 매장 코드 : 99998\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 프로모션 : PRO202605051083\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 10,000 (C판매 적립포인트)\n"
            "- 시나리오 단계 : 준비 (1) C 판매 — 적립포인트 생성"
        ),
    ),
    # 부수 : SAL04061 잘못 등록
    "a12JO000000MhuOYAS": (
        "(부수) SAL04061",
        "SAL04061 잘못 등록 (시나리오에 없음, 이후 취소 정리)",
        (
            "판매 번호 : SAL04061  (시간 : 2026-05-10 23:07:49 KST)\n"
            "- 회원 : TEST03007\n"
            "- 실결제 금액 : 2,020,000원\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 20,000\n"
            "- 시나리오 단계 : 부수 — 시나리오에 없는 등록 (직후 입금 삭제 + 판매 전체 삭제로 정리)"
        ),
    ),
    # 부수 : SAL04061 입금 삭제
    "a12JO000000MhxZYAS": (
        "(부수) 삭제",
        "판매 입금 삭제 — DELETE DEP040031 (SAL04061 정리)",
        (
            "Apex 클래스 : GdSaleDepositApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-10 23:08:17 KST\n"
            "- 대상 입금 번호 : DEP040031  /  DeposiSeqNo : 1\n"
            "- 응답 : 「판매 입금 삭제 삭제 완료」 — SAL04061 의 PT_USE 20,000 복원\n"
            "- 시나리오 단계 : 부수 — SAL04061 (잘못 등록건) 의 입금 삭제"
        ),
    ),
    # 부수 : SAL04061 판매 전체 삭제
    "a12JO000000Mi0qYAC": (
        "(부수) 삭제",
        "판매 전체 삭제 — DELETE SAL04061 (잘못 등록건 정리 완료)",
        (
            "Apex 클래스 : GdSaleApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-10 23:08:39 KST\n"
            "- 대상 판매 번호 : SAL04061\n"
            "- URL : {\"SaleNo__c\":\"SAL04061\"} (RequestBody 없음)\n"
            "- 응답 : 「판매 전체 삭제 완료」 (code 200)\n"
            "- 시나리오 단계 : 부수 — SAL04061 정리 완료, 다음 D 판매로 이동"
        ),
    ),
    # 준비 (2) D 판매 — SAL04062 (A+B 일부 끌어와 사용)
    "a12JO000000Mi2SYAS": (
        "준비 (2) D",
        "D 판매 등록 — SAL04062 (A+B(일부) 포인트 끌어와 사용, 20,000P 사용)",
        (
            "판매 번호 : SAL04062  (시간 : 2026-05-10 23:09:05 KST)\n"
            "- 회원 : TEST03007\n"
            "- 매장 코드 : 99998\n"
            "- 실결제 금액 : 2,020,000원\n"
            "- 사용 포인트 : 20,000 (기타 포인트 5,000 + B판매 적립 15,000)\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 20,000 (D판매 적립포인트)\n"
            "- 시나리오 단계 : 준비 (2) — A+B(일부) 포인트 끌어와 D 판매처리\n"
            "- ※ 이 SAL04062 가 B판매(SAL04059) 적립 포인트를 사용한 대상 — 이후 B 취소 시도시 에러"
        ),
    ),
    # 준비 (3) E 판매 — SAL04063 (B 나머지 + C 끌어와 사용)
    "a12JO000000Mi8yYAC": (
        "준비 (3) E",
        "E 판매 등록 — SAL04063 (B(나머지) + C 포인트 끌어와 사용, 15,000P 사용)",
        (
            "판매 번호 : SAL04063  (시간 : 2026-05-10 23:11:50 KST)\n"
            "- 회원 : TEST03007\n"
            "- 매장 코드 : 99998\n"
            "- 실결제 금액 : 2,015,000원\n"
            "- 사용 포인트 : 15,000 (C판매 1만 + B판매 1,666 + A판매 3,334)\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 20,000 (E판매 적립포인트)\n"
            "- 시나리오 단계 : 준비 (3) — B(나머지)+C 포인트 끌어와 E 판매처리\n"
            "- ※ D, E 판매로 생성된 적립포인트는 남아있는 상태 (준비 (4))"
        ),
    ),
    # 진행 (1) B 판매 취소 시도 — ERROR
    "a12JO000000MiFJYA0": (
        "진행 (1) ⚠️",
        "B 판매 입금 삭제 시도 — DELETE DEP040029 ❌ ERROR (B 판매 취소 불가)",
        (
            "Apex 클래스 : GdSaleDepositApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-10 23:19:25 KST\n"
            "- 대상 입금 번호 : DEP040029  /  DeposiSeqNo : 1\n"
            "  (= B 판매 SAL04059 의 입금 row)\n"
            "- URL : {\"DeposiSeqNo__c\":\"1\",\"DepositNo__c\":\"DEP040029\"} (RequestBody 없음)\n"
            "- 응답 : ❌ code 400\n"
            "    「해당 입금 건은 삭제할 수 없습니다. 적립된 포인트가 다른 주문/판매에서\n"
            "     사용되었습니다. 판매번호: SAL04062 / 입금번호: DEP040034-2」\n"
            "- 시나리오 단계 : 진행 (1) — B 판매 취소 시도 → 취소 불가 ✓ PASS\n"
            "- 검증 : D 판매 (SAL04062) 가 B 판매 적립 포인트 사용 중 → B 취소 차단 정상"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 50, 95, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
