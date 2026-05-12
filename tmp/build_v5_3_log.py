"""V5-3 Log.xlsx — 회원 TEST02009 / [TEST] 통합시나리오V5_사은2배 (PRO202605041080).

주문에서 입력 한 프로모션 번호의 품목코드를 수정 → 타 매장 인도건 판매처리 → 판매 삭제 테스트.
"""
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V5-3 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v5_3_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V5-3"
MEMBER_NO = "TEST02009"
SHEET_NAME = "V5-3"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower().replace("_", "")
    if "memberhelpinfo" in ac:
        return "회원 도움창"
    if "memberapi" in ac:
        return "회원"
    if "voucherhelp" in ac:
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac:
        return "포인트 (적용 조회)"
    if "saledeposit" in ac:
        return "판매 입금"
    if "returndeposit" in ac:
        return "반품 입금"
    if "orderapi" in ac:
        return "주문"
    if "saleapi" in ac:
        return "판매"
    if "returnapi" in ac:
        return "반품"
    if "pointcredit" in ac:
        return "포인트 (지급)"
    if "pointdebit" in ac:
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    # 테스트 준비 (1)~(3) : 주문 등록
    "a12JO000000Mfe4YAC": (
        "준비 (1)~(3)",
        "주문 등록 — SOR04014 (선수금 100% / 주문매장 99998 / 사은2배 프로모션)",
        (
            "주문 번호 : SOR04014  (시간 : 2026-05-10 19:58:46 KST)\n"
            "- 회원 : TEST02009\n"
            "- 주문 매장 코드 : 99998 (백화점)\n"
            "- 실결제 금액 : 900,000원\n"
            "- 프로모션 : PRO202605041080 ([TEST] 통합시나리오V5_사은2배)\n"
            "- 결제 구성 : 선수금 100% (현금 50% + 카드 50%)\n"
            "- 주문 셋팅 : 기획제품 2개 + 사은포인트 2배\n"
            "- 응답 : 「주문 등록 완료」\n"
            "- 시나리오 단계 : 테스트 준비 (1)~(3)"
        ),
    ),
    # 진행 (2) : 품목 변경 후 타 매장 인도건 판매처리
    "a12JO000000MfiuYAC": (
        "진행 (2)",
        "판매 처리 — SAL04048 (품목 변경 + 인도매장 99995 / 사은2배 정상 적용)",
        (
            "판매 번호 : SAL04048  (시간 : 2026-05-10 20:02:31 KST)\n"
            "- 원거래 주문 번호 : SOR04014\n"
            "- 인도매장 코드 : 99995 (타 매장)\n"
            "- 실결제 금액 : 900,000원\n"
            "- 프로모션 : PRO202605041080 (사은2배)\n"
            "- 처리 내용 : 주문서의 기획제품 2개를 다른 품목코드로 변경 후 판매처리\n"
            "  (본사 출고 시점에 품목 변경 — 테스트 환경에서는 강제 UPDATE 로 진행)\n"
            "- 응답 : 「판매 등록 완료」 — 구매포인트 PT_TARGET 18,000P (사은2배 정상 적용)\n"
            "- 시나리오 단계 : 진행 (2-1) — 인도매장이 달라도 판매 매장은 주문매장으로 SAL 생성\n"
            "- 검증 : 변경된 품목으로 로열티에 잘 바인딩, 사은포인트 2배수 정상 동작"
        ),
    ),
    # 판매 삭제 사전 : 입금 삭제 (현금 50%)
    "a12JO000000MfkTYAS": (
        "진행 (3) 사전",
        "판매 입금 삭제 — DELETE DEP040002 (현금 50%)",
        (
            "Apex 클래스 : GdSaleDepositApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-10 20:05:19 KST\n"
            "- 대상 입금 번호 : DEP040002  /  DeposiSeqNo : 1\n"
            "- URL : {\"DeposiSeqNo__c\":\"1\",\"DepositNo__c\":\"DEP040002\"} (RequestBody 없음)\n"
            "- 응답 : 「판매 입금 삭제 삭제 완료」 — PT_USE 9,000 복원\n"
            "- 시나리오 단계 : 진행 (3) 사전 정리 — 판매 삭제 전 입금 row 삭제 (현금)"
        ),
    ),
    # 판매 삭제 사전 : 입금 삭제 (카드 50%)
    "a12JO000000Mfm5YAC": (
        "진행 (3) 사전",
        "판매 입금 삭제 — DELETE DEP040003 (카드 50%)",
        (
            "Apex 클래스 : GdSaleDepositApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-10 20:05:26 KST  (현금 삭제 후 7초)\n"
            "- 대상 입금 번호 : DEP040003  /  DeposiSeqNo : 2\n"
            "- URL : {\"DeposiSeqNo__c\":\"2\",\"DepositNo__c\":\"DEP040003\"} (RequestBody 없음)\n"
            "- 응답 : 「판매 입금 삭제 삭제 완료」 — PT_USE 9,000 복원\n"
            "- 시나리오 단계 : 진행 (3) 사전 정리 — 판매 삭제 전 입금 row 삭제 (카드)"
        ),
    ),
    # 진행 (3) : 판매 삭제
    "a12JO000000MfpMYAS": (
        "진행 (3)",
        "판매 삭제 — DELETE SAL04048 (사은포인트 2배수 정상 차감 확인)",
        (
            "Apex 클래스 : GdSaleApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-10 20:05:45 KST  (입금 삭제 완료 후 19초)\n"
            "- 대상 판매 번호 : SAL04048\n"
            "- URL : {\"SaleNo__c\":\"SAL04048\"} (RequestBody 없음)\n"
            "- 응답 : 「판매 전체 삭제 완료」 (code 200)\n"
            "- 시나리오 단계 : 진행 (3) — 판매 삭제 시 사은포인트 2배수 정상 적용\n"
            "- 검증 : 인도매장이 달라도 사은포인트 2배수 정상 차감 (적립포인트 환불)"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 50, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
