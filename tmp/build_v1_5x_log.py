"""V1 (5-x) Log.xlsx — 회원 TEST02006.

5. 주문 시 쿠폰/포인트 적용하지 않고 주문 후 판매 처리 시 쿠폰/포인트 적용하여 판매 처리
   - 판매 처리 이후 부분 반품 진행

※ 시나리오 페이지 회원번호 212501600243 은 마이그 전 표기. 실제 로그 회원: TEST02006.
"""
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V1 (5-x) Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v1_5x_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V1"
MEMBER_NO = "TEST02006"
SHEET_NAME = "V1 (5-x)"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower().replace("_", "")
    if "memberhelpinfo" in ac:
        return "회원 도움창"
    if "memberapi" in ac:
        return "회원"
    if "voucherhelp" in ac:
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac:
        return "포인트 (적용 조회)"
    if "saledeposit" in ac:
        return "판매 입금"
    if "returndeposit" in ac:
        return "반품 입금"
    if "orderapi" in ac:
        return "주문"
    if "saleapi" in ac:
        return "판매"
    if "returnapi" in ac:
        return "반품"
    if "pointcredit" in ac:
        return "포인트 (지급)"
    if "pointdebit" in ac:
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    # 5-1 시도 — SOR02V013 (DUPLICATE_VALUE 에러)
    "a12JO000000MdxDYAS": (
        "5-1 시도",
        "주문 등록 1차 시도 — SOR02V013 ❌ ERROR DUPLICATE_VALUE",
        (
            "주문 번호 : SOR02V013  (시간 : 2026-05-10 18:44:56 KST)\n"
            "- 회원 : TEST02006\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 응답 : ❌ code 500 — DUPLICATE_VALUE (DepositNo__c 중복)\n"
            "- 시나리오 단계 : 5-1 1차 시도 (DepositNo 중복 에러)"
        ),
    ),
    # 5-1 — SOR02V014 주문 등록 (쿠폰/포인트 적용 X)
    "a12JO000000Mei0YAC": (
        "5-1",
        "주문 등록 — SOR02V014 (쿠폰/포인트 적용 X)",
        (
            "주문 번호 : SOR02V014  (시간 : 2026-05-10 18:45:08 KST)\n"
            "- 회원 : TEST02006  /  매장 코드 : 99998\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 프로모션 : PRO202604161041\n"
            "- 응답 : 「주문 등록 완료」 — DebitPointList: 빈 배열\n"
            "  (= 쿠폰/포인트 적용 안 됨 확인)\n"
            "- 시나리오 단계 : 5-1 주문 시 쿠폰/포인트 적용하지 않고 주문"
        ),
    ),
    # 5-2 — SAL02V015 판매 처리 (쿠폰/포인트 적용)
    "a12JO000000MelEYAS": (
        "5-2",
        "판매 처리 — SAL02V015 (판매 시 쿠폰/포인트 적용, PT_TARGET 18,000P)",
        (
            "판매 번호 : SAL02V015  (시간 : 2026-05-10 18:47:52 KST)\n"
            "- 원거래 주문 번호 : SOR02V014\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 프로모션 : PRO202604161041\n"
            "- 결제 구성 : 판매 시점에 쿠폰/포인트 적용하여 처리\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 18,000P\n"
            "- 시나리오 단계 : 5-2 판매시 쿠폰 / 포인트 적용하여 판매 처리\n"
            "  (회원 도움창 적립 내역상 SAL02V015 가 이전 SAL02V003/V004/V008 의\n"
            "   적립 포인트를 사용한 것으로 표시됨)"
        ),
    ),
    # 5-3 시도 — SAL02V016 반품 1차 (DUPLICATE_VALUE 에러)
    "a12JO000000MeDMYA0": (
        "5-3 시도",
        "부분반품 1차 시도 — SAL02V016 ❌ ERROR DUPLICATE_VALUE",
        (
            "반품 번호 : SAL02V016  /  원거래 판매 번호 : SAL02V015\n"
            "- 시간 : 2026-05-10 19:00:37 KST\n"
            "- 실결제 금액 : 500,000원 (부분반품 1개)\n"
            "- 응답 : ❌ code 500 — DUPLICATE_VALUE (DepositNo__c 중복)\n"
            "- 시나리오 단계 : 5-3 1차 시도 (DepositNo 중복 에러)"
        ),
    ),
    # 5-3 — SAL02V016 부분반품 성공
    "a12JO000000MeuuYAC": (
        "5-3 ✓",
        "부분반품 — SAL02V016 (생일쿠폰 원복, 포인트 복원 / PT_USE 8,400)",
        (
            "반품 번호 : SAL02V016  /  원거래 판매 번호 : SAL02V015\n"
            "- 시간 : 2026-05-10 19:00:48 KST  (1차 시도 후 11초)\n"
            "- 실결제 금액 : 500,000원 (부분반품 1개, -1 수량)\n"
            "- 응답 : 「반품 등록 완료」 — PT_USE 8,400 복원\n"
            "- 시나리오 단계 : 5-3 부분반품 ✓ PASS\n"
            "- 검증 결과 (시나리오 메모) :\n"
            "  · 생일 쿠폰 원복 ✓\n"
            "  · 포인트 20,000 복원 (실제 응답 PT_USE 8,400)\n"
            "- 메모 : 처음 시나리오 의도는 「주문→판매→부분반품 흐름에서 쿠폰/포인트\n"
            "  처리가 잘 되는지」 확인 (단순 시나리오는 주문 판매 부분반품 로직에서 쿠폰\n"
            "  포인트 처리가 잘 되는지 확인이었으나 회상이 화상이라고 하심...)"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 50, 95, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
