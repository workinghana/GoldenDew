import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "tmp" / "apex_audit_trail_export_20260508.xlsx"
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "scenario_audit_trail_20260508.xlsx"

EXCEL_CELL_LIMIT = 32767

SCENARIO_VERSION = "v4-1_7차"
MEMBER_NO = "269999800209"

# 도메인 분류 (ApexClass / URL 패턴 기반)
def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    u = (url or "").lower()
    if "member" in ac and "help" in ac:
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "erpasync" in ac.replace("_", ""):
        return "ERP 비동기 (쿠폰)"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (차감)"
    return ""


def http_method(apex_class, url):
    # 본 시나리오의 모든 REST 엔드포인트는 @HttpPost 패턴 — 외부 callout만 별도
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    return "POST"


# 시나리오 매핑: Salesforce 로그 Id -> (번호, 확인내용, 확인 기준)
SCENARIO = {
    # ============================================================
    # 1. 아울렛에서 신규 회원 가입 + 자동 쿠폰 발행
    # ============================================================
    "a12JO000000M5XJYA0": (
        "1-1",
        "회원가입 (SMS 수신 허용)",
        (
            "신규 회원 등록 — 아울렛 매장에서 가입\n"
            "- MemberNo__c : 269999800209\n"
            "- 응답 message : 회원 등록 성공.\n"
            "- code : 200"
        ),
    ),
    "a12JO000000M5YvYAK": (
        "1-2",
        "신규 회원 혜택쿠폰 자동 발행 (ERP 비동기)",
        (
            "ErpAsyncJob → callout:GoldenDewApiNC/Crm/CouponDetail\n"
            "- 신규 가입 혜택 쿠폰 2건 자동 발행 (백화점 3% + 아울렛 1만원)\n"
            "- 발행된 쿠폰 (TJ 기준):\n"
            "    · COP2026MY0014 (아울렛 1만원)\n"
            "    · COP2026MY0017 (백화점 3%)\n"
            "- TJ Welcome Coupon Issue 확인 (TransactionJournal 0lVJO000000Bv9R2AS, 00:31:18)"
        ),
    ),
    # ============================================================
    # 2. 아울렛 주문 (1만원 쿠폰 + 10% 입금)
    # ============================================================
    "a12JO000000M3FOYA0": (
        "2-1",
        "아울렛 주문 시 사용 가능 쿠폰 조회",
        (
            "GdVoucherHelpApiControllerV1 — /gd/v1/voucher/help\n"
            "- MemberNo__c : 269999800209\n"
            "- ActualPaymentAmount__c : 500,000\n"
            "- 응답 : 사용 가능한 쿠폰 조회 성공"
        ),
    ),
    "a12JO000000M5h2YAC": (
        "2-2",
        "아울렛 주문 등록 (A 주문)",
        (
            "A 주문 : SOR202605080040  (Salesforce Id : 801JO00000I9ugaYAB)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 입금 금액 : 50,000원  (10% / 현금-PaymentMethod 01 · DepositType 01 · DEP202605080002-1)\n"
            "- 쿠폰 사용 : 10,000원  (COP2026MY0014 / PaymentMethod 90 · DEP202605080002-2)  ※ TJ에서 확인 (audit trail 본 요청에는 CouponNo 빈값)\n"
            "  → 추후 C 판매(SAL202605080056)에 매칭 (SaleId 갱신)\n"
            "- A 주문 합계 입금 : 60,000원 (현금 50,000 + 쿠폰 10,000)\n"
            "- OrderStatus : 01 (정상 주문)"
        ),
    ),
    # ============================================================
    # 3. 백화점 주문번호1 (3% 쿠폰 + 10% 입금)
    # ============================================================
    "a12JO000000M4mYYAS": (
        "3-1",
        "백화점 주문 시 사용 가능 쿠폰 조회 (1차)",
        (
            "GdVoucherHelpApiControllerV1 — /gd/v1/voucher/help\n"
            "- MemberNo__c : 269999800209\n"
            "- ActualPaymentAmount__c : 1,000,000\n"
            "- 응답 : 사용 가능한 쿠폰 조회 성공  (백화점 3% 쿠폰 조회 확인)"
        ),
    ),
    "a12JO000000M60LYAS": (
        "3-2",
        "백화점 주문 시 사용 가능 쿠폰 조회 (2차/재조회)",
        (
            "동일 회원/금액 재조회 — 응답 동일 (1초 차)"
        ),
    ),
    "a12JO000000M620YAC": (
        "3-3",
        "백화점 주문번호1 등록 (3% 쿠폰 + 10% 입금)",
        (
            "백화점 주문번호1 : SOR202605080041  (Salesforce Id : 801JO00000IA0UFYA1)\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 입금 금액 : 100,000원  (10% / 현금-PaymentMethod 01 · DepositType 01 · DEP202605080003-1)\n"
            "- 쿠폰 사용 : 30,000원  (COP2026MY0017 — 백화점 3% / PaymentMethod 90 · DEP202605080003-2)  ※ TJ 기준\n"
            "- 2 items / OrderStatus : 01"
        ),
    ),
    # ============================================================
    # 5. 백화점 주문 수정 (입금 환불 시도 + 품목 삭제 불가 + 3% 쿠폰 환불만 저장)
    # ============================================================
    "a12JO000000M65CYAS": (
        "5-1",
        "백화점 주문번호1 수정 (3% 쿠폰 환불만 저장)",
        (
            "GdOrderApiControllerV1 — /gd/v1/order  /  응답 message : 「주문 수정 완료」\n"
            "수정 대상 : SOR202605080041  (Salesforce Id : 801JO00000IA0UFYA1)\n"
            "시나리오 동작:\n"
            "  1) 입금탭에서 모든 입금 환불 등록 (저장 없이)\n"
            "  2) 품목탭에서 1개 제품 삭제 시도 → 삭제 불가 확인\n"
            "  3) 입금탭으로 복귀 → 3% 쿠폰 환불 등록 분만 남기고 나머지 환불 모두 삭제\n"
            "  4) 원주문 상태에서 3% 쿠폰만 환불 등록한 채 최종 저장\n"
            "TJ 기록:\n"
            "  - 쿠폰 사용 취소 30,000원  (COP2026MY0017 / Coupon Redeem Cancel / DepositType 03 · DEP202605080004-1, 00:37:48)\n"
            "  - 입금/품목은 변동 없음 (주문 100,000원 입금 유지)"
        ),
    ),
    "a12JO000000M52kYAC": (
        "5-2",
        "백화점 주문번호2 등록 (정상가 3개 / 부분 판매 진행용)",
        (
            "백화점 주문번호2 : SOR202605080042  (Salesforce Id : 801JO00000I9vo3YAB)\n"
            "- 실결제 금액 : 1,400,000원  (정상가 판매 제품 3개)\n"
            "- 입금 금액 : 140,000원  (10% / 현금-PaymentMethod 01 · DepositType 01 · DEP202605080005-1)\n"
            "- 쿠폰 사용 : 없음  (CouponNo 빈값 — 부분 판매 단계에서 3% 쿠폰 재적용 예정)\n"
            "- 3 items / OrderStatus : 01\n"
            "  ※ 추후 B 판매(SAL202605080053) 등록 시 SaleId 갱신"
        ),
    ),
    # ============================================================
    # 6. 백화점 주문번호2 → 부분 판매 2개 (3% 쿠폰 재적용)
    # ============================================================
    "a12JO000000M6LJYA0": (
        "6-1",
        "부분 판매 전 사용 가능 쿠폰 조회 (1,000,000 기준)",
        (
            "GdVoucherHelpApiControllerV1 — /gd/v1/voucher/help\n"
            "- ActualPaymentAmount__c : 1,000,000  (조회 시점 임시 금액)"
        ),
    ),
    "a12JO000000M6Q9YAK": (
        "6-2",
        "A 판매(부분 판매) 직전 사용 가능 쿠폰 조회 (900,000)",
        (
            "GdVoucherHelpApiControllerV1 — /gd/v1/voucher/help\n"
            "- ActualPaymentAmount__c : 900,000  (실 부분 판매 금액 기준)"
        ),
    ),
    "a12JO000000M6RlYAK": (
        "6-3",
        "A 판매(부분 판매) 직전 쿠폰 조회 재조회",
        "동일 회원/금액 재조회 — 응답 동일 (10초 차)",
    ),
    "a12JO000000M52oYAC": (
        "6-4",
        "A 판매 등록 — 부분 판매 2개 + 3% 쿠폰 재적용",
        (
            "A 판매 : SAL202605080046  /  원주문 : SOR202605080042 (백화점 주문번호2)\n"
            "(Salesforce Id : 801JO00000I9vo5YAB)\n"
            "- 실결제 금액 : 900,000원  (3개 중 2개 부분 판매)\n"
            "- 판매 입금 : 873,000원  (현금-PaymentMethod 01 · DepositType 02 · DEP202605080006-1)\n"
            "- 쿠폰 사용 : 27,000원  (COP2026MY0017 — 3% 재적용 / PaymentMethod 90 · DEP202605080006-2)\n"
            "- 적립포인트(쇼핑포인트) : 34,920P  (TJ Purchase Accrual / CD_TYPE_POINT 01)\n"
            "  ※ API 응답 PT_TARGET(1% 기준) : 8,730P / TX_REMARK : 구매포인트\n"
            "- A 판매 합계 : 873,000(현금) + 27,000(쿠폰) = 900,000원 ✓"
        ),
    ),
    "a12JO000000M6UzYAK": (
        "6-5",
        "A 판매 후 회원 도움창 조회",
        (
            "GdMemberHelpInfoApiControllerV1\n"
            "- MemberNo__c : 269999800209\n"
            "- 적립포인트(PT_TARGET) : 8,730P / TotalAmount 500,000 / TransactionAmount 10,000\n"
            "- 응답 message : 회원도움정보 조회 성공"
        ),
    ),
    "a12JO000000M6bRYAS": (
        "6-6",
        "A 판매 후 회원 도움창 재조회",
        "동일 회원 재조회 — 응답 동일 (적립포인트 8,730P 변동 없음)",
    ),
    # ============================================================
    # 7. 나머지 1개 부분 판매 후 반품
    # ============================================================
    "a12JO000000M6jYYAS": (
        "7-1",
        "B 판매 등록 — 나머지 1개 부분 판매",
        (
            "B 판매 : SAL202605080053  /  원주문 : SOR202605080042 (백화점 주문번호2)\n"
            "(Salesforce Id : 801JO00000I9pAsYAJ)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 판매 입금(주문 입금 차감) : 140,000원  (DepositType 01 · PaymentMethod 01 · DEP202605080005-1)\n"
            "- 판매 입금(잔여) : 360,000원  (DepositType 02 · PaymentMethod 01 · DEP202605080007-1)\n"
            "- 적립포인트(쇼핑포인트) : 20,000P  (TJ Purchase Accrual / CD_TYPE_POINT 01)\n"
            "  ※ API 응답 PT_TARGET(1% 기준) : 5,000P\n"
            "- B 판매 합계 : 140,000(B주문 차감) + 360,000(판매 입금) = 500,000원 ✓\n"
            "- (백화점 주문번호2 1,400,000 = A 판매 900,000 + B 판매 500,000 으로 분할 소진)"
        ),
    ),
    "a12JO000000M54JYAS": (
        "7-2",
        "이벤트포인트 강제 지급 (B 판매 기준)",
        (
            "GdPointCreditApiControllerV1 — /gd/v1/point/credit\n"
            "- B 판매(SAL202605080053) 기준 이벤트포인트 강제 지급 (Api Forced Credit)\n"
            "- TX_REMARK : 이벤트포인트\n"
            "- 적립포인트(이벤트) : 3,000P\n"
            "- CD_TYPE_SAVE : Credit / CD_TYPE_POINT : 01\n"
            "- 유효기간 : 2026-05-08 09:00 ~ 2028-05-08 (소멸예정 2028-05-09)\n"
            "- CD_JOURNAL : 05wJO0000004AaHYAU"
        ),
    ),
    "a12JO000000M6l7YAC": (
        "7-3",
        "B 판매 + 이벤트포인트 후 회원 도움창 조회",
        (
            "GdMemberHelpInfoApiControllerV1\n"
            "- B 판매(SAL202605080053) 등록 및 이벤트포인트(3,000P) 지급 후 도움창 호출\n"
            "- 회원 포인트/거래 정보 응답 확인"
        ),
    ),
    "a12JO000000M6LNYA0": (
        "7-4",
        "B 판매 반품 등록",
        (
            "반품 거래 : SAL202605080055  /  원거래 : SAL202605080053 (B 판매)\n"
            "(Salesforce Id : 801JO00000IA0qsYAD)\n"
            "- Quantity : -1 (반품)\n"
            "- 환불 금액 : 500,000원  (판매 입금 취소 / DepositType 03 · PaymentMethod 01 · DEP202605080008-1)\n"
            "  ※ Sale Deposit Reversal — B 판매 시 입금된 500,000원 전액 환불\n"
            "- 복원 포인트(구매 적립 취소) : 20,000P  (TJ Cancel Purchase Accrual)\n"
            "- 이벤트포인트 강제 차감 : 3,000P  (TJ Api Forced Debit / JournalReason : FULL_RETURN_ECO_POINT_CANCEL)\n"
            "  ※ B 판매 시 지급된 이벤트포인트 3,000P 회수\n"
            "- 시나리오 검증: 「주문 시 등록했던 3% 쿠폰은 주문 단계에서 환불 등록했으므로 부분 판매 처리 및 환불 가능」"
        ),
    ),
    "a12JO000000M6oLYAS": (
        "7-5",
        "B 판매 반품 후 회원 도움창 조회",
        (
            "GdMemberHelpInfoApiControllerV1\n"
            "- 반품 직후 도움창 호출\n"
            "- 복원 포인트(20,000P) + 이벤트포인트 강제 차감(3,000P) 결과 확인"
        ),
    ),
    "a12JO000000M6pxYAC": (
        "7-6",
        "B 판매 반품 후 회원 도움창 재조회",
        "동일 회원 재조회 — 응답 동일",
    ),
    # ============================================================
    # 8. 아울렛 판매 (A 주문 기반)
    # ============================================================
    "a12JO000000M6uqYAC": (
        "8-1",
        "아울렛 판매 등록 (C 판매)",
        (
            "C 판매 : SAL202605080056  /  원주문 : SOR202605080040 (A 주문 — 아울렛)\n"
            "(Salesforce Id : 801JO00000IA1OfYAL)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 판매 입금(주문 입금 차감) : 50,000원  (DepositType 01 · PaymentMethod 01 · DEP202605080002-1)\n"
            "- 쿠폰 사용(주문 시 사용분 매칭) : 10,000원  (COP2026MY0014 / PaymentMethod 90 · DEP202605080002-2)\n"
            "- 판매 입금(잔여) : 440,000원  (DepositType 02 · PaymentMethod 01 · DEP202605080009-1)\n"
            "- C 판매 합계 : 50,000 + 10,000 + 440,000 = 500,000원 ✓\n"
            "- 시나리오 검증: 「아울렛 제품 판매 처리 시 백화점 구매 시 적립된 포인트 사용 시도 → 아울렛은 포인트 적용 불가하게 조회되지 않는 결과」"
        ),
    ),
    # ============================================================
    # 9. 아울렛 판매 후 — 포인트 적립 안 됨 확인
    # ============================================================
    "a12JO000000M6wPYAS": (
        "9-1",
        "아울렛 판매 후 회원 도움창 조회 (포인트 적립 검증)",
        (
            "GdMemberHelpInfoApiControllerV1\n"
            "- C 판매(SAL202605080056) 등록 후 도움창 호출 (시나리오 최종)\n"
            "- 시나리오 검증: 「아울렛 판매 건은 포인트 적립되지 않음」 확인\n"
            "- 누적 적립/사용/복원 이력 응답 확인"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    wb_src = load_workbook(SRC, read_only=True)
    ws_src = wb_src.active
    headers = [c.value for c in next(ws_src.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    src_rows = []
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        rid = row[idx["Id"]]
        if rid not in SCENARIO:
            continue
        src_rows.append({
            "Id": rid,
            "CreatedDate": row[idx["CreatedDate"]],
            "ApexClass__c": row[idx["ApexClass__c"]],
            "RequestUrl__c": row[idx["RequestUrl__c"]],
            "RequestBody__c": row[idx["RequestBody__c"]] or "",
            "ResponseBody__c": row[idx["ResponseBody__c"]] or "",
        })

    # SCENARIO 의 번호 기준으로 정렬 (자연 정렬)
    def sort_key(rec):
        no = SCENARIO[rec["Id"]][0]
        major, _, minor = no.partition("-")
        return (int(major), int(minor) if minor.isdigit() else 0, str(rec["CreatedDate"] or ""))

    src_rows.sort(key=sort_key)
    print(f"[matched] {len(src_rows)} of {len(SCENARIO)} expected scenario rows")

    wb = Workbook()
    ws = wb.active
    ws.title = SCENARIO_VERSION

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "ApexClass__c",
        "로그 ID",
        "CreatedDate",
        "RequestBody__c",
        "ResponseBody__c",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)
    for r_idx, rec in enumerate(src_rows, start=2):
        no, label, criterion = SCENARIO[rec["Id"]]
        url = rec["RequestUrl__c"] or ""
        domain = classify_domain(rec["ApexClass__c"], url)
        method = http_method(rec["ApexClass__c"], url)

        values = [
            SCENARIO_VERSION,
            no,
            label,
            criterion,
            MEMBER_NO,
            domain,
            trim(url),
            method,
            trim(rec["ApexClass__c"]),
            trim(rec["Id"]),
            trim(rec["CreatedDate"]),
            trim(rec["RequestBody__c"]),
            trim(rec["ResponseBody__c"]),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 7, 36, 80, 18, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    DOWNLOADS.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[saved] {OUT}")


if __name__ == "__main__":
    main()
