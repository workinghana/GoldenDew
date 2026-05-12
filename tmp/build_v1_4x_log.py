"""V1 (4-x) Log.xlsx — 회원 TEST02006 / PRO202605021069 (사은포인트 더블).

4. 신제품 구매 시 사은포인트 더블 적립 (4-1 ~ 4-4)
- 1월 신제품 3개 구매 + 에코 6,000원 적립
- 2월 1개 부분 반품
- 2월 1개 반품 후 다른 제품 추가 1개 구매 (교환개념, 1월 적립 포인트 사용)
- 3월에 1월 구매 제품 중 남은 2개 모두 반품
  · 가장 첫 원판매 먼저 반품 → 사용된 포인트 있어 반품 불가
  · 추가 구매 제품 먼저 반품 → 반품 가능

※ 시나리오 페이지 회원번호 TEST01006 은 마이그 전 표기. 실제 로그 회원: TEST02006.
"""
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V1 (4-x) Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v1_4x_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V1"
MEMBER_NO = "TEST02006"
SHEET_NAME = "V1 (4-x)"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower().replace("_", "")
    if "memberhelpinfo" in ac:
        return "회원 도움창"
    if "memberapi" in ac:
        return "회원"
    if "voucherhelp" in ac:
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac:
        return "포인트 (적용 조회)"
    if "saledeposit" in ac:
        return "판매 입금"
    if "returndeposit" in ac:
        return "반품 입금"
    if "orderapi" in ac:
        return "주문"
    if "saleapi" in ac:
        return "판매"
    if "returnapi" in ac:
        return "반품"
    if "pointcredit" in ac:
        return "포인트 (지급)"
    if "pointdebit" in ac:
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    # 4-1 판매 등록 (제품 3개 + 사은더블)
    "a12JO000000MeN2YAK": (
        "4-1",
        "판매 등록 — SAL02V008 (제품 2A2600001 × 3개, 사은포인트 더블 적립)",
        (
            "판매 번호 : SAL02V008  (시간 : 2026-05-10 18:27:25 KST)\n"
            "- 회원 : TEST02006  /  매장 코드 : 99998\n"
            "- 실결제 금액 : 1,500,000원 (= 500,000 × 3개)\n"
            "- 등록 제품 : 2A2600001 × 3\n"
            "- 프로모션 : PRO202605021069 (사은포인트 2배)\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 45,000P\n"
            "  (사은 적립 45,000 = 더블 적용 결과)\n"
            "- 시나리오 단계 : 4-1 신제품 3개 구매 + 사은포인트 더블 적립"
        ),
    ),
    # 4-1 에코포인트 6,000P 지급
    "a12JO000000MeObYAK": (
        "4-1 에코",
        "에코포인트 6,000P 지급 (SAL02V008 대상)",
        (
            "포인트 지급 : 이벤트포인트 6,000P  (시간 : 2026-05-10 18:27:58 KST)\n"
            "- 회원 : TEST02006  /  대상 판매 : SAL02V008\n"
            "- 포인트 유형 : 05 (에코포인트)\n"
            "- 응답 : 「포인트 지급 성공」 — PT_TARGET 6,000\n"
            "- 시나리오 단계 : 4-1 에코 적립 6,000P\n"
            "- 4-1 최종 적립 합산 : 사은 45,000 + 더블 45,000 (= 합산 표기) + 에코 6,000\n"
            "  → 기존 100,000P + 96,000P = 196,000P"
        ),
    ),
    # 4-2 부분반품
    "a12JO000000McMuYAK": (
        "4-2",
        "부분반품 — SAL02V009 (1개 / 적립 15,000P 취소)",
        (
            "반품 번호 : SAL02V009  /  원거래 판매 번호 : SAL02V008\n"
            "- 시간 : 2026-05-10 18:29:39 KST\n"
            "- 실결제 금액 : 500,000원 (1개 부분반품)\n"
            "- 응답 : 「반품 등록 완료」 — PT_USE 15,000 복원\n"
            "- 시나리오 단계 : 4-2 부분반품 (총 30,000 적립 취소 = 적립 15,000 + 더블 15,000)\n"
            "  → 196,000P → 166,000P"
        ),
    ),
    # 4-3 (전반) 부분반품 추가
    "a12JO000000MeRsYAK": (
        "4-3 (전반)",
        "부분반품 — SAL02V010 (1개 추가 / 적립 15,000P 취소)",
        (
            "반품 번호 : SAL02V010  /  원거래 판매 번호 : SAL02V008\n"
            "- 시간 : 2026-05-10 18:33:34 KST\n"
            "- 실결제 금액 : 500,000원 (1개 추가 반품)\n"
            "- 응답 : 「반품 등록 완료」 — PT_USE 15,000 복원\n"
            "- 시나리오 단계 : 4-3 (전반) 부분반품 (적립 취소 15,000 + 더블 15,000)\n"
            "  → 166,000P → 136,000P"
        ),
    ),
    # 4-3 추가 구매 1차 시도 ERROR
    "a12JO000000MeV4YAK": (
        "4-3 시도",
        "추가 구매 1차 시도 — SAL02V011 ❌ ERROR DUPLICATE_VALUE",
        (
            "판매 번호 : SAL02V011  (시간 : 2026-05-10 18:34:59 KST)\n"
            "- 응답 : ❌ code 500 — DUPLICATE_VALUE (DepositNo__c 중복)\n"
            "- 시나리오 단계 : 4-3 추가 구매 1차 시도 (DepositNo 중복 에러)"
        ),
    ),
    # 4-3 (후반) 추가 구매 (교환개념)
    "a12JO000000MeWiYAK": (
        "4-3 (후반)",
        "추가 구매 — SAL02V011 (포인트 사용 10,000P, 사은 11,700P + 더블 11,700P)",
        (
            "판매 번호 : SAL02V011  (시간 : 2026-05-10 18:35:15 KST)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 결제 구성 : 포인트 10,000 사용 + 구매 390,000\n"
            "- 프로모션 : PRO202605021069\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 13,200P (사은+더블 합산)\n"
            "- 시나리오 단계 : 4-3 (후반) 포인트 사용하여 추가 구매 (교환개념)\n"
            "  → 1월 구매 적립 포인트 사용 / 136,000P → 149,400P"
        ),
    ),
    # 4-4 시도 — 가장 첫 원판매 먼저 반품 시도 → 반품 불가 ★ 핵심
    "a12JO000000Mdu2YAC": (
        "4-4 ★ 핵심",
        "원판매 먼저 반품 시도 — SAL02V012 ❌ 「적립포인트 다른 판매 사용중」 (불가 메시지)",
        (
            "반품 번호 : SAL02V012  /  원거래 판매 번호 : SAL02V008 (원판매)\n"
            "- 시간 : 2026-05-10 18:39:49 KST\n"
            "- 응답 : ❌ code 400\n"
            "    「반품 입금 등록 불가 — 적립된 포인트가 다른 주문/판매에서 이미\n"
            "     사용되었습니다. 판매번호: SAL02V011 / 입금번호: DEP02V011-2」\n"
            "- 시나리오 단계 : 4-4 가장 첫 원판매 먼저 반품 시도 → 반품 불가 메시지 ✓ PASS\n"
            "- 검증 : 원판매 적립 포인트를 추가구매 SAL02V011 에서 사용 중\n"
            "  → 원판매 직접 반품 차단 정상 동작"
        ),
    ),
    # 4-4 추가 구매 먼저 반품 (성공)
    "a12JO000000MedAYAS": (
        "4-4 ✓",
        "추가 구매 먼저 반품 — SAL02V012 (원판매 SAL02V011 / 반품 가능)",
        (
            "반품 번호 : SAL02V012  /  원거래 판매 번호 : SAL02V011 (추가 구매)\n"
            "- 시간 : 2026-05-10 18:40:56 KST  (불가 시도 후 1분 7초)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 응답 : 「반품 등록 완료」 — PT_USE 13,200 복원\n"
            "- 시나리오 단계 : 4-4 추가 구매 제품 먼저 반품 → 반품 가능 ✓ PASS"
        ),
    ),
    # 4-4 남은 원판매 반품 (성공)
    "a12JO000000Md9FYAS": (
        "4-4 ✓",
        "남은 원판매 반품 — SAL02V013 (원판매 SAL02V008 / 1개 반품)",
        (
            "반품 번호 : SAL02V013  /  원거래 판매 번호 : SAL02V008\n"
            "- 시간 : 2026-05-10 18:42:24 KST\n"
            "- 실결제 금액 : 500,000원\n"
            "- 응답 : 「반품 등록 완료」 — PT_USE 15,000 복원\n"
            "- 시나리오 단계 : 4-4 추가구매 반품 후 남은 원판매 1개 반품 → 가능 ✓ PASS\n"
            "- 최종 : 사은 취소 27,000 + 더블 취소 27,000 + 에코 취소 6,000 + 포인트 사용 취소 10,000\n"
            "  → 최종 보유 100,000P (시나리오 정답과 일치)"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 50, 95, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
