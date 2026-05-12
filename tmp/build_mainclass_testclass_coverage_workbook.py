import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
RUN_DIR = ROOT / "tmp" / "coverage_all_20260504"
RESULT_JSON = RUN_DIR / "test-result-707JO00000ESofA.json"
CLASS_COVERAGE_JSON = RUN_DIR / "test-result-codecoverage.json"
CLASSES_DIR = ROOT / "force-app" / "main" / "default" / "classes"


def load_json(path):
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def read_text(path):
    return path.read_text(encoding="utf-8", errors="ignore")


def is_test_class(path):
    text = read_text(path)
    return bool(re.search(r"@isTest\b|testMethod\b", text, re.IGNORECASE))


def pct(covered, total):
    if not total:
        return None
    return round(covered * 100 / total, 2)


def status_label(coverage):
    if coverage is None:
        return "75% 미만"
    return "75% 이상" if coverage >= 75 else "75% 미만"


def comma(values):
    values = [v for v in values if v]
    return ", ".join(sorted(set(values)))


def infer_test_classes(main_class, test_class_names):
    candidates = []
    exact_patterns = [
        f"{main_class}Test",
        f"Test{main_class}",
        f"{main_class}CoverageTest",
        f"{main_class}SeeAllTest",
    ]
    for name in exact_patterns:
        if name in test_class_names:
            candidates.append(name)

    for test_name in test_class_names:
        if test_name in candidates:
            continue
        normalized = test_name.lower()
        target = main_class.lower()
        if target in normalized and ("test" in normalized or normalized.startswith("test")):
            candidates.append(test_name)
    return sorted(set(candidates))


def add_sheet(wb, title, headers, rows, tab_color=None):
    ws = wb.create_sheet(title)
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, idx, header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(value, str) and len(value) > 60)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, header in enumerate(headers, start=1):
        width = len(str(header))
        for row_idx in range(2, min(ws.max_row, 250) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, min(len(str(value)), 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 82)
    return ws


def main(output_path):
    result = load_json(RESULT_JSON)
    result = result.get("result", result)
    summary = result["summary"]
    coverage_records = result["coverage"]["records"]
    class_coverage = load_json(CLASS_COVERAGE_JSON)

    local_main = {}
    local_tests = {}
    for path in CLASSES_DIR.glob("*.cls"):
        rel = str(path.relative_to(ROOT))
        if is_test_class(path):
            local_tests[path.stem] = rel
        else:
            local_main[path.stem] = rel

    coverage_by_class = {}
    for row in class_coverage:
        name = row["name"]
        lines = row.get("lines") or {}
        total = row.get("totalLines") or len(lines)
        covered = row.get("totalCovered")
        if covered is None:
            covered = sum(1 for value in lines.values() if value)
        coverage_by_class[name] = {
            "total": total,
            "covered": covered,
            "uncovered": max(total - covered, 0),
            "percent": row.get("coveredPercent", pct(covered, total)),
        }

    test_outcomes = defaultdict(lambda: {"Pass": 0, "Fail": 0, "Skip": 0, "Other": 0})
    for test in result.get("tests", []):
        test_class = (test.get("ApexClass") or {}).get("Name")
        outcome = test.get("Outcome") or "Other"
        bucket = outcome if outcome in ("Pass", "Fail", "Skip") else "Other"
        test_outcomes[test_class][bucket] += 1

    by_main_test = defaultdict(lambda: defaultdict(lambda: {
        "methods": set(),
        "covered_lines": set(),
        "uncovered_lines": set(),
        "covered_sum": 0,
        "uncovered_sum": 0,
    }))

    for rec in coverage_records:
        main_class = (rec.get("ApexClassOrTrigger") or {}).get("Name")
        test_class = (rec.get("ApexTestClass") or {}).get("Name")
        if main_class not in local_main or not test_class:
            continue
        data = by_main_test[main_class][test_class]
        method = rec.get("TestMethodName")
        if method:
            data["methods"].add(method)
        covered_lines = rec.get("Coverage", {}).get("coveredLines") or []
        uncovered_lines = rec.get("Coverage", {}).get("uncoveredLines") or []
        data["covered_lines"].update(covered_lines)
        data["uncovered_lines"].update(uncovered_lines)
        data["covered_sum"] += rec.get("NumLinesCovered") or len(covered_lines)
        data["uncovered_sum"] += rec.get("NumLinesUncovered") or len(uncovered_lines)

    all_rows = []
    below_rows = []
    above_rows = []
    test_detail_rows = []

    for main_class in sorted(local_main):
        cov = coverage_by_class.get(main_class, {})
        total = cov.get("total")
        covered = cov.get("covered")
        uncovered = cov.get("uncovered")
        coverage_pct = cov.get("percent")
        status = status_label(coverage_pct)
        actual_tests = sorted(by_main_test.get(main_class, {}).keys())
        inferred_tests = infer_test_classes(main_class, set(local_tests))
        test_classes = actual_tests or inferred_tests
        test_exists = "Y" if test_classes else "N"
        actual_test_text = comma(actual_tests)
        inferred_test_text = comma(inferred_tests)
        note = ""
        if not coverage_pct and not actual_tests:
            note = "No detailed coverage records in this run"
        elif inferred_tests and not actual_tests:
            note = "Matching test class exists, but no coverage record returned"

        row = [
            status,
            main_class,
            coverage_pct,
            total,
            covered,
            uncovered,
            test_exists,
            actual_test_text,
            inferred_test_text,
            len(actual_tests),
            local_main[main_class],
            note,
        ]
        all_rows.append(row)
        if status == "75% 이상":
            above_rows.append(row)
        else:
            below_rows.append(row)

        if actual_tests:
            for test_class in actual_tests:
                data = by_main_test[main_class][test_class]
                unique_covered = len(data["covered_lines"])
                outcomes = test_outcomes[test_class]
                test_detail_rows.append([
                    status,
                    main_class,
                    coverage_pct,
                    total,
                    test_class,
                    len(data["methods"]),
                    unique_covered,
                    pct(unique_covered, total),
                    data["covered_sum"],
                    data["uncovered_sum"],
                    outcomes["Pass"],
                    outcomes["Fail"],
                    local_tests.get(test_class, ""),
                ])
        elif inferred_tests:
            for test_class in inferred_tests:
                outcomes = test_outcomes[test_class]
                test_detail_rows.append([
                    status,
                    main_class,
                    coverage_pct,
                    total,
                    test_class,
                    None,
                    None,
                    None,
                    None,
                    None,
                    outcomes["Pass"],
                    outcomes["Fail"],
                    local_tests.get(test_class, ""),
                ])
        else:
            test_detail_rows.append([
                status,
                main_class,
                coverage_pct,
                total,
                "",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "",
            ])

    summary_rows = [
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Source Test Run ID", summary.get("testRunId")],
        ["Source Test Start Time", summary.get("testStartTime")],
        ["Source Org", summary.get("username")],
        ["Run Outcome", summary.get("outcome")],
        ["Tests Ran", summary.get("testsRan")],
        ["Passing", summary.get("passing")],
        ["Failing", summary.get("failing")],
        ["Org Wide Coverage", summary.get("orgWideCoverage")],
        ["Test Run Coverage", summary.get("testRunCoverage")],
        ["Local Main Classes", len(local_main)],
        ["Local Test Classes", len(local_tests)],
        ["Main Classes >= 75%", len(above_rows)],
        ["Main Classes < 75%", len(below_rows)],
        ["Rows With Actual Coverage Test Class", sum(1 for r in all_rows if r[7])],
        ["Note", "Apex 본 클래스 파일은 수정하지 않고, 기존 전체 테스트 실행 결과로 엑셀만 생성함"],
    ]

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "Summary", ["항목", "값"], summary_rows)
    headers = [
        "구분",
        "본 클래스",
        "본 클래스 전체 커버리지 %",
        "총 라인",
        "커버 라인",
        "미커버 라인",
        "테스트 클래스 존재",
        "실제 커버한 테스트 클래스",
        "이름상 매칭 테스트 클래스",
        "실제 커버 테스트 클래스 수",
        "본 클래스 파일",
        "비고",
    ]
    add_sheet(wb, "75미만_수정대상", headers, below_rows, "C00000")
    add_sheet(wb, "75이상", headers, above_rows, "00A65A")
    add_sheet(wb, "전체_본클래스", headers, all_rows)
    add_sheet(wb, "테스트클래스별_상세", [
        "구분",
        "본 클래스",
        "본 클래스 전체 커버리지 %",
        "본 클래스 총 라인",
        "테스트 클래스",
        "커버 발생 테스트 메서드 수",
        "해당 테스트 클래스 고유 커버 라인",
        "해당 테스트 클래스 기준 커버리지 %",
        "커버 라인 합계",
        "미커버 라인 합계",
        "테스트 클래스 성공 메서드 수",
        "테스트 클래스 실패 메서드 수",
        "테스트 클래스 파일",
    ], test_detail_rows)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(output)
    print(f"main={len(local_main)} below75={len(below_rows)} above75={len(above_rows)} detailRows={len(test_detail_rows)}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python build_mainclass_testclass_coverage_workbook.py <output.xlsx>")
    main(sys.argv[1])
