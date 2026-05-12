"""V4-6 Log.xlsx — 회원 123002600463 SVIP, 시나리오 V4 / 6-1 ~ 6-2.

「6. 예술의전당 프로모션 이용」 — [TEST] 통합시나리오V4_골드바_행사
- 6-1 판매처리 — 구매 + 사은 포인트 미적립 (골드바 적립 제외)
- 6-2 에코포인트 지급 불가 — 골드제품 구매 시 에코포인트창 미노출 (ERP 수정 5/5 최과장님)
"""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V4-6 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v4_6_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V4"
MEMBER_NO = "123002600463"
SHEET_NAME = "V4-6"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac.replace("_", ""):
        return "포인트 (적용 조회)"
    if "saledeposit" in ac.replace("_", ""):
        return "판매 입금"
    if "returndeposit" in ac.replace("_", ""):
        return "반품 입금"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    return "POST"


SCENARIO = {
    # ───── 6-1 : 판매처리 — 구매 + 사은 포인트 미적립 ─────
    "a12JO000000MDl3YAG": (
        "6-1",
        "골드바 판매처리 — 구매 + 사은 포인트 미적립 확인",
        (
            "판매 번호 : SAL202605080161  /  원거래 주문 번호 : SOR202605080146\n"
            "- 시간 : 2026-05-08 14:38:08 KST\n"
            "- 매장 코드 : 99998 (현대 백화점)\n"
            "- 프로모션 번호 : PRO202604171052  ([TEST] 통합시나리오V4_골드바_행사)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 판매 항목 :\n"
            "    1) 상품 코드 382600002 (골드바)  /  수량 1  /  순매가 단가 500,000  /  할인율 0%\n"
            "- 거래 원장 (입금 번호 DEP202605080169) :\n"
            "    1) 결제 수단 01 (현금)  /  입금 구분 01 (계약금)  /  거래 금액 500,000원 (전액 현금 결제)\n"
            "- 응답 메시지 : 「판매 등록 완료」  /  success: true  /  code: 200\n"
            "- 응답 적립 포인트 목록 : (없음) ✓ — 「CreditPointList」 항목 자체가 없음\n"
            "- 시나리오 검증 : 「구매 + 사은 포인트 미적립」 ✓\n"
            "- ※ 일반 판매라면 PT_TARGET 「구매포인트 × 2.00」 + 「사은포인트 × 3.00」 가 등장해야 하나, 골드바 행사 프로모션 + 골드제품 분류로 적립 자체가 트리거되지 않음"
        ),
    ),
    # ───── 6-2 : 에코포인트 지급 불가 (골드바 isGoldenBar:True) ─────
    "a12JO000000MDWXYA4": (
        "6-2",
        "에코포인트 지급 불가 — 골드제품 구매 시 에코포인트창 미노출 (ERP 차단)",
        (
            "주문 번호 : SOR202605080146  (시간 : 2026-05-08 14:36:30 KST)\n"
            "- 매장 코드 : 99998  /  유형 : 01 (계약)\n"
            "- 프로모션 번호 : PRO202604171052\n"
            "- 실결제 금액 : 500,000원\n"
            "- 주문 항목 : 상품 코드 382600002 (골드바) / 수량 1 / 순매가 단가 500,000 / 할인율 0%\n"
            "- 거래 원장 (입금 번호 DEP202605080169) :\n"
            "    1) 결제 수단 01 (현금)  /  입금 구분 01  /  거래 금액 500,000원\n"
            "- 응답 핵심 데이터 : 「isGoldenBar: True」 ✓ ← 골드바 식별 플래그 정상 노출\n"
            "- 응답 메시지 : 「주문 등록 완료」\n"
            "- 시나리오 검증 :\n"
            "    1) 골드제품 (isGoldenBar:True) 식별 ✓\n"
            "    2) 6-1 판매처리 후 후속 GdPointCreditApiControllerV1 호출이 audit 에 없음 → ERP 화면에서 「에코포인트창」 미노출로 사용자가 호출하지 않은 것으로 추정 ✓\n"
            "    3) 비교 : V4-2-2 ([TEST] 사은포인트3배) 같은 일반 행사에서는 판매 직후 에코포인트 9,000P PointCredit 호출이 발생 — 해당 호출 부재 = 에코포인트 차단 효과\n"
            "- ※ ERP 수정일자 5/5 최과장님 — 골드제품 분기에서 에코포인트창 비표시 처리 확인"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    def num_sort_key(item):
        no = item[1][0]
        parts = no.split("-")
        try:
            return tuple(int(p) for p in parts)
        except ValueError:
            return (99, 99, 99)

    sorted_items = sorted(SCENARIO.items(), key=num_sort_key)

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 10, 40, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
