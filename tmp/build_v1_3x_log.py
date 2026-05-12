"""V1 (3-x) Log.xlsx — 회원 TEST02006 / VIP 무제한 정률쿠폰 COP2026MY0011.

3. VIP 회원 무제한 정률쿠폰 사용 건 확인 (3-1 ~ 3-4)
- 사용 시 자동 발행
- 환불 시 쿠폰 복원 로직 확인
- 1차 쿠폰 사용분 반품 시 2차 쿠폰 사라지고 1차 쿠폰 복원되는지 확인

※ 시나리오 페이지 회원번호 193001000334 은 마이그 전 표기. 실제 로그 회원: TEST02006.
"""
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V1 (3-x) Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v1_3x_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V1"
MEMBER_NO = "TEST02006"
SHEET_NAME = "V1 (3-x)"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower().replace("_", "")
    if "memberhelpinfo" in ac:
        return "회원 도움창"
    if "memberapi" in ac:
        return "회원"
    if "voucherhelp" in ac:
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac:
        return "포인트 (적용 조회)"
    if "saledeposit" in ac:
        return "판매 입금"
    if "returndeposit" in ac:
        return "반품 입금"
    if "orderapi" in ac:
        return "주문"
    if "saleapi" in ac:
        return "판매"
    if "returnapi" in ac:
        return "반품"
    if "pointcredit" in ac:
        return "포인트 (지급)"
    if "pointdebit" in ac:
        return "포인트 (사용/차감)"
    if "erpasync" in ac or "asyncjob" in ac:
        return "ERP 콜아웃 (쿠폰 동기화)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    # 3-1 판매 저장 (무제한 쿠폰 사용)
    "a12JO000000Me8aYAC": (
        "3-1",
        "판매 저장 — SAL02V006 (1차 무제한 쿠폰 COP2026MY0011 사용)",
        (
            "판매 번호 : SAL02V006  (시간 : 2026-05-10 18:18:37 KST)\n"
            "- 회원 : TEST02006  /  매장 코드 : 99998\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 프로모션 : PRO202604161041\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 18,000P\n"
            "- 시나리오 단계 : 3-1 판매 저장 (VIP 무제한 정률쿠폰 사용)"
        ),
    ),
    # 3-2 무제한 쿠폰 재발행 — ERP 콜아웃 1차 (mil 00218247)
    "a12JO000000MeI9YAK": (
        "3-2 (1차)",
        "ERP 쿠폰 동기화 콜아웃 — no_man_coupon_mil: 00218247 (1차 쿠폰)",
        (
            "Apex 클래스 : ErpAsyncJob (CALLOUT)\n"
            "- 시간 : 2026-05-10 18:16:41 KST\n"
            "- URL : callout:GoldenDewApiNC/Crm/CouponDetail\n"
            "- 요청 : {\"no_mbr\":\"TEST02006\", \"no_man_coupon_mil\":\"00218247\",\n"
            "         \"no_man_coupon\":\"COP2026MY0011\", \"id_insert\":\"milvus\"}\n"
            "- 응답 : ❌ {\"code\":0, \"success\":false,\n"
            "         \"message\":\"SQL - 회원이 존재하지 않습니다.\"}\n"
            "- 시나리오 단계 : 3-2 — 1차 쿠폰 사용 시 ERP 동기화 콜아웃 (회원 미존재로 실패)\n"
            "- ※ 시나리오 메모상 1차 쿠폰 mil 코드 = 9TKSHRYN 으로 표기 / 실제 로그 = 00218247"
        ),
    ),
    # 3-2 무제한 쿠폰 재발행 — ERP 콜아웃 2차 (mil AKU33FUQ)
    "a12JO000000Mdh4YAC": (
        "3-2 (2차)",
        "ERP 쿠폰 동기화 콜아웃 — no_man_coupon_mil: AKU33FUQ (2차 재발행 쿠폰)",
        (
            "Apex 클래스 : ErpAsyncJob (CALLOUT)\n"
            "- 시간 : 2026-05-10 18:18:37 KST\n"
            "- URL : callout:GoldenDewApiNC/Crm/CouponDetail\n"
            "- 요청 : {\"no_mbr\":\"TEST02006\", \"no_man_coupon_mil\":\"AKU33FUQ\",\n"
            "         \"no_man_coupon\":\"COP2026MY0011\", \"id_insert\":\"milvus\"}\n"
            "- 응답 : ❌ {\"code\":0, \"success\":false,\n"
            "         \"message\":\"SQL - 회원이 존재하지 않습니다.\"}\n"
            "- 시나리오 단계 : 3-2 무제한 쿠폰 재발행 — 2차 쿠폰 AKU33FUQ 자동 발행\n"
            "  (다시 무제한 쿠폰 조회 시 신규 쿠폰 등록 콜아웃)"
        ),
    ),
    # 3-3 무제한 쿠폰 사용 판매건 반품 등록
    "a12JO000000MeJoYAK": (
        "3-3",
        "무제한 쿠폰 사용 판매건 반품 등록 — SAL02V007 (원판매 SAL02V006)",
        (
            "반품 번호 : SAL02V007  /  원거래 판매 번호 : SAL02V006\n"
            "- 시간 : 2026-05-10 18:22:18 KST\n"
            "- 실결제 금액 : 1,000,000원 (완전 반품, -2 수량)\n"
            "- 응답 : 「반품 등록 완료」 — PT_USE 18,000 복원\n"
            "- 시나리오 단계 : 3-3 무제한 쿠폰 사용한 판매건 반품 등록\n"
            "  (판매 SAL02V006 / 반품 SAL02V007)\n"
            "- 검증 메모 : 「2차 쿠폰 발행 이후에 1차 쿠폰 사용분 반품 처리 시\n"
            "  2차 쿠폰 사라지고 / 1차 쿠폰 다시 복원 됨」\n"
            "- 3-4 (반품등록시 이전+신규 동시 조회) 는 UI 호출이라 audit log 미기록"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 50, 95, 16, 24, 42, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
