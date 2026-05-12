import csv
import json
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
TMP = ROOT / "tmp"
OUT_PATH = TMP / "apex_audit_trail_export_20260508.xlsx"
TARGET_ORG = "goldendew-sb"
SF_CMD = r"C:\Program Files\sf\bin\sf.cmd"

EXCEL_CELL_LIMIT = 32767  # xlsx hard limit per cell

# 두 시간 사이(가장 이른 시작 ~ 가장 늦은 종료) 단일 구간
START = "2026-05-07T09:31:00.000+09:00"
END = "2026-05-08T09:53:00.000+09:00"

# 두 묶음의 번호를 모두 합쳐 매칭 (중복 제거)
NUMBERS = sorted({
    "269999800209",
    "SOR202605080040",
    "SOR202605080041",
    "SOR202605080042",
    "SAL202605080046",
    "SAL202605080053",
    "SAL202605080056",
    "SAL202605080055",
    "SAL202605070221",
    "SAL202605070223",
    "SAL202605070224",
    "SAL202605070231",
})


SOQL = (
    "SELECT Id, CreatedDate, ApexClass__c, RequestUrl__c, RequestBody__c, ResponseBody__c "
    "FROM ApexAuditTrail__c "
    f"WHERE CreatedDate >= {START} AND CreatedDate <= {END} "
    "ORDER BY CreatedDate ASC"
)


def run_bulk_query(soql):
    """Run SOQL via sf CLI bulk mode and return list of dict rows.

    LongTextArea fields cannot be filtered in WHERE, so we fetch the full
    time-range and post-filter in Python.
    """
    csv_out = TMP / "_apex_audit_trail_raw.csv"
    if csv_out.exists():
        csv_out.unlink()

    # sf data query with --bulk writes CSV to stdout when result-format=csv
    cmd = [
        SF_CMD,
        "data",
        "query",
        "--query",
        soql,
        "--target-org",
        TARGET_ORG,
        "--result-format",
        "csv",
    ]
    print(f"[run] sf data query --result-format csv (soql len={len(soql)})", flush=True)
    proc = subprocess.run(cmd, capture_output=True, shell=False)
    if proc.returncode != 0:
        sys.stderr.buffer.write(b"[stdout]\n")
        sys.stderr.buffer.write(proc.stdout[:4000])
        sys.stderr.buffer.write(b"\n[stderr]\n")
        sys.stderr.buffer.write(proc.stderr[:4000])
        raise RuntimeError(f"sf data query failed (exit {proc.returncode})")

    csv_out.write_bytes(proc.stdout)
    print(f"[saved-raw] {csv_out} ({csv_out.stat().st_size:,} bytes)", flush=True)

    rows = []
    with csv_out.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append(r)
    return rows


def matched_numbers(row, numbers):
    req = row.get("RequestBody__c") or ""
    res = row.get("ResponseBody__c") or ""
    hits = [n for n in numbers if (n in req) or (n in res)]
    return hits


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    rows = run_bulk_query(SOQL)
    print(f"[fetched] total {len(rows)} rows in range", flush=True)

    matched = []
    for r in rows:
        hits = matched_numbers(r, NUMBERS)
        if hits:
            matched.append((hits, r))
    print(f"[matched] {len(matched)} rows contain target numbers", flush=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "ApexAuditTrail"

    headers = [
        "MatchedNumbers",
        "Id",
        "CreatedDate",
        "ApexClass__c",
        "RequestUrl__c",
        "RequestBody__c",
        "ResponseBody__c",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, (hits, rec) in enumerate(matched, start=2):
        ws.cell(row=row_idx, column=1, value=", ".join(hits))
        ws.cell(row=row_idx, column=2, value=trim(rec.get("Id")))
        ws.cell(row=row_idx, column=3, value=trim(rec.get("CreatedDate")))
        ws.cell(row=row_idx, column=4, value=trim(rec.get("ApexClass__c")))
        ws.cell(row=row_idx, column=5, value=trim(rec.get("RequestUrl__c")))
        ws.cell(row=row_idx, column=6, value=trim(rec.get("RequestBody__c")))
        ws.cell(row=row_idx, column=7, value=trim(rec.get("ResponseBody__c")))

    widths = [30, 22, 28, 32, 50, 80, 80]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(OUT_PATH)

    # per-number hit counts (for sanity)
    counts = {n: 0 for n in NUMBERS}
    for hits, _ in matched:
        for n in hits:
            counts[n] += 1
    print()
    print("=== Summary ===")
    print(f"  range : {START} ~ {END}")
    print(f"  fetched: {len(rows)}")
    print(f"  matched: {len(matched)}")
    print(f"  saved : {OUT_PATH}")
    print("  hits per number:")
    for n in NUMBERS:
        print(f"    {n}: {counts[n]}")


if __name__ == "__main__":
    main()
