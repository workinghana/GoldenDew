"""V5-7 Log.xlsx — 회원 TEST02009 / [TEST]_프로모션없음.

포인트 분할로 사용된 판매 건 부분 반품 진행 시 — 정리 대상 팝업 호출 및
사전 정리 후 1차 판매 부분 반품 성공 여부 확인.
"""
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V5-7 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v5_7_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V5-7"
MEMBER_NO = "TEST02009"
SHEET_NAME = "V5-7"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower().replace("_", "")
    if "memberhelpinfo" in ac:
        return "회원 도움창"
    if "memberapi" in ac:
        return "회원"
    if "voucherhelp" in ac:
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac:
        return "포인트 (적용 조회)"
    if "saledeposit" in ac:
        return "판매 입금"
    if "returndeposit" in ac:
        return "반품 입금"
    if "orderapi" in ac:
        return "주문"
    if "saleapi" in ac:
        return "판매"
    if "returnapi" in ac:
        return "반품"
    if "pointcredit" in ac:
        return "포인트 (지급)"
    if "pointdebit" in ac:
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    # 1) 1차 판매 SAL04052
    "a12JO000000MfvsYAC": (
        "1) 1차",
        "1차 판매 등록 — SAL04052 (적립포인트 40,000P 생성)",
        (
            "판매 번호 : SAL04052  (시간 : 2026-05-10 20:34:28 KST)\n"
            "- 회원 : TEST02009  /  매장 코드 : 99998\n"
            "- 실결제 금액 : 2,000,000원\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 40,000P (구매포인트 2배 적용)\n"
            "- 시나리오 단계 : 1) 1차 판매 건으로 10만원 포인트 발생 (실제로는 4만원 분개)\n"
            "  ※ 이후 2~5차 판매에서 이 1차 적립 포인트를 분할 사용함"
        ),
    ),
    # 2) 2차 판매 SAL04053
    "a12JO000000MfZIYA0": (
        "2) 2차",
        "2차 판매 등록 — SAL04053 (1차 적립 3만원 사은 사용)",
        (
            "판매 번호 : SAL04053  (시간 : 2026-05-10 20:38:40 KST)\n"
            "- 실결제 금액 : 230,000원\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 4,000P\n"
            "- 시나리오 단계 : 2) 2차 판매 — 1차 적립 포인트 30,000 사은에서 사용처리"
        ),
    ),
    # 3) 3차 판매 SAL04054
    "a12JO000000MgNEYA0": (
        "3) 3차",
        "3차 판매 등록 — SAL04054 (1차 적립 4만원 사은+구매 사용)",
        (
            "판매 번호 : SAL04054  (시간 : 2026-05-10 20:49:38 KST)\n"
            "- 실결제 금액 : 540,000원\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 10,000P\n"
            "- 시나리오 단계 : 3) 3차 판매 — 1차 적립 포인트 40,000 사은+구매에서 사용처리"
        ),
    ),
    # 4) 4차 판매 SAL04055
    "a12JO000000MgTgYAK": (
        "4) 4차",
        "4차 판매 등록 — SAL04055 (1차 적립 3만원 구매 사용)",
        (
            "판매 번호 : SAL04055  (시간 : 2026-05-10 20:57:50 KST)\n"
            "- 실결제 금액 : 330,000원\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 6,000P\n"
            "- 시나리오 단계 : 4) 4차 판매 — 1차 적립 포인트 30,000 구매에서 사용처리"
        ),
    ),
    # 5) 5차 판매 SAL04056
    "a12JO000000MgYWYA0": (
        "5) 5차",
        "5차 판매 등록 — SAL04056 (1차 판매 이후 판매건의 사은 1만원 사용)",
        (
            "판매 번호 : SAL04056  (시간 : 2026-05-10 21:00:50 KST)\n"
            "- 실결제 금액 : 210,000원\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 4,000P\n"
            "- 시나리오 단계 : 5) 5차 판매 — 1차 판매건 이후 판매건의 사은에서 사용처리 10,000\n"
            "  ※ 이후 부분반품 에러 메시지에는 SAL04056 이 정리 대상에서 빠짐 (1차 적립 직접 사용 X)"
        ),
    ),
    # 6) 1차 부분반품 1차 시도 — ERROR DUPLICATE_VALUE
    "a12JO000000MggYYAS": (
        "6) 시도 1차 ⚠️",
        "1차 부분반품 1차 시도 — SAL04118 ❌ ERROR DUPLICATE_VALUE",
        (
            "반품 번호 : SAL04118  /  원거래 판매 번호 : SAL04052\n"
            "- 시간 : 2026-05-10 21:09:23 KST\n"
            "- 응답 : ❌ code 500 — DUPLICATE_VALUE (DepositNo__c 중복: 0lVJO000000CIKY2A4)\n"
            "- 시나리오 단계 : 6) 1차 시도 (DepositNo 중복 에러로 실패)"
        ),
    ),
    # 6) 1차 부분반품 2차 시도 — ERROR 「다른 판매 사용중」 ★ 시나리오 핵심
    "a12JO000000MgiAYAS": (
        "6) ★ 핵심",
        "1차 부분반품 재시도 — SAL04118 ❌ 「적립포인트 다른 판매 사용 중」 (PASS 시나리오)",
        (
            "반품 번호 : SAL04118  /  원거래 판매 번호 : SAL04052\n"
            "- 시간 : 2026-05-10 21:09:49 KST  (1차 시도 후 26초)\n"
            "- 응답 : ❌ code 400\n"
            "    「반품 입금 등록 불가 — 적립된 포인트가 다른 주문/판매에서 이미\n"
            "     사용되었습니다. 판매번호: SAL04053, SAL04054, SAL04055 /\n"
            "     입금번호: DEP040011-2, DEP040013-2, DEP040015-2」\n"
            "- 시나리오 단계 : 6) 1차 판매 부분 반품 진행 시 정리 대상 팝업 호출 ✓ PASS\n"
            "- 검증 : 2~6차 판매 중 1차 적립 포인트를 사용 한 SAL04053/54/55 가 표시됨\n"
            "  (SAL04056 은 1차 적립 직접 사용 X 라서 목록에서 제외)"
        ),
    ),
    # 7-1) SAL04053 사전 정리 — DEP040010 삭제
    "a12JO000000Mg0fYAC": (
        "7-1) 사전",
        "판매 입금 삭제 — DELETE DEP040010 (SAL04053 구매포인트)",
        (
            "Apex 클래스 : GdSaleDepositApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-10 21:12:52 KST\n"
            "- 대상 입금 번호 : DEP040010  /  DeposiSeqNo : 1\n"
            "- 응답 : 「판매 입금 삭제 삭제 완료」 — PT_USE 4,000 복원 (SAL04053)\n"
            "- 시나리오 단계 : 7-1) SAL04053 사전 정리 (구매포인트 입금 삭제)"
        ),
    ),
    # 7-1) SAL04053 사전 정리 — DEP040011 삭제 (사은포인트 30,000 복원)
    "a12JO000000MgNFYA0": (
        "7-1) 사전",
        "판매 입금 삭제 — DELETE DEP040011 (SAL04053 사은포인트 30,000P 복원)",
        (
            "Apex 클래스 : GdSaleDepositApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-10 21:12:57 KST\n"
            "- 대상 입금 번호 : DEP040011  /  DeposiSeqNo : 2\n"
            "- 응답 : 「판매 입금 삭제 삭제 완료」\n"
            "    RecreditPointList : 사은포인트 30,000 복원 (ExpiryDate 2028-05-04)\n"
            "- 시나리오 단계 : 7-1) SAL04053 사전 정리 (사은포인트 입금 삭제 → 1차 적립 복원)"
        ),
    ),
    # 7-1) SAL04053 판매 전체 삭제
    "a12JO000000MfscYAC": (
        "7-1) ✓",
        "판매 전체 삭제 — DELETE SAL04053 (시나리오 7-1)",
        (
            "Apex 클래스 : GdSaleApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-10 21:13:05 KST\n"
            "- 대상 판매 번호 : SAL04053\n"
            "- URL : {\"SaleNo__c\":\"SAL04053\"} (RequestBody 없음)\n"
            "- 응답 : 「판매 전체 삭제 완료」 (code 200)\n"
            "- 시나리오 단계 : 7-1) SAL04053 판매 삭제 완료"
        ),
    ),
    # 7-2) SAL04054 완전 반품
    "a12JO000000MglQYAS": (
        "7-2) ✓",
        "SAL04054 완전 반품 — SAL04119 (반품 등록 완료)",
        (
            "반품 번호 : SAL04119  /  원거래 판매 번호 : SAL04054\n"
            "- 시간 : 2026-05-10 21:25:29 KST\n"
            "- 실결제 금액 : 540,000원 (완전 반품)\n"
            "- 응답 : 「반품 등록 완료」 — PT_USE 10,000 복원\n"
            "- 시나리오 단계 : 7-2) SAL04054 완전 반품 완료"
        ),
    ),
    # SAL04119 중복 에러 (오타 — OriginalOrderId 가 SAL04055 로 변경됐는데 SaleNo__c 는 그대로)
    "a12JO000000MgmzYAC": (
        "(오타) ⚠️",
        "SAL04119 재사용 시도 — ❌ ERROR \"SaleNo__c already exists\"",
        (
            "반품 번호 : SAL04119  /  원거래 판매 번호 : SAL04055\n"
            "- 시간 : 2026-05-10 21:27:07 KST\n"
            "- 응답 : ❌ code 400 「SaleNo__c already exists」\n"
            "- 시나리오 단계 : 부수 — SAL04119 는 이미 SAL04054 반품에 사용됨\n"
            "  → SAL04120 으로 번호 변경하여 재시도"
        ),
    ),
    # 7-3) SAL04055 완전 반품
    "a12JO000000MgoeYAC": (
        "7-3) ✓",
        "SAL04055 완전 반품 — SAL04120 (반품 등록 완료)",
        (
            "반품 번호 : SAL04120  /  원거래 판매 번호 : SAL04055\n"
            "- 시간 : 2026-05-10 21:27:31 KST\n"
            "- 실결제 금액 : 330,000원 (완전 반품)\n"
            "- 응답 : 「반품 등록 완료」 — PT_USE 6,000 복원\n"
            "- 시나리오 단계 : 7-3) SAL04055 완전 반품 완료"
        ),
    ),
    # 7-5) 1차 판매 부분 반품
    "a12JO000000MgqGYAS": (
        "7-5) ★ ✓",
        "1차 판매 반품 진행 — SAL04121 (부분 반품 성공 PASS)",
        (
            "반품 번호 : SAL04121  /  원거래 판매 번호 : SAL04052\n"
            "- 시간 : 2026-05-10 21:29:27 KST\n"
            "- 실결제 금액 : 2,000,000원\n"
            "- 응답 : 「반품 등록 완료」 — PT_USE 40,000 복원 (1차 적립 전체 환수)\n"
            "- 시나리오 단계 : 7-5) 1차 판매 부분 반품 진행 ✓ PASS\n"
            "- 검증 : 정리 대상 (SAL04053/54/55) 사전 처리 후 1차 판매 반품 정상 동작"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 50, 95, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
