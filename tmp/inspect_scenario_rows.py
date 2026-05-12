import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "tmp" / "apex_audit_trail_export_20260508.xlsx"

TARGETS = {
    "SOR202605080040",
    "SOR202605080042",
    "SAL202605080046",
    "SAL202605080053",
    "SAL202605080055",
    "SAL202605080056",
}

KEY_FIELDS = [
    "SaleNo__c", "OrderNo__c", "F_OrderNo__c", "F_SaleNo__c",
    "OriginalOrderId", "OriginalOrderId__c", "originalOrderId", "orderId",
    "ActualPaymentAmount__c", "TotalAmount", "MemberNo__c", "F_MemberNo__c",
    "OrderStatus__c", "SaleStatus__c", "DeleteDateTime__c", "IsDeleted__c",
    "IsReductionOrder__c", "IsReductionOrder",
    "code", "success", "message", "reason",
    "point_processing", "PointCreditList", "PointDebitList",
    "TX_REMARK", "PT_TARGET", "CD_TYPE_SAVE", "CD_TYPE_POINT",
    "DepositNo__c", "DepositType__c", "TransactionAmount", "PaymentMethod",
]


def short(v, n=180):
    s = json.dumps(v, ensure_ascii=False) if not isinstance(v, str) else v
    s = s.replace("\n", " ")
    return s if len(s) <= n else s[:n] + "..."


def extract_keys(text):
    if not text:
        return {}
    out = {}
    try:
        obj = json.loads(text)
    except Exception:
        return {"_raw_excerpt": short(text, 200)}

    def walk(node, path=""):
        if isinstance(node, dict):
            for k, v in node.items():
                if k in KEY_FIELDS:
                    if k not in out:
                        out[k] = v
                if isinstance(v, (dict, list)):
                    walk(v, path + "." + k)
        elif isinstance(node, list):
            for i, item in enumerate(node[:3]):
                walk(item, path + f"[{i}]")

    walk(obj)
    return out


def main():
    wb = load_workbook(SRC, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        matched_text = row[idx["MatchedNumbers"]] or ""
        matched_set = {m.strip() for m in matched_text.split(",") if m.strip()}
        if not (matched_set & TARGETS):
            continue
        rows.append({
            "MatchedNumbers": matched_text,
            "Id": row[idx["Id"]],
            "CreatedDate": row[idx["CreatedDate"]],
            "ApexClass__c": row[idx["ApexClass__c"]],
            "RequestUrl__c": row[idx["RequestUrl__c"]],
            "RequestBody__c": row[idx["RequestBody__c"]] or "",
            "ResponseBody__c": row[idx["ResponseBody__c"]] or "",
        })

    rows.sort(key=lambda r: (str(r["CreatedDate"] or ""), str(r["Id"] or "")))

    for i, r in enumerate(rows, start=1):
        print("=" * 100)
        print(f"[{i:>2}] CreatedDate = {r['CreatedDate']}")
        print(f"     Id           = {r['Id']}")
        print(f"     ApexClass    = {r['ApexClass__c']}")
        print(f"     RequestUrl   = {r['RequestUrl__c']}")
        print(f"     Matched      = {r['MatchedNumbers']}")
        rk = extract_keys(r["RequestBody__c"])
        print(f"     RequestKeys  : {rk}")
        sk = extract_keys(r["ResponseBody__c"])
        print(f"     ResponseKeys : {sk}")


if __name__ == "__main__":
    main()
