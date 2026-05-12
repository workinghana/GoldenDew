"""V5-1 Log.xlsx — 회원 TEST03006 / [TEST] 통합시나리오V5_구매2배 (PRO202605041079).

원판매(주문)에 포인트와 쿠폰을 사용했는데, 최종 통 반품 시점에 포인트와 쿠폰을
100% 입력하지 않고 저장하는 경우.
"""
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V5-1 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v5_1_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V5-1"
MEMBER_NO = "TEST03006"
SHEET_NAME = "V5-1"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower().replace("_", "")
    if "memberhelpinfo" in ac:
        return "회원 도움창"
    if "memberapi" in ac:
        return "회원"
    if "voucherhelp" in ac:
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac:
        return "포인트 (적용 조회)"
    if "saledeposit" in ac:
        return "판매 입금"
    if "returndeposit" in ac:
        return "반품 입금"
    if "orderapi" in ac:
        return "주문"
    if "saleapi" in ac:
        return "판매"
    if "returnapi" in ac:
        return "반품"
    if "pointcredit" in ac:
        return "포인트 (지급)"
    if "pointdebit" in ac:
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    # 사전 셋팅 (1)~(3) : 주문 SOR04011 등록
    "a12JO000000Me24YAC": (
        "사전 (1)~(3)",
        "주문 등록 — SOR04011 (포인트 10,000 + 정액쿠폰 30,000 + 선수금 카드 10%)",
        (
            "주문 번호 : SOR04011  (시간 : 2026-05-10 18:02:52 KST)\n"
            "- 회원 : TEST03006\n"
            "- 매장 코드 : 99998 (백화점)\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 프로모션 : PRO202605041079 ([TEST] 통합시나리오V5_구매2배)\n"
            "- 결제 구성 : 포인트 10,000원 + 정액쿠폰 30,000원 + 선수금 카드 10%\n"
            "- 응답 : 「주문 등록 완료」 — PT_USE 10,000\n"
            "- 시나리오 단계 : 사전 셋팅 (1)~(3) (주문에서 포인트/쿠폰/선수금 입력)"
        ),
    ),
    # 사전 셋팅 (3) : 주문을 끌어와 판매처리하여 원판매번호 생성
    "a12JO000000Me3kYAC": (
        "사전 (3)",
        "판매 처리 — SAL04044 (원판매번호 생성, 구매포인트 2배 적용)",
        (
            "판매 번호 : SAL04044  (시간 : 2026-05-10 18:07:25 KST)\n"
            "- 원거래 주문 번호 : SOR04011\n"
            "- 매장 코드 : 99998\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 프로모션 : PRO202605041079 (구매포인트 2배)\n"
            "- 응답 : 「판매 등록 완료」 — 구매포인트 9,600P 적립 (= 4,800 × 2배)\n"
            "- 시나리오 단계 : 사전 셋팅 (3) — 주문 → 판매 처리하여 원판매번호 생성"
        ),
    ),
    # 사전 셋팅 (4) : 에코포인트 6,000원 지급
    "a12JO000000McmcYAC": (
        "사전 (4)",
        "에코포인트 6,000원 지급 (SAL04044 대상)",
        (
            "포인트 지급 — 이벤트포인트 (에코포인트) 6,000원\n"
            "- 시간 : 2026-05-10 18:09:23 KST\n"
            "- 회원 : TEST03006\n"
            "- 대상 판매 번호 : SAL04044\n"
            "- 포인트 유형 : 05 (이벤트포인트 / 에코포인트)\n"
            "- 응답 : 「포인트 지급 성공」 — PT_TARGET 6,000\n"
            "- 시나리오 단계 : 사전 셋팅 (4) — 에코포인트 6,000원 지급"
        ),
    ),
    # 진행 (1) 1차 시도 — ERROR (DUPLICATE_VALUE)
    "a12JO000000MeF0YAK": (
        "진행 (1) 1차",
        "부분 반품 1차 시도 — SAL04116 (포인트 5,000 입력) ⚠️ ERROR",
        (
            "반품 번호 : SAL04116  /  원거래 판매 번호 : SAL04044\n"
            "- 시간 : 2026-05-10 18:14:03 KST\n"
            "- 실결제 금액 : 500,000원 (부분반품 1개 품목)\n"
            "- 결제 구성 : 포인트만 5,000원 입력\n"
            "- 응답 : ❌ ERROR — code 500\n"
            "    Insert failed. DUPLICATE_VALUE (DepositNo__c 중복: 0lVJO000000CGx32AG)\n"
            "- 시나리오 단계 : 진행 (1) 1차 시도 (DepositNo 중복 에러로 실패)"
        ),
    ),
    # 진행 (1) 재시도 — 성공
    "a12JO000000MdR2YAK": (
        "진행 (1)",
        "부분 반품 재시도 — SAL04116 (포인트 5,000 입력) ✓ PASS",
        (
            "반품 번호 : SAL04116  /  원거래 판매 번호 : SAL04044\n"
            "- 시간 : 2026-05-10 18:14:13 KST  (1차 시도 후 10초)\n"
            "- 실결제 금액 : 500,000원 (부분반품 1개 품목)\n"
            "- 결제 구성 : 포인트만 5,000원 입력\n"
            "- 응답 : 「반품 등록 완료」 — 적립포인트 4,950P 복원\n"
            "- 시나리오 단계 : 진행 (1) 부분 반품 성공"
        ),
    ),
    # 진행 (2-3) 통반품 — PASS
    "a12JO000000Mdh8YAC": (
        "진행 (2-3)",
        "통 반품 — SAL04117 (포인트 5,000 + 쿠폰 30,000 입력) ✓ PASS",
        (
            "반품 번호 : SAL04117  /  원거래 판매 번호 : SAL04044\n"
            "- 시간 : 2026-05-10 18:18:45 KST\n"
            "- 실결제 금액 : 500,000원 (나머지 1개 품목 통반품)\n"
            "- 결제 구성 : 포인트 5,000 + 쿠폰 30,000 입력 후 저장 (2-3 케이스)\n"
            "- 응답 : 「반품 등록 완료」 — 적립포인트 4,650P 복원\n"
            "- 시나리오 단계 : 진행 (2-3) — 포인트 5,000 / 쿠폰 30,000 입력 후 저장 (PASS)\n"
            "- 참고 : (2-1) 저장 된 경우 Error / (2-2) 포인트 X 쿠폰 30,000 Error 는\n"
            "  로컬티 테스트 환경에서 금액 부족 시 저장 막힘 미적용으로 진행 X"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 12, 50, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
