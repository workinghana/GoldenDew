import os
from pathlib import Path

from openpyxl import load_workbook


DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
TARGET = DOWNLOADS / "scenario_audit_trail_20260508.xlsx"

wb = load_workbook(TARGET, read_only=True)
print(f"파일: {TARGET}")
print(f"시트 목록: {wb.sheetnames}")
print()

for name in wb.sheetnames:
    ws = wb[name]
    n = ws.max_row - 1  # exclude header
    print(f"  · {name}  → {n}행")
print()

if "v4-2_7차" not in wb.sheetnames:
    print("[ERROR] v4-2_7차 시트가 없습니다!")
    raise SystemExit(1)

ws = wb["v4-2_7차"]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print("=" * 100)
print(f"시트 [v4-2_7차] 헤더: {headers}")
print("=" * 100)
print()

idx = {h: i for i, h in enumerate(headers)}

for row in ws.iter_rows(min_row=2, values_only=True):
    no = row[idx["번호"]]
    label = row[idx["확인내용"]]
    member_no = row[idx["회원번호"]]
    member_name = row[idx["회원명"]]
    log_id = row[idx["로그 ID"]]
    apex = row[idx["ApexClass__c"]]
    url = row[idx["URL"]]
    print(f"[{no}] {label}")
    print(f"     회원: {member_no} ({member_name})")
    print(f"     로그: {log_id}  /  {apex}  /  {url[:60] if url else ''}")
    crit = row[idx["확인 기준 (정답지 - 오류 내용 제외)"]] or ""
    first_line = crit.split('\n')[0]
    print(f"     기준 첫줄: {first_line[:120]}")
    print()
