"""V5-4 Log.xlsx — 시나리오 V5-4 (프로모션 비활성화/만료 → 판매처리 → 판매삭제 적립내역 미삭제 이슈).

「4. 주문에서 입력한 프로모션 번호를 비활성화(삭제) 처리 → 판매처리 테스트」
- 메인 시나리오 회원: 173000400470 / SOR064·SOR065 (audit 미발견 — Salesforce 직접 생성 추정)
- (05.08) 적립내역 미삭제 이슈 검증 회원: 043004800023 / SAL202605080266 (audit 매칭)
"""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V5-4 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "_v5_4_raw.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V5-4"
SHEET_NAME = "V5-4"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac.replace("_", ""):
        return "포인트 (적용 조회)"
    if "saledeposit" in ac.replace("_", ""):
        return "판매 입금"
    if "returndeposit" in ac.replace("_", ""):
        return "반품 입금"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용/차감)"
    if "promotion" in ac.replace("_", ""):
        return "프로모션 동기화"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        # SaleNo only URL (no /gd/v1/...) → DELETE
        if '"SaleNo__c"' in url and '"DepositNo__c"' not in url:
            return "DELETE"
        return "DELETE"
    return "POST"


# (Salesforce log Id) -> (번호, 회원번호, 확인내용, 확인 기준)
SCENARIO = {
    # (05.08-A) Order SOR202605110006 — 프로모션 기간 종료 시점 주문 등록
    "a12JO000000MEpDYAW": (
        "(05.08)-A",
        "043004800023",
        "주문 등록 (프로모션 기간 종료 후 검증 주문) — SOR202605110006",
        (
            "주문 번호 : SOR202605110006  (시간 : 2026-05-08 16:42:51 KST)\n"
            "- 회원 번호 : 043004800023\n"
            "- 매장 코드 : 99998 (백화점)  /  유형 : 01 (계약)\n"
            "- 프로모션 번호 : PRO202605041081\n"
            "- 실결제 금액 : 500,000원 (2A2600001 × 1, 순매가 단가 500,000)\n"
            "- 거래 원장 (입금 번호 DEP202605080295) :\n"
            "    1) 결제 수단 01 (현금)  /  입금 구분 01 (계약금)  /  거래 금액 50,000원\n"
            "    2) 결제 수단 81 (포인트)  /  입금 구분 01  /  거래 금액 10,000원\n"
            "- 응답 메시지 : 「주문 등록 완료」  /  success: true\n"
            "- 검증 : 주문 정상 등록 (프로모션 기간 종료 시점에서 사은포인트만 적립되는지 후속 확인 대상)"
        ),
    ),
    # (05.08-B) Sale SAL202605080266 — 판매 처리
    "a12JO000000MJ5gYAG": (
        "(05.08)-B",
        "043004800023",
        "판매 처리 — SAL202605080266 (적립내역 행 검증 대상)",
        (
            "판매 번호 : SAL202605080266  /  원거래 주문 번호 : SOR202605110006\n"
            "- 시간 : 2026-05-08 16:49:42 KST\n"
            "- 회원 번호 : 043004800023\n"
            "- 매장 코드 : 99998\n"
            "- 실결제 금액 : 500,000원\n"
            "- 프로모션 번호 : PRO202605041081\n"
            "- 거래 원장 (계약금 + 잔금 분개) — 응답 본문 확인 필요\n"
            "- 응답 적립 포인트 : 시나리오상 「프로모션 만료되더라도 구매포인트 적립 안되고, 사은포인트만 적립됨」 (확인 완료)\n"
            "- 시나리오 검증 : 진행 (2-2) 「구매포인트 적립 안됨 - PASS」 ✓"
        ),
    ),
    # (05.08-C) Sale SAL202605080266 DELETE — 적립내역 행 사라지지 않는 이슈 검증
    "a12JO000000ME1BYAW": (
        "(05.08)-C",
        "043004800023",
        "판매 삭제 — SAL202605080266 DELETE / 적립내역 행 미삭제 이슈 검증",
        (
            "Apex 클래스 : GdSaleApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-08 16:50:20 KST  (B 판매 후 38초)\n"
            "- 대상 판매 번호 : SAL202605080266\n"
            "- URL : {\"SaleNo__c\":\"SAL202605080266\"} (RequestBody 없음)\n"
            "- 응답 메시지 : 응답 본문 별도 확인\n"
            "- 시나리오 이슈 (05.08) :\n"
            "    · 「프로모션 기간 종료 후, 판매 삭제 처리시 포인트 합은 맞으나 회원정보도움창의 적립내역의 행이 사라지지 않음」\n"
            "    · 즉 audit 상으로는 SaleDelete 가 정상 호출되었으나 회원정보도움창 적립내역 표시는 미반영\n"
            "    · → 화면(회원정보도움창) 측 데이터 동기화 로직 점검 필요\n"
            "- ※ 이 audit 가 시나리오 표기 「(05.08) 판매 삭제 처리시 ...」 의 직접 검증 단서"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Add note row at top
    note_text = (
        "※ 시나리오 V5-4 메인 (회원 173000400470 / SOR202605060064 진행(1) / SOR202605060065 진행(2)) 의 audit 는 "
        "본 캐시(2026-05-06 ~ 2026-05-08) 에서 발견되지 않음 — 5/6 이전 또는 Salesforce 직접 생성된 것으로 추정. "
        "본 파일은 (05.08) 판매 삭제 시 적립내역 행 미삭제 이슈 검증 audit (회원 043004800023) 만 수록."
    )

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    note_fill = PatternFill("solid", fgColor="FFF2CC")

    # Row 1 : note (merged across)
    ws.cell(row=1, column=1, value=note_text)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out_headers))
    ws.cell(row=1, column=1).fill = note_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=1, column=1).font = Font(color="9C5700", bold=False, italic=True)
    ws.row_dimensions[1].height = 50

    # Row 2 : header
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, member_no, label, criterion)) in enumerate(sorted_items, start=3):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            member_no,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 40, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C3"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
