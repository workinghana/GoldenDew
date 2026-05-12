import csv
import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
TARGET_FILE = DOWNLOADS / "scenario_audit_trail_20260508.xlsx"
MATCHED_CSV = ROOT / "tmp" / "scenario2_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "v4-2_7차"
MEMBER_NO = "103006200124"
MEMBER_NAME = "한상희 VVIP"
SHEET_NAME = "v4-2_7차"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용)"
    if "erpasync" in ac.replace("_", ""):
        return "ERP 비동기"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    return "POST"


# 시나리오 2 매핑
# Id 가 "_NO_LOG_" 로 시작하면 audit trail 호출이 없는 단계 (UI/내부 처리)
SCENARIO_2 = {
    # ========== 4-2-1: 백화점 쿠폰 금액 변경(10만원→7만원) + 판매 + 반품 → 쿠폰 복원 검증
    "a12JO000000M3QjYAK": (
        "4-2-1-1",
        "백화점 주문 등록 (VVIP 10만원 생일쿠폰 → 7만원 수정 저장)",
        (
            "주문 : SOR202605080077  (Salesforce Id : 801JO00000I9rseYAB)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 입금 금액 : 50,000원  (10% / PaymentMethod 03 · DepositType 01 · DEP202605080010-1)\n"
            "- audit trail 본 요청 CouponNo : '' (빈값)\n"
            "- 시나리오 동작 : VVIP 10만원 생일쿠폰 입금 등록 → ERP UI 에서 7만원으로 금액 수정 후 저장\n"
            "  ※ 쿠폰 금액 수정은 별도 ERP 작업으로 본 audit trail 단일 호출엔 직접 노출되지 않음 (TJ 추적 필요)"
        ),
    ),
    "a12JO000000M7HQYA0": (
        "4-2-1-2",
        "해당 주문 건 판매 처리 (A 판매)",
        (
            "판매 : SAL202605080058  /  원주문 : SOR202605080077\n"
            "(Salesforce Id : 801JO00000I9wXFYAZ)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 판매 입금(주문 입금 차감) : 50,000원  (DepositType 01 · PaymentMethod 03 · DEP202605080010-1)\n"
            "- 적립포인트(구매포인트, VVIP 2.0배율) : 8,600P  (PT_TARGET / API 응답 기준)\n"
            "- TX_REMARK : 구매포인트 / CD_TYPE_POINT : 01\n"
            "- SaleStatus : 01 (정상 판매)"
        ),
    ),
    "a12JO000000M7R6YAK": (
        "4-2-1-3",
        "판매 반품 등록 (쿠폰 복원 검증)",
        (
            "반품 거래 : SAL202605080059  /  원거래 : SAL202605080058 (A 판매)\n"
            "(Salesforce Id : 801JO00000IA1dBYAT)\n"
            "- Quantity : -1 (반품)\n"
            "- 환불 금액 : 70,000원  (판매 입금 취소 / DepositType 03 · PaymentMethod 90 (쿠폰) · DEP202605080012-1)\n"
            "  ※ 사용된 쿠폰 7만원 환불 (Sale Deposit Reversal — 쿠폰)\n"
            "- 환불 쿠폰 : COP2026MY0012;-;9S8WTK5N  (VVIP 생일쿠폰, 10만원 → 7만원 수정 후 사용분)\n"
            "- 시나리오 검증 : 「반품 등록 시 VVIP 생일 쿠폰 7만원 입금처리된 것 환불, 쿠폰 복원은 VVIP 생일쿠폰 10만원으로 복원되는지 확인」\n"
            "  · 환불 7만원 ✓\n"
            "  · 쿠폰 10만원 복원은 별도 Voucher 객체 검증 필요\n"
            "- originalOrderId(internal) : 801JO00000I9wXFYAZ → orderId 801JO00000IA1dBYAT"
        ),
    ),
    # ========== 4-2-2: 매장 재고 판매(쿠폰 금액 수정) + 판매 삭제 → 쿠폰 복원 검증
    "a12JO000000M5QvYAK": (
        "4-2-2-1",
        "백화점 매장 재고 판매 (VVIP 10만원 → 7만원 수정 쿠폰 사용)",
        (
            "판매 : SAL202605080060  (매장 재고 판매 — 원주문 없음)\n"
            "(Salesforce Id : 801JO00000I9r7nYAB)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 판매 입금(쿠폰) : 70,000원  (DepositType 02 · PaymentMethod 90 (쿠폰) · DEP202605080013-1)\n"
            "- 사용 쿠폰 : COP2026MY0012;-;9S8WTK5N  (VVIP 10만원 생일쿠폰 → 7만원으로 수정 후 사용)\n"
            "- 적립포인트(구매포인트, VVIP 2.0배율) : 8,600P  (PT_TARGET)\n"
            "- TX_REMARK : 구매포인트 / CD_TYPE_POINT : 01\n"
            "- 시나리오 동작 : VVIP 회원 생일 10만원 입금 등록 → 쿠폰 금액 7만원으로 수정 저장"
        ),
    ),
    "a12JO000000M7kSYAS": (
        "4-2-2-2",
        "판매 삭제 (월마감 전 — 반품 처리 흐름으로 기록)",
        (
            "판매 삭제(반품) : SAL202605080061  /  원거래 : SAL202605080060\n"
            "(Salesforce Id : 801JO00000IA0cKYAT)\n"
            "- Quantity : -1\n"
            "- 환불 금액 : 70,000원  (판매 입금 취소 / DepositType 03 · PaymentMethod 90 (쿠폰) · DEP202605080014-1)\n"
            "- 환불 쿠폰 : COP2026MY0012;-;9S8WTK5N\n"
            "- 시나리오 검증 : 「판매 삭제 시에도 VVIP 생일 쿠폰 7만원 환불, 쿠폰 복원은 VVIP 생일쿠폰 10만원으로 복원되는지 확인」\n"
            "  · 환불 7만원 ✓\n"
            "  · 쿠폰 10만원 복원은 Voucher 객체 검증 필요\n"
            "- ※ 본 환경에서는 판매 삭제도 GdReturnApiControllerV1 흐름으로 처리됨 (응답 message: 「반품 등록 완료」)"
        ),
    ),
    # ========== 4-2-3: 아울렛 수선 (15만원 → 10만원 수정) + 수선 판매 + 반품 시도 차단
    "a12JO000000M5NmYAK": (
        "4-2-3-1",
        "아울렛 수선 주문 등록 (SVIP 수선쿠폰 15만원 → 10만원 수정 저장)",
        (
            "수선 주문 : SOR202605080078  (Salesforce Id : 801JO00000IA09RYAT)\n"
            "- 실결제 금액 : 70,000원 (수선료)\n"
            "- 입금 금액 : 7,000원 (10% / PaymentMethod 01 · DepositType 01 · DEP202605080016-1)\n"
            "- audit trail 본 요청 CouponNo : '' (빈값)\n"
            "- 시나리오 동작 : SVIP 수선 쿠폰 15만원 입금 적용 → 10만원으로 수정 저장 (ERP UI 작업)\n"
            "  ※ 쿠폰 금액 수정은 별도 ERP 작업"
        ),
    ),
    "a12JO000000M85QYAS": (
        "4-2-3-2",
        "수선 판매 처리 (아울렛 — 적립포인트 없음 검증)",
        (
            "수선 판매 : SAL202605080064  /  원주문 : SOR202605080078\n"
            "(Salesforce Id : 801JO00000I9pPGYAZ)\n"
            "- 실결제 금액 : 70,000원\n"
            "- 판매 입금 : 7,000원  (DepositType 01 · PaymentMethod 01 · DEP202605080016-1)\n"
            "- 적립포인트 : 없음 (PT_TARGET 미존재)\n"
            "- 시나리오 검증 : 「아울렛은 적립포인트 없음」 ✓ (응답에 PT_TARGET·TX_REMARK·CD_TYPE_POINT 모두 미존재 확인)"
        ),
    ),
    "_NO_LOG_4-2-3-3": (
        "4-2-3-3",
        "수선 판매 반품 시도 → 「수선 판매내역은 반품처리 할 수 없습니다」 메시지 노출",
        (
            "시나리오 검증 : 「수선 판매내역은 반품처리 할 수 없습니다」 메시지 노출 (확인완료 — IMC팀·전산팀 확인)\n"
            "- 수선 판매내역은 일반적으로 반품 등록 불가\n"
            "- UI 단계에서 차단되어 audit trail 호출이 발생하지 않음 → 본 단계 로그 없음\n"
            "- 시나리오 의도 : 수선 쿠폰 10만원 환불 + SVIP 수선 쿠폰 15만원 복원 검증 — 본 환경에서는 차단되어 진행되지 않음"
        ),
    ),
    # ========== 4-2-4: 백화점 SVIP 부분반품(10만원→5만원 수정) + 쿠폰 복원
    "a12JO000000M8OmYAK": (
        "4-2-4-1",
        "백화점 주문 등록 (SVIP 10만원 생일쿠폰 → 5만원 수정 저장)",
        (
            "주문 : SOR202605080081  (Salesforce Id : 801JO00000I9seuYAB)\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 입금 금액 : 100,000원  (10% / PaymentMethod 01 · DepositType 01 · DEP202605080023-1)\n"
            "- audit trail 본 요청 CouponNo : '' (빈값)\n"
            "- 시나리오 동작 : SVIP 10만원 생일쿠폰 입금 등록 → 5만원으로 수정 저장 (ERP UI 작업)"
        ),
    ),
    "a12JO000000M3QnYAK": (
        "4-2-4-2",
        "해당 주문 건 판매 처리",
        (
            "판매 : SAL202605080068  /  원주문 : SOR202605080081\n"
            "(Salesforce Id : 801JO00000I9rsoYAB)\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 판매 입금(주문 입금 차감) : 100,000원  (DepositType 01 · PaymentMethod 01 · DEP202605080023-1)\n"
            "- 적립포인트(구매포인트, VVIP 2.0배율) : 19,800P  (PT_TARGET)\n"
            "- TX_REMARK : 구매포인트 / CD_TYPE_POINT : 01\n"
            "- SaleStatus : 01 (정상 판매)"
        ),
    ),
    "a12JO000000M6YJYA0": (
        "4-2-4-3",
        "부분 반품 등록 (SVIP 5만원 변경 사용분 환불)",
        (
            "부분 반품 거래 : SAL202605080070  /  원거래 : SAL202605080068\n"
            "(Salesforce Id : 801JO00000I9yimYAB)\n"
            "- Quantity : -1 (부분 반품)\n"
            "- 부분 반품 ActualPaymentAmount__c : 500,000원\n"
            "- 환불 금액(쿠폰) : 10,000원  (DepositType 03 · PaymentMethod 90 (쿠폰) · DEP202605080027-1)\n"
            "  ※ 본 audit trail 호출에 노출된 환불 금액 — 부분 반품 비율/순서에 따른 처리 결과\n"
            "- 환불 쿠폰 : COP2026JA0029;-;169  (SVIP 생일쿠폰)\n"
            "- 시나리오 검증 :\n"
            "    1) 사용 쿠폰 환불 「불러오기」 버튼 클릭 시 → SVIP 생일 쿠폰 5만원으로 변경된 금액으로 정확히 호출 (확인완료)\n"
            "    2) 「쿠폰반품등록」 버튼 클릭하여 해당 쿠폰 다시 등록 시 마이너스 금액으로 끌고와지지 않고 0으로 끌고와짐\n"
            "       + 현금/쿠폰 순서를 달리해도 연산되지 않음"
        ),
    ),
    "_NO_LOG_4-2-4-4": (
        "4-2-4-4",
        "부분 반품 완료 후 SVIP 생일쿠폰 10만원 복원 검증",
        (
            "시나리오 검증 : 「부분 반품 등록 완료 후 쿠폰 복원이 SVIP 생일쿠폰 10만원으로 복원되는지 확인」\n"
            "- 별도 audit trail 호출 없음 — 쿠폰 복원은 반품 처리 시 자동 수행 → 본 단계 로그 없음\n"
            "- TJ(TransactionJournal) 또는 Loyalty Voucher 객체에서 복원 결과 직접 검증 필요"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def load_matched_rows():
    """matched csv에서 SCENARIO_2 의 Id 들에 해당하는 행만 dict로 매핑."""
    by_id = {}
    if not MATCHED_CSV.exists():
        raise FileNotFoundError(f"matched csv not found: {MATCHED_CSV}")
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            by_id[r["Id"]] = r
    return by_id


def main():
    by_id = load_matched_rows()

    if not TARGET_FILE.exists():
        raise FileNotFoundError(f"target file not found: {TARGET_FILE}")
    wb = load_workbook(TARGET_FILE)

    # 기존 동일 시트 있으면 제거
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(title=SHEET_NAME)

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "회원명",
        "도메인",
        "URL",
        "METHOD",
        "ApexClass__c",
        "로그 ID",
        "CreatedDate",
        "RequestBody__c",
        "ResponseBody__c",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    no_log_fill = PatternFill("solid", fgColor="FFF2CC")  # 연노랑 — 로그 없는 단계 표시
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    # 정렬: 번호 기준
    def num_sort_key(item):
        no = item[1][0]
        # "4-2-1-1" -> (4,2,1,1)
        parts = no.split("-")
        try:
            return tuple(int(p) for p in parts)
        except ValueError:
            return (99, 99, 99, 99)

    sorted_items = sorted(SCENARIO_2.items(), key=num_sort_key)

    matched_count = 0
    no_log_count = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        if sf_id.startswith("_NO_LOG_"):
            url = ""
            method = ""
            apex = ""
            log_id = "(로그 없음)"
            created = ""
            req_body = ""
            res_body = ""
            domain = "(audit trail 호출 없음)"
            no_log_count += 1
        else:
            rec = by_id.get(sf_id)
            if not rec:
                print(f"[warn] {no}: log Id {sf_id} not in matched csv")
                continue
            matched_count += 1
            url = rec.get("RequestUrl__c") or ""
            apex = rec.get("ApexClass__c") or ""
            method = http_method(apex, url)
            domain = classify_domain(apex, url)
            log_id = rec.get("Id") or ""
            created = rec.get("CreatedDate") or ""
            req_body = rec.get("RequestBody__c") or ""
            res_body = rec.get("ResponseBody__c") or ""

        values = [
            SCENARIO_VERSION,
            no,
            label,
            criterion,
            MEMBER_NO,
            MEMBER_NAME,
            domain,
            trim(url),
            method,
            trim(apex),
            trim(log_id),
            trim(created),
            trim(req_body),
            trim(res_body),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align
            if sf_id.startswith("_NO_LOG_"):
                cell.fill = no_log_fill

    widths = [12, 10, 40, 80, 16, 14, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(TARGET_FILE)
    print(f"[saved] {TARGET_FILE}")
    print(f"[summary] matched_with_log={matched_count}  no_log_steps={no_log_count}  total={len(SCENARIO_2)}")


if __name__ == "__main__":
    main()
