"""V5-8 Log.xlsx — V5 시리즈 7-1 ~ 7-5 시나리오 (회원별 다양).

7-1 (밀버스1 / 269999800188) — 로그 미확인 (캐시 범위 외)
7-2 (밀버스2 / 269999800189) — 덩어리 포인트 20000P, 부분/나머지 반품 (에코 차감 X 확인)
7-3 (밀버스7 / 269999800194) — 덩어리 200000P, 1차+2차 판매, 부분/나머지 반품
7-4 — 반품 삭제 로직 수정 대기 (미실행)
7-5 (밀버스5 / 269999800192) — 로그 미확인 (캐시 범위 외)
"""
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V5-8 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v5_8_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V5-8"
SHEET_NAME = "V5-8"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower().replace("_", "")
    if "memberhelpinfo" in ac:
        return "회원 도움창"
    if "memberapi" in ac:
        return "회원"
    if "voucherhelp" in ac:
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac:
        return "포인트 (적용 조회)"
    if "saledeposit" in ac:
        return "판매 입금"
    if "returndeposit" in ac:
        return "반품 입금"
    if "orderapi" in ac:
        return "주문"
    if "saleapi" in ac:
        return "판매"
    if "returnapi" in ac:
        return "반품"
    if "pointcredit" in ac:
        return "포인트 (지급)"
    if "pointdebit" in ac:
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


# (log_id, member_no, scenario_no, label, criterion)
SCENARIO = {
    # 7-2 (밀버스2 / 269999800189)
    "a12JO000000MjhgYAC": (
        "269999800189",
        "7-2 (밀버스2)",
        "SAL04064 판매 등록 (PT_TARGET 25,000P)",
        (
            "판매 번호 : SAL04064  (시간 : 2026-05-11 01:00:47 KST)\n"
            "- 회원 : 269999800189 (밀버스2)\n"
            "- 실결제 금액 : 2,500,000원\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 25,000P\n"
            "- 시나리오 7-2 : 덩어리 포인트 20,000P / 부분반품 → 5,000P 차감"
        ),
    ),
    "a12JO000000Mj09YAC": (
        "269999800189",
        "7-2 (밀버스2)",
        "판매 입금 삭제 — DELETE DEP040042 (SAL04064)",
        (
            "- 시간 : 2026-05-11 01:05:40 KST\n"
            "- 응답 : 「판매 입금 삭제 삭제 완료」 — PT_USE 25,000 복원\n"
            "- 시나리오 7-2 : 입금 삭제"
        ),
    ),
    "a12JO000000MjpkYAC": (
        "269999800189",
        "7-2 (밀버스2)",
        "SAL04065 판매 등록 (재판매, PT_TARGET 2,500P)",
        (
            "판매 번호 : SAL04065  (시간 : 2026-05-11 01:07:55 KST)\n"
            "- 회원 : 269999800189 (밀버스2)\n"
            "- 실결제 금액 : 250,000원\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 2,500P\n"
            "- 시나리오 7-2 : 재판매"
        ),
    ),
    "a12JO000000MkAgYAK": (
        "269999800189",
        "7-2 (밀버스2)",
        "SAL04066 판매 등록 1차 시도 ❌ ERROR DUPLICATE_VALUE",
        (
            "판매 번호 : SAL04066  (시간 : 2026-05-11 01:10:38 KST)\n"
            "- 응답 : ❌ code 500 — DUPLICATE_VALUE (DepositNo__c 중복)\n"
            "- 시나리오 7-2 : 1차 시도 실패"
        ),
    ),
    "a12JO000000MkCKYA0": (
        "269999800189",
        "7-2 (밀버스2)",
        "SAL04066 판매 등록 재시도 (PT_TARGET 2,500P)",
        (
            "판매 번호 : SAL04066  (시간 : 2026-05-11 01:11:01 KST)\n"
            "- 실결제 금액 : 270,000원\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 2,500P\n"
            "- 시나리오 7-2 : 재시도 성공"
        ),
    ),
    "a12JO000000MkIjYAK": (
        "269999800189",
        "7-2 (밀버스2)",
        "판매 입금 삭제 — DELETE DEP040044 (SAL04066)",
        (
            "- 시간 : 2026-05-11 01:17:36 KST\n"
            "- 응답 : 「판매 입금 삭제 삭제 완료」 — PT_USE 2,500 복원\n"
            "- 시나리오 7-2 : 입금 삭제 (반품 사전 정리)\n"
            "- ※ 이미지 메모상 부분반품 / 나머지 반품 / 에코포인트 차감 안됨 확인 단계 로그는\n"
            "  본 캐시 시간 범위에서 미확인 — 별도 시간대 데이터 필요"
        ),
    ),
    # 7-3 (밀버스7 / 269999800194)
    "a12JO000000Mk13YAC": (
        "269999800194",
        "7-3 (밀버스7)",
        "SAL04067 1차 판매 등록 (PT_TARGET 40,000P 적립)",
        (
            "판매 번호 : SAL04067  (시간 : 2026-05-11 01:28:15 KST)\n"
            "- 회원 : 269999800194 (밀버스7)\n"
            "- 실결제 금액 : 2,000,000원\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 40,000P (구매포인트 2배)\n"
            "- 시나리오 7-3 : 1차 판매 진행"
        ),
    ),
    "a12JO000000MkYuYAK": (
        "269999800194",
        "7-3 (밀버스7)",
        "SAL04068 2차 판매 등록 (PT_TARGET 40,000P)",
        (
            "판매 번호 : SAL04068  (시간 : 2026-05-11 01:31:05 KST)\n"
            "- 실결제 금액 : 2,200,000원\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 40,000P\n"
            "- 시나리오 7-3 : 2차 판매 진행 (포인트 20만원 사용)"
        ),
    ),
    "a12JO000000MkaTYAS": (
        "269999800194",
        "7-3 (밀버스7)",
        "에코포인트 6,000P 지급 (SAL04068)",
        (
            "- 시간 : 2026-05-11 01:31:35 KST\n"
            "- 응답 : 「포인트 지급 성공」 — 이벤트포인트 6,000P\n"
            "- 시나리오 7-3 : 에코 6,000P 적립 (총 보유 포인트 106,000P)"
        ),
    ),
    "a12JO000000Mkc8YAC": (
        "269999800194",
        "7-3 (밀버스7)",
        "SAL04129 2차 판매 부분 반품 (PT_USE 20,000 복원)",
        (
            "반품 번호 : SAL04129  /  원거래 판매 번호 : SAL04068\n"
            "- 시간 : 2026-05-11 01:37:24 KST\n"
            "- 실결제 금액 : 1,100,000원\n"
            "- 응답 : 「반품 등록 완료」 — PT_USE 20,000 복원\n"
            "- 시나리오 7-3 : 2차 판매 부분 반품 (10만 포인트 사용 취소 일부)"
        ),
    ),
    "a12JO000000MkdkYAC": (
        "269999800194",
        "7-3 (밀버스7)",
        "SAL04130 2차 판매 나머지 반품 (PT_USE 20,000 복원)",
        (
            "반품 번호 : SAL04130  /  원거래 판매 번호 : SAL04068\n"
            "- 시간 : 2026-05-11 01:41:26 KST\n"
            "- 실결제 금액 : 1,100,000원\n"
            "- 응답 : 「반품 등록 완료」 — PT_USE 20,000 복원\n"
            "- 시나리오 7-3 : 2차 판매 나머지 반품 진행"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (member_no, no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            member_no,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 50, 95, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
