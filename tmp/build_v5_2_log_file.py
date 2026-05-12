"""V5-2 Log.xlsx — 회원 163001700337, 시나리오 V5-2.

「2. 판매처리 후 입금수정 단계에서 포인트/쿠폰을 삭제하고 해당 삭제금액만큼 현금/카드로 금액 수정 후 저장 테스트」
- [TEST]_프로모션없음 (audit PromotionNo: PRO202604161041)
- 시간: 2026-05-08 16:13:25 ~ 오류 (UTC 07:13:25 ~ 07:24:01)
- 주문: SOR202605080168  /  판매: SAL202605080242 → 재판매 SAL202605080246
"""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V5-2 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v5_1_matched.csv"  # same query covers V5-2 too

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V5-2"
MEMBER_NO = "163001700337"
SHEET_NAME = "V5-2"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac.replace("_", ""):
        return "포인트 (적용 조회)"
    if "saledeposit" in ac.replace("_", ""):
        return "판매 입금"
    if "returndeposit" in ac.replace("_", ""):
        return "반품 입금"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    # ───── 테스트 준비 ─────
    "a12JO000000MFbbYAG": (
        "준비(1)(2)",
        "주문 등록 (포인트 10K + 정률쿠폰 50K + 선수금 카드 10% / 프로모션 없음)",
        (
            "주문 번호 : SOR202605080168  (시간 : 2026-05-08 16:15:33 KST)\n"
            "- 매장 코드 : 99998 (백화점)  /  유형 : 01 (계약)\n"
            "- 실결제 금액 : 500,000원  (2A2600001 × 1, 순매가 단가 500,000)\n"
            "- 프로모션 번호 : PRO202604161041  (시나리오 기재 「[TEST]_프로모션없음」)\n"
            "- 거래 원장 (입금 번호 DEP202605080271) :\n"
            "    1) 결제 수단 01 (현금)  /  입금 구분 01 (계약금)  /  거래 금액 50,000원  ← 선수금 10%\n"
            "    2) 결제 수단 90 (쿠폰)  /  입금 구분 01  /  거래 금액 50,000원  /  쿠폰 COP2026MY0003 (정률쿠폰)\n"
            "    3) 결제 수단 81 (포인트)  /  입금 구분 01  /  거래 금액 10,000원\n"
            "- 응답 : 「주문 등록 완료」  /  차감 포인트 10,000P (CD 04)\n"
            "- 검증 : 준비 (1) 포인트/쿠폰/선수금 입력 정상 / (2) 「프로모션 없음」 효과로 후속 적립 비교 기준점 설정"
        ),
    ),
    "a12JO000000ME61YAG": (
        "준비(3)",
        "판매 처리 (주문 끌어와 잔금 카드 완불)",
        (
            "판매 번호 : SAL202605080242  /  원거래 주문 번호 : SOR202605080168\n"
            "- 시간 : 2026-05-08 16:20:32 KST\n"
            "- 실결제 금액 : 500,000원\n"
            "- 거래 원장 :\n"
            "    · DEP202605080271-1 : 현금 01 / 계약금 01 / 50,000원 (선수금)\n"
            "    · DEP202605080271-2 : 쿠폰 90 / 계약금 01 / 50,000원 (COP2026MY0003)\n"
            "    · DEP202605080271-3 : 포인트 81 / 계약금 01 / 10,000원\n"
            "    · DEP202605080275-1 : 결제 수단 01 / 잔금 02 / 390,000원  ← 잔금 카드 완불\n"
            "- 응답 적립 포인트 : 구매포인트 8,800P + 사은포인트 13,200P (프로모션 없음 → 표준 배율)\n"
            "- 검증 : 준비 (3) 원판매번호 SAL242 정상 생성"
        ),
    ),
    # ───── 진행 ─────
    "a12JO000000MIWAYA4": (
        "진행(1)(2)",
        "판매 입금내역 포인트/쿠폰 삭제 후 저장 — 「판매 수정 완료」",
        (
            "판매 번호 : SAL202605080242  (수정)  /  시간 : 2026-05-08 16:21:48 KST\n"
            "- 거래 원장 (포인트 81 / 쿠폰 90 분개 삭제 후 저장) :\n"
            "    · DEP202605080271-1 : 현금 01 / 계약금 01 / 50,000원  ← 카드 row 만 남김\n"
            "    · DEP202605080275-1 : 결제 수단 01 / 잔금 02 / 450,000원  ← 잔금 50K 증가 (포인트 10K + 쿠폰 50K → 카드 잔금 흡수 후 일부만 잔액에 반영)\n"
            "  ※ 시나리오 진행 (1) 「포인트 삭제 + 쿠폰 삭제」 의도가 audit 에 그대로 반영 (Pay 81/90 분개 부재)\n"
            "  ※ 잔금 합계 50K + 450K = 500K 매칭되어 시스템상 「판매 수정 완료」 응답 반환\n"
            "- 응답 메시지 : 「판매 수정 완료」 / success: true (audit 상 정상 처리)\n"
            "- 응답 적립 포인트 (재계산) : 구매포인트 10,000P + 사은포인트 15,000P\n"
            "    ※ 포인트 사용액(10,000) 만큼 적립 베이스 증가 → 결과 적립 포인트 증가\n"
            "- 시나리오 검증 : 진행 (2-1) 「저장 성공 시 - Error」 ↔ audit 「수정 완료」 — 시스템이 차단해야 할 케이스가 차단되지 않은 점 확인 필요"
        ),
    ),
    "a12JO000000MFoQYAW": (
        "진행(4)",
        "판매 입금 삭제 — DEP202605080271-1 DELETE",
        (
            "Apex 클래스 : GdSaleDepositApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-08 16:23:08 KST\n"
            "- 대상 입금 : DepositNo DEP202605080271 / DeposiSeqNo 1 (계약금 카드 50K 행)\n"
            "- URL : {\"DeposiSeqNo__c\":\"1\",\"DepositNo__c\":\"DEP202605080271\"} (RequestBody 없음)\n"
            "- 응답 메시지 : 「판매 입금 삭제 삭제 완료」 / success: true\n"
            "- 응답 적립 포인트 소멸 : CD 01 / 1,000P + CD 06 / 1,500P (50K 카드 행 분 적립 소멸)\n"
            "- 검증 : 진행 (4) 판매 입금 행 삭제 → 재판매 진행 가능 상태"
        ),
    ),
    "a12JO000000MHOrYAO": (
        "진행(4)",
        "재판매 처리 — SAL202605080246 (주문 입금 다시 불러오기)",
        (
            "판매 번호 : SAL202605080246  /  원거래 주문 번호 : SOR202605080168  /  시간 : 2026-05-08 16:23:45 KST\n"
            "- 실결제 금액 : 500,000원\n"
            "- 거래 원장 (주문에서 사용한 포인트·쿠폰 다시 불러옴) :\n"
            "    · DEP202605080271-1 : 현금 01 / 계약금 01 / 50,000원\n"
            "    · DEP202605080271-2 : 쿠폰 90 / 계약금 01 / 50,000원 (COP2026MY0003)\n"
            "    · DEP202605080271-3 : 포인트 81 / 계약금 01 / 10,000원\n"
            "    · DEP202605080278-1 : 결제 수단 01 / 잔금 02 / 390,000원\n"
            "- 응답 메시지 : 「판매 등록 완료」 / success: true\n"
            "- 응답 적립 포인트 : 구매포인트 8,800P + 사은포인트 13,200P (준비(3) SAL242 와 동일)\n"
            "- 시나리오 검증 : 진행 (4) 「판매 입금 전부 삭제 후 재판매처리」 → 주문에서 사용한 포인트/쿠폰 정상 복원·재적용 ✓"
        ),
    ),
    "a12JO000000MEdrYAG": (
        "재판매 후 오류",
        "재판매 후 에코포인트 지급 시도 — 「지급할 포인트가 0 이하」 오류",
        (
            "Apex 클래스 : GdPointCreditApiControllerV1 — /gd/v1/point/credit\n"
            "- 시간 : 2026-05-08 16:24:01 KST  (재판매 SAL246 후 16초)\n"
            "- 대상 판매 번호 : SAL202605080246\n"
            "- PointsList : [ point 0.0 / cd_type_point 05 (에코포인트) ]\n"
            "- 응답 메시지 : 「지급할 포인트가 0 이하입니다.」 / success: false / code: 400\n"
            "- ※ 시나리오 종료시간 「오류」 와 일치 — 마지막 호출이 에코포인트 0원 입력 → 검증 실패로 종료\n"
            "- 검증 : 운영 의도상 0P 입력은 차단되어야 하므로 오류 응답이 정상 동작 (단, 화면에서 0P 입력이 가능했던 부분은 별도 UI 점검 필요)"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 40, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
