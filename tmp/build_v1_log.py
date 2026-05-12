"""V1 Log.xlsx — 회원 TEST02006 / V1 시나리오 (2-1 ~ 2-5).

원판매 포인트 사용 후 환불 시나리오 (A판매 1,000만 → B판매 100만 → C판매 200만
→ A판매 부분반품 불가 확인).

※ 시나리오 페이지의 회원번호 133005100137 은 마이그 전 표기. 실제 로그 회원: TEST02006.
"""
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V1 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v1_test02006_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V1"
MEMBER_NO = "TEST02006"
SHEET_NAME = "V1"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower().replace("_", "")
    if "memberhelpinfo" in ac:
        return "회원 도움창"
    if "memberapi" in ac:
        return "회원"
    if "voucherhelp" in ac:
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac:
        return "포인트 (적용 조회)"
    if "saledeposit" in ac:
        return "판매 입금"
    if "returndeposit" in ac:
        return "반품 입금"
    if "orderapi" in ac:
        return "주문"
    if "saleapi" in ac:
        return "판매"
    if "returnapi" in ac:
        return "반품"
    if "pointcredit" in ac:
        return "포인트 (지급)"
    if "pointdebit" in ac:
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    # 2-1 A 판매 등록
    "a12JO000000Me3gYAC": (
        "2-1",
        "A 판매 등록 — SAL02V002 (1,000만원 GIP, 구매포인트 200,000P 적립)",
        (
            "판매 번호 : SAL02V002  (시간 : 2026-05-10 18:03:59 KST)\n"
            "- 회원 : TEST02006  /  매장 코드 : 99998 (백화점)\n"
            "- 실결제 금액 : 10,000,000원\n"
            "- 프로모션 : PRO202604161041\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 200,000P (구매포인트 2배)\n"
            "- 시나리오 단계 : 2-1 A 판매 등록 (적립포인트 생성)"
        ),
    ),
    # 2-2 A판매 에코포인트 저장
    "a12JO000000Me5FYAS": (
        "2-2",
        "A판매 에코포인트 저장 — SAL02V002 9,000P 지급",
        (
            "포인트 지급 : 이벤트포인트 9,000P  (시간 : 2026-05-10 18:04:53 KST)\n"
            "- 회원 : TEST02006\n"
            "- 대상 판매 번호 : SAL02V002\n"
            "- 포인트 유형 : 05 (이벤트포인트 / 에코포인트)\n"
            "- 응답 : 「포인트 지급 성공」 — PT_TARGET 9,000\n"
            "- 시나리오 단계 : 2-2 A판매건 에코포인트 저장 (현재 보유 포인트 9,000P)"
        ),
    ),
    # 2-3 B 판매 1차 시도 (ERROR)
    "a12JO000000MeA6YAK": (
        "2-3 시도",
        "B 판매 등록 1차 시도 — SAL02V003 ❌ ERROR DUPLICATE_VALUE",
        (
            "판매 번호 : SAL02V003  (시간 : 2026-05-10 18:07:41 KST)\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 응답 : ❌ code 500 — DUPLICATE_VALUE (DepositNo__c 중복)\n"
            "- 시나리오 단계 : 2-3 B 판매 1차 시도 (DepositNo 중복 에러)"
        ),
    ),
    # 2-3 B 판매 재시도 (성공)
    "a12JO000000MeBkYAK": (
        "2-3",
        "B 판매 등록 — SAL02V003 (100만원, 생일쿠폰 3만 + 포인트 30만 사용, 13,400P 적립)",
        (
            "판매 번호 : SAL02V003  (시간 : 2026-05-10 18:07:50 KST)\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 결제 구성 : 생일쿠폰 30,000 + 포인트 300,000 사용 → 입금 670,000\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 13,400P\n"
            "- 시나리오 단계 : 2-3 B 판매 등록"
        ),
    ),
    # 2-4 C 판매 1차 시도 (포인트 부족 ERROR)
    "a12JO000000MeDKYA0": (
        "2-4 시도",
        "C 판매 등록 1차 시도 — SAL02V004 ❌ 「포인트 잔액 부족」",
        (
            "판매 번호 : SAL02V004  (시간 : 2026-05-10 18:09:58 KST)\n"
            "- 실결제 금액 : 2,000,000원\n"
            "- 응답 : ❌ code 400\n"
            "    「포인트 잔액이 부족합니다. 사용 가능 금액: 878,652P (요청: 1,000,000P)」\n"
            "- 시나리오 단계 : 2-4 C 판매 1차 시도 (요청 100만P > 잔액 878,652P)"
        ),
    ),
    # 2-4 C 판매 재시도 (성공)
    "a12JO000000MdnZYAS": (
        "2-4",
        "C 판매 등록 — SAL02V004 (200만원, 포인트 156,000P 사용, 22,600P 적립)",
        (
            "판매 번호 : SAL02V004  (시간 : 2026-05-10 18:10:32 KST)\n"
            "- 실결제 금액 : 2,000,000원\n"
            "- 결제 구성 : 포인트 156,000 사용 (잔액 조정 후)\n"
            "- 응답 : 「판매 등록 완료」 — PT_TARGET 22,600P\n"
            "- 시나리오 단계 : 2-4 C 판매 등록"
        ),
    ),
    # 2-5 A 부분반품 1차 시도 (SAL02V004 already exists 에러)
    "a12JO000000MeGXYA0": (
        "2-5 시도",
        "A 부분반품 1차 시도 — SAL02V004 ❌ 「SaleNo__c already exists」",
        (
            "반품 번호 : SAL02V004  /  원거래 판매 번호 : SAL02V002\n"
            "- 시간 : 2026-05-10 18:13:58 KST\n"
            "- 실결제 금액 : 9,000,000원 (부분환불)\n"
            "- 응답 : ❌ code 400 「SaleNo__c already exists」\n"
            "- 시나리오 단계 : 2-5 1차 시도 (반품번호로 SAL02V004 잘못 입력 → 이미 C판매에 사용중)"
        ),
    ),
    # 2-5 A 부분반품 재시도 (★ 핵심 PASS — 적립포인트 사용중 에러)
    "a12JO000000MdUBYA0": (
        "2-5 ★ 핵심",
        "A 부분반품 진행 — SAL02V005 ❌ 「적립포인트 다른 판매에서 사용중」 ✓ PASS",
        (
            "반품 번호 : SAL02V005  /  원거래 판매 번호 : SAL02V002\n"
            "- 시간 : 2026-05-10 18:14:05 KST  (1차 시도 후 7초)\n"
            "- 실결제 금액 : 9,000,000원 (부분환불 — 18 수량)\n"
            "- 응답 : ❌ code 400\n"
            "    「반품 입금 등록 불가 — 적립된 포인트가 다른 주문/판매에서 이미\n"
            "     사용되었습니다. 판매번호: SAL02V004 / 입금번호: DEP02V004-2」\n"
            "- 시나리오 단계 : 2-5 A 부분반품 불가 ✓ PASS\n"
            "- 검증 : A판매(SAL02V002)에 적립된 포인트를 C판매(SAL02V004)에서 사용 중\n"
            "  → A 부분반품 차단 정상 동작"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 50, 95, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
