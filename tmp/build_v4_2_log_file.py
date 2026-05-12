"""V4-2 Log.xlsx — 회원 103006200124 (한상희 VVIP).

시나리오 V4-2 / 단계 4-2-1 ~ 4-2-4
백화점·아울렛에서 쿠폰 금액 변경하여 사용 후 판매삭제 / 반품 시 쿠폰으로 복원되는지 확인
"""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V4-2 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v4_2_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V4-2"
MEMBER_NO = "103006200124"
SHEET_NAME = "V4-2"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac.replace("_", ""):
        return "포인트 (적용 조회)"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    return "POST"


# (Salesforce log Id) -> (번호, 확인내용, 확인 기준)
SCENARIO = {
    # ───── 4-2-1 : 백화점 주문 + 판매 + 반품 (VVIP 생일쿠폰 10만원→7만원 수정 사용) ─────
    "a12JO000000M3QjYAK": (
        "4-2-1-a",
        "백화점 주문 등록 (VVIP 생일쿠폰 10만원 → 7만원 수정 사용)",
        (
            "주문 번호 : SOR202605080077  (시간 : 2026-05-08 09:59:52 KST)\n"
            "- 매장 코드 : 99998 (백화점)\n"
            "- 유형 : 01 (계약)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 프로모션 번호 : PRO202604161041\n"
            "- 주문 항목 : 상품 코드 2A2600001 / 수량 1 / 순매가 단가 500,000\n"
            "- 거래 원장 (입금 번호 DEP202605080010) :\n"
            "    1) 결제 수단 03 (카드)  /  입금 구분 01 (계약금)  /  거래 금액 50,000원\n"
            "    2) 결제 수단 90 (쿠폰)  /  입금 구분 01 (계약금)  /  거래 금액 70,000원  /  쿠폰 번호 COP2026MY0012;-;9S8WTK5N  ← VVIP 생일쿠폰 10만원 → 7만원으로 수정\n"
            "- 검증 : 주문 단계에서 생일쿠폰 7만원으로 수정 적용 (계약금 단계)"
        ),
    ),
    "a12JO000000M7HQYA0": (
        "4-2-1-b",
        "해당 주문 건 판매 처리",
        (
            "판매 번호 : SAL202605080058  /  원거래 주문 번호 : SOR202605080077\n"
            "- 시간 : 2026-05-08 10:02:20 KST\n"
            "- 실결제 금액 : 500,000원\n"
            "- 프로모션 번호 : PRO202604161041\n"
            "- 거래 원장 :\n"
            "    1) DEP202605080010-1 : 결제 수단 03 (카드) / 입금 구분 01 / 50,000원\n"
            "    2) DEP202605080010-2 : 결제 수단 90 (쿠폰) / 입금 구분 01 / 70,000원 (쿠폰 COP2026MY0012)\n"
            "    3) DEP202605080011-1 : 결제 수단 01 (현금) / 입금 구분 02 (잔금) / 380,000원\n"
            "- 응답 적립 포인트 목록 :\n"
            "    · 적립 사유 「구매포인트 × 2.00」  /  적립 포인트 8,600P\n"
            "    · 적립 사유 「사은포인트 × 3.00」  /  적립 포인트 12,900P\n"
            "- 검증 : 판매 단계에서 쿠폰 사용 7만원 + 잔금 38만원 정상 처리"
        ),
    ),
    "a12JO000000M7R6YAK": (
        "4-2-1-c",
        "해당 판매 건 반품 등록 → 쿠폰 환불 + 쿠폰 복원 검증",
        (
            "반품 번호 : SAL202605080059  /  원거래 판매 번호 : SAL202605080058\n"
            "- 시간 : 2026-05-08 10:03:40 KST\n"
            "- 실결제 금액 : 500,000원\n"
            "- 판매 항목 : 상품 코드 2A2600001 / 수량 -1 (반품)\n"
            "- 거래 원장 (입금 번호 DEP202605080012) :\n"
            "    1) 결제 수단 90 (쿠폰)  /  입금 구분 03 (반품)  /  거래 금액 70,000원  /  쿠폰 COP2026MY0012\n"
            "    2) 결제 수단 01 (현금)  /  입금 구분 03 (반품)  /  거래 금액 430,000원\n"
            "- 검증 :\n"
            "    1) VVIP 생일쿠폰 7만원 입금처리된 것 환불 ✓ (DEP202605080012-1)\n"
            "    2) 쿠폰 복원 = VVIP 생일쿠폰 10만원으로 복원되는지 확인  ※ ERP 쿠폰 마스터 별도 확인 필요"
        ),
    ),
    # ───── 4-2-2 : 매장 재고 직접 판매 → 판매 삭제(반품 형태) (10만원→7만원) ─────
    "a12JO000000M5QvYAK": (
        "4-2-2-a",
        "백화점 매장 재고 직접 판매 처리 (VVIP 생일쿠폰 10만원 → 7만원)",
        (
            "판매 번호 : SAL202605080060  (주문 없이 매장 재고 직접 판매)\n"
            "- 시간 : 2026-05-08 10:05:57 KST\n"
            "- 매장 코드 : 99998 (백화점)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 프로모션 번호 : PRO202604161041\n"
            "- 판매 항목 : 상품 코드 2A2600001 / 수량 1 / 순매가 단가 500,000\n"
            "- 거래 원장 (입금 번호 DEP202605080013) :\n"
            "    1) 결제 수단 90 (쿠폰)  /  입금 구분 02 (납입/잔금)  /  거래 금액 70,000원  /  쿠폰 COP2026MY0012;-;9S8WTK5N\n"
            "    2) 결제 수단 01 (현금)  /  입금 구분 02  /  거래 금액 430,000원\n"
            "- 응답 적립 포인트 목록 : 구매포인트 8,600P + 사은포인트 12,900P\n"
            "- ※ 매장 재고 직접 판매라 주문 단계 없음 — 입금 구분 02 로 한 번에 처리"
        ),
    ),
    "a12JO000000M7kSYAS": (
        "4-2-2-b",
        "해당 판매 건 원마감 전 판매 삭제(반품 처리) → 쿠폰 환불 + 쿠폰 복원 검증",
        (
            "반품(판매 삭제) 번호 : SAL202605080061  /  원거래 판매 번호 : SAL202605080060\n"
            "- 시간 : 2026-05-08 10:07:07 KST\n"
            "- 실결제 금액 : 500,000원\n"
            "- 판매 항목 : 상품 코드 2A2600001 / 수량 -1\n"
            "- 거래 원장 (입금 번호 DEP202605080014) :\n"
            "    1) 결제 수단 90 (쿠폰)  /  입금 구분 03 (반품)  /  거래 금액 70,000원  /  쿠폰 COP2026MY0012\n"
            "    2) 결제 수단 01 (현금)  /  입금 구분 03  /  거래 금액 430,000원\n"
            "- 검증 :\n"
            "    1) 판매 삭제 시에도 VVIP 생일쿠폰 7만원 입금 처리된 것 환불 ✓\n"
            "    2) 쿠폰 복원 = VVIP 생일쿠폰 10만원으로 복원되는지 확인  ※ ERP 쿠폰 마스터 별도 확인 필요"
        ),
    ),
    # ───── 4-2-3 : 아울렛 수선 (수선 판매내역은 반품 불가 — PASS) ─────
    "a12JO000000M5NmYAK": (
        "4-2-3-a",
        "아울렛 수선 접수 (SVIP 수선쿠폰 사용)",
        (
            "수선 주문 번호 : SOR202605080078  (시간 : 2026-05-08 10:20:29 KST)\n"
            "- 매장 코드 : 99997 (아울렛)\n"
            "- 유형 : 03 (수선)\n"
            "- 실결제 금액 : 70,000원\n"
            "- 프로모션 번호 : (없음)\n"
            "- 주문 항목 : 상품 코드 2A2600001 / 수량 1 / 순매가 단가 70,000\n"
            "- 거래 원장 (입금 번호 DEP202605080016) :\n"
            "    1) 결제 수단 01 (현금)  /  입금 구분 01 (계약금)  /  거래 금액 7,000원\n"
            "    2) 결제 수단 90 (쿠폰)  /  입금 구분 01  /  거래 금액 63,000원  /  쿠폰 COP2026MY0021;-;V10Z19PZ\n"
            "- 시나리오 검증 : SVIP 수선쿠폰 사용 (수정 적용된 금액으로 입금 처리)"
        ),
    ),
    "a12JO000000M85QYAS": (
        "4-2-3-b",
        "해당 수선 건 판매 처리 — 아울렛 적립포인트 없음 / 수선 반품 불가 (PASS)",
        (
            "수선 판매 번호 : SAL202605080064  /  원거래 주문 번호 : SOR202605080078\n"
            "- 시간 : 2026-05-08 10:40:29 KST\n"
            "- 매장 코드 : 99997 (아울렛)\n"
            "- 실결제 금액 : 70,000원\n"
            "- 거래 원장 : DEP202605080016 (계약 시 사용한 것 그대로 — 7,000원 + 쿠폰 63,000원)\n"
            "- 응답 적립 포인트 목록 : (없음) — 아울렛 수선은 적립 대상 아님\n"
            "- 시나리오 검증 :\n"
            "    · 「수선 판매내역은 반품처리 할 수 없습니다」 메시지 노출 (수선 판매내역은 일반적으로 반품 등록 불가, IMC팀과 전산팀 확인완료) → 결과 PASS\n"
            "    · 이후 반품 등록 시도가 차단되므로 4-2-3-c 반품 단계는 발생하지 않음"
        ),
    ),
    # ───── 4-2-4 : 백화점 주문(2건) + 판매 + 부분반품 (SVIP 생일쿠폰 변경 사용) ─────
    "a12JO000000M8OmYAK": (
        "4-2-4-a",
        "백화점 주문 등록 (SVIP 생일쿠폰 10만원 → 변경 사용)",
        (
            "주문 번호 : SOR202605080081  (시간 : 2026-05-08 10:47:46 KST)\n"
            "- 매장 코드 : 99998 (백화점)\n"
            "- 유형 : 01 (계약)\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 프로모션 번호 : PRO202604161041\n"
            "- 주문 항목 (2건) :\n"
            "    1) 상품 코드 2A2600001  /  수량 1  /  순매가 단가 500,000\n"
            "    2) 상품 코드 3A2600005  /  수량 1  /  순매가 단가 500,000\n"
            "- 거래 원장 (입금 번호 DEP202605080023) :\n"
            "    1) 결제 수단 01 (현금)  /  입금 구분 01 (계약금)  /  거래 금액 100,000원\n"
            "    2) 결제 수단 90 (쿠폰)  /  입금 구분 01  /  거래 금액 10,000원  /  쿠폰 COP2026JA0029;-;169  ← SVIP 생일쿠폰 (시나리오 기재 5만원, audit 거래 금액 10,000원)\n"
            "- ※ 시나리오 텍스트는 「10만원→5만원 수정」 — audit 의 거래 금액은 10,000원으로 기록됨 (실 입력값과 시나리오 기재 금액 비교 확인 필요)"
        ),
    ),
    "a12JO000000M3QnYAK": (
        "4-2-4-b",
        "해당 주문 건 판매 처리",
        (
            "판매 번호 : SAL202605080068  /  원거래 주문 번호 : SOR202605080081\n"
            "- 시간 : 2026-05-08 10:50:16 KST\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 프로모션 번호 : PRO202604161041\n"
            "- 판매 항목 (2건) : 2A2600001 + 3A2600005 (각 500,000원)\n"
            "- 거래 원장 :\n"
            "    1) DEP202605080023-1 : 현금 01 / 입금 구분 01 / 100,000원\n"
            "    2) DEP202605080023-2 : 쿠폰 90 / 입금 구분 01 / 10,000원 (쿠폰 COP2026JA0029)\n"
            "    3) DEP202605080025-1 : 현금 01 / 입금 구분 02 (잔금) / 890,000원\n"
            "- 응답 적립 포인트 목록 :\n"
            "    · 구매포인트 × 2.00  →  적립 포인트 19,800P\n"
            "    · 사은포인트 × 3.00  →  적립 포인트 29,700P\n"
            "- 검증 : 쿠폰 적용된 결제 정상 판매 처리"
        ),
    ),
    "a12JO000000M6YJYA0": (
        "4-2-4-c",
        "해당 판매 건 부분 반품 등록 (제품 1건) → 쿠폰 환불 + 쿠폰 복원 검증",
        (
            "부분 반품 번호 : SAL202605080070  /  원거래 판매 번호 : SAL202605080068\n"
            "- 시간 : 2026-05-08 10:52:41 KST\n"
            "- 실결제 금액 : 500,000원  (1,000,000원 중 1건만 반품)\n"
            "- 판매 항목 : 상품 코드 2A2600001 / 수량 -1\n"
            "- 거래 원장 (입금 번호 DEP202605080027) :\n"
            "    1) 결제 수단 90 (쿠폰)  /  입금 구분 03 (반품)  /  거래 금액 10,000원  /  쿠폰 COP2026JA0029\n"
            "    2) 결제 수단 01 (현금)  /  입금 구분 03  /  거래 금액 490,000원\n"
            "- 검증 :\n"
            "    1) 부분 반품 시 SVIP 생일쿠폰 사용분(10,000원)도 환불 처리 ✓\n"
            "    2) 부분 반품 등록 시 사용 쿠폰 환불 불러오기 버튼 클릭 시 변경된 금액으로 정상 호출됨 (확인 완료)\n"
            "    3) 쿠폰 복원 = SVIP 생일쿠폰 10만원으로 복원되는지 확인 ※ ERP 쿠폰 마스터 별도 확인 필요\n"
            "- ※ 부분 반품 「쿠폰반품등록」 버튼으로 쿠폰 재등록 시 마이너스 금액 미연산, 0으로 끌고와짐 + 현금/쿠폰 순서 달라질 때도 미연산 — 별도 이슈"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 10, 38, 80, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
