"""V2-1 Log.xlsx — 회원 103006200153, 시나리오 V2-1 (사은불가 프로모션 / 판매-에코포인트-부분반품-완전반품)."""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V2-1 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v2_1_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V2-1"
MEMBER_NO = "103006200153"
MEMBER_NAME = ""  # 이미지에 회원명 명시 없음
SHEET_NAME = "V2-1"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac.replace("_", ""):
        return "포인트 (적용 조회)"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    return "POST"


SCENARIO = {
    "a12JO000000M8AHYA0": (
        "1-1",
        "판매 저장 (구매포인트 적립)",
        (
            "판매 : SAL202605080066  (Salesforce Id : 801JO00000IA07iYAD)\n"
            "- 시간 : 2026-05-08 10:44:38 KST\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 판매 입금 : 1,000,000원  (DepositType 02 · PaymentMethod 01 (현금) · DEP202605080022-1)\n"
            "- PromotionNo__c : PRO202604161042  ([TEST] 사은불가_프로모션 추정)\n"
            "- 적립포인트(구매포인트) : 20,000P  (TX_REMARK 구매포인트 × 2.00 / CD_TYPE_POINT 01 / CD_TYPE_SAVE Credit)\n"
            "- ※ 사은불가 프로모션이므로 사은포인트는 별도 지급되지 않음 (구매포인트만)\n"
            "- 시나리오 검증 : 「구매 포인트 적립 : 20,000P」 ✓"
        ),
    ),
    "a12JO000000M8IHYA0": (
        "1-2",
        "에코포인트 저장 (이벤트포인트 강제 지급)",
        (
            "GdPointCreditApiControllerV1 — /gd/v1/point/credit\n"
            "- 시간 : 2026-05-08 10:44:51 KST  (1-1 판매 직후 13초)\n"
            "- 적용 대상 : SaleNo__c SAL202605080066\n"
            "- TX_REMARK : 이벤트포인트  ※ 화면 표기로는 「에코포인트」\n"
            "- PT_TARGET : 6,000P  /  CD_TYPE_SAVE : Credit  /  CD_TYPE_POINT : 01\n"
            "- 응답 message : 「포인트 지급 성공」\n"
            "- 시나리오 검증 : 「에코 포인트 적립 : 6,000P」 ✓"
        ),
    ),
    "a12JO000000M6wTYAS": (
        "1-3",
        "부분 반품 (구매적립가능 품목만 / 구매포인트 차감)",
        (
            "부분 반품 : SAL202605080067  /  원거래 : SAL202605080066\n"
            "(Salesforce Id : 801JO00000I9yShYAJ)\n"
            "- 시간 : 2026-05-08 10:48:00 KST\n"
            "- Quantity : -1 (반품)\n"
            "- 환불 금액 : 500,000원  (판매 입금 취소 / DepositType 03 · PaymentMethod 01 (현금) · DEP202605080024-1)\n"
            "- 부분 반품 ActualPaymentAmount__c : 500,000원  (1-1 판매 1,000,000 중 절반)\n"
            "- 시나리오 검증 : 「구매 포인트 차감 : 10,000P」 — 1-1 적립 20,000P 의 절반 차감\n"
            "- ※ 구매적립가능 품목만 부분반품이므로 에코포인트는 유지 (1-4 완전반품 시 차감)"
        ),
    ),
    "a12JO000000M519YAC": (
        "1-4",
        "완전반품 (나머지 구매포인트 + 에코포인트 차감)",
        (
            "완전 반품 : SAL202605080069  /  원거래 : SAL202605080066\n"
            "(Salesforce Id : 801JO00000I9zQHYAZ)\n"
            "- 시간 : 2026-05-08 10:52:39 KST\n"
            "- Quantity : -1 (반품)\n"
            "- 환불 금액 : 500,000원  (판매 입금 취소 / DepositType 03 · PaymentMethod 01 (현금) · DEP202605080026-1)\n"
            "- 시나리오 검증 :\n"
            "    1) 「나머지 구매포인트 10,000P 차감」 — 1-1 적립 20,000P 중 1-3 차감 후 잔여분\n"
            "    2) 「에코포인트 차감 -6,000P」 — 1-2 적립 이벤트포인트 6,000P 회수 (완전반품 시 자동 차감)"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "ApexClass__c",
        "로그 ID",
        "CreatedDate",
        "RequestBody__c",
        "ResponseBody__c",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    def num_sort_key(item):
        no = item[1][0]
        parts = no.split("-")
        try:
            return tuple(int(p) for p in parts)
        except ValueError:
            return (99, 99)

    sorted_items = sorted(SCENARIO.items(), key=num_sort_key)

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 8, 38, 80, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
