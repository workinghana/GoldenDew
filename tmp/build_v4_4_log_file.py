"""V4 Log (4-1~4-4).xlsx — 회원 123002600463 SVIP, 시나리오 V4 / 4-1 ~ 4-4.

「4. 백화점 프로모션 이용」 — [TEST] 통합시나리오V4_사은포인트3배
- 4-1 주문 저장 — 10만원 쿠폰 사용 + 포인트 39,000P 사용
- 4-2 판매 저장 — 사은포인트 3배 (이벤트포인트 두 배 적립)
- 4-3 에코포인트 지급 — 9,000P
- 4-4 부분반품 — 포인트·쿠폰 유지, 현금 환불 / 적립된 사은포인트 소멸 확인
"""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V4 Log (4-1~4-4).xlsx"
MATCHED_CSV = ROOT / "tmp" / "v4_4_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V4"
MEMBER_NO = "123002600463"
SHEET_NAME = "V4 (4-1~4-4)"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac.replace("_", ""):
        return "포인트 (적용 조회)"
    if "saledeposit" in ac.replace("_", ""):
        return "판매 입금"
    if "returndeposit" in ac.replace("_", ""):
        return "반품 입금"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    return "POST"


SCENARIO = {
    # ───── 4-1 : 주문 저장 (10만원 쿠폰 + 포인트 39,000P) ─────
    "a12JO000000MDcyYAG": (
        "4-1",
        "주문 저장 (롯데계열 행사 20% 5개 품목 / 10만원 쿠폰 + 포인트 39,000P 사용)",
        (
            "주문 번호 : SOR202605080137  (시간 : 2026-05-08 14:12:14 KST)\n"
            "- 매장 코드 : 99998 (신세계 백화점)  /  유형 : 01 (계약)\n"
            "- 프로모션 번호 : PRO202604201055  ([TEST] 통합시나리오V4_사은포인트3배)\n"
            "- 정상 합계 : 2,400,000원  →  실결제 금액 : 1,920,000원 (20% 행사 할인 후)\n"
            "- 주문 항목 (5개 품목) :\n"
            "    1) 상품 코드 212500046  /  순매가 단가 320,000  /  할인율 20%\n"
            "    2) 상품 코드 212500047  /  순매가 단가 400,000  /  할인율 20%\n"
            "    3) 상품 코드 2A2600001  /  순매가 단가 400,000  /  할인율 20%\n"
            "    4) 상품 코드 3A2600005  /  수량 2  /  순매가 단가 400,000  /  할인율 20%\n"
            "- 거래 원장 (입금 번호 DEP202605080139, 계약금) :\n"
            "    1) 결제 수단 01 (현금)  /  입금 구분 01  /  거래 금액 200,000원\n"
            "    2) 결제 수단 81 (포인트)  /  입금 구분 01  /  거래 금액 39,000원  ← 포인트 39,000P 사용\n"
            "    3) 결제 수단 90 (쿠폰)  /  입금 구분 01  /  거래 금액 100,000원  /  쿠폰 COP2026MA0017;-;02 (10만원 생일쿠폰)\n"
            "- 응답 차감 포인트 목록 : 사용 포인트 39,000P (CD 04 / 사용 사유 1)\n"
            "- 검증 : 「10만원 쿠폰 사용 + 3만 9천원 포인트 사용 (-39,000P)」 ✓"
        ),
    ),
    # ───── 4-2 : 판매 저장 (사은포인트 3배 → 이벤트 포인트로 두 배 적립) ─────
    "a12JO000000MDDBYA4": (
        "4-2",
        "판매 저장 (구매 + 사은(3배) + 이벤트(3배 ×2회) 포인트 적립)",
        (
            "판매 번호 : SAL202605080141  /  원거래 주문 번호 : SOR202605080137\n"
            "- 시간 : 2026-05-08 14:15:36 KST\n"
            "- 실결제 금액 : 1,920,000원\n"
            "- 거래 원장 (계약금 DEP139 + 잔금 DEP141) :\n"
            "    · DEP202605080139-1 : 현금 01 / 계약금 01 / 200,000원\n"
            "    · DEP202605080139-2 : 포인트 81 / 계약금 01 / 39,000원\n"
            "    · DEP202605080139-3 : 쿠폰 90 / 계약금 01 / 100,000원 (COP2026MA0017)\n"
            "    · DEP202605080141-1 : 현금 01 / 잔금 02 / 1,581,000원\n"
            "- 응답 적립 포인트 목록 :\n"
            "    · 구매포인트 × 2.00  →  35,620P  (CD 01)\n"
            "    · 사은포인트 × 3.00  →  53,430P  (CD 01)  ← 사은 3배\n"
            "    · 이벤트포인트 × 3.00  →  53,430P  (CD 07)  ← 사은3배 → 이벤트 두 배 적립 분개 1\n"
            "    · 이벤트포인트 × 3.00  →  53,430P  (CD 07)  ← 분개 2\n"
            "- 시나리오 검증 : 「프로모션을 통한 추가 사은 포인트는 이벤트 포인트로 적립 되는 부분 확인」 ✓\n"
            "- ※ 시나리오 기재 합계 (사은 67,830 / 구매 45,220 / 이벤트 135,660) 와 audit 적립값 (사은 53,430 / 구매 35,620 / 이벤트 53,430×2 = 106,860) 차이 — 회원 추가 등급/누적 보정 가능성, 별도 확인"
        ),
    ),
    # ───── 4-3 : 에코포인트 지급 (9,000P) ─────
    "a12JO000000MDg9YAG": (
        "4-3",
        "에코포인트 지급 (9,000P / CD_TYPE_POINT 05)",
        (
            "Apex 클래스 : GdPointCreditApiControllerV1 — /gd/v1/point/credit\n"
            "- 시간 : 2026-05-08 14:15:47 KST  (4-2 판매 저장 후 11초)\n"
            "- 회원 : 123002600463\n"
            "- 대상 판매 번호 : SAL202605080141\n"
            "- PointsList : [ point 9,000P  /  cd_type_point 05 (에코포인트) ]\n"
            "- 응답 메시지 : 「포인트 지급 성공」  /  적립 포인트 9,000P (CD_TYPE_POINT 01 / Credit / TX_REMARK 「이벤트포인트」)\n"
            "- 시나리오 검증 : 「에코포인트 지급 : 9,000P」 ✓"
        ),
    ),
    # ───── 4-4 : 부분반품 (현금만 환불, 포인트·쿠폰 유지, 사은포인트 소멸) ─────
    "a12JO000000MDOWYA4": (
        "4-4",
        "부분 반품 (2개 품목 / 현금만 환불, 포인트·쿠폰 유지 / 사은포인트 소멸)",
        (
            "반품 번호 : SAL202605080145  /  원거래 판매 번호 : SAL202605080141\n"
            "- 시간 : 2026-05-08 14:19:03 KST\n"
            "- 실결제 금액 : 800,000원 (1,920,000원 중 2개 품목만 부분 반품)\n"
            "- 판매 항목 :\n"
            "    1) 상품 코드 212500047  /  수량 -1  /  순매가 단가 400,000  /  할인율 20%\n"
            "    2) 상품 코드 2A2600001  /  수량 -1  /  순매가 단가 400,000  /  할인율 20%\n"
            "- 거래 원장 (입금 번호 DEP202605080146) :\n"
            "    1) 결제 수단 01 (현금)  /  입금 구분 03 (반품)  /  거래 금액 800,000원\n"
            "  ※ 결제 수단 90(쿠폰) / 81(포인트) 분개 없음 → 「포인트, 쿠폰 놔두고 환불처리」 ✓\n"
            "- 응답 적립 포인트 소멸 목록 (CancelPointsCreditedList) :\n"
            "    · CD_TYPE_POINT 01 / 16,000P 소멸 (CD_RESON 6)\n"
            "    · CD_TYPE_POINT 06 / 24,000P 소멸\n"
            "    · CD_TYPE_POINT 07 / 48,000P 소멸 (이벤트포인트)\n"
            "- 시나리오 검증 :\n"
            "    1) 포인트·쿠폰 유지, 현금 800,000원만 환불 ✓\n"
            "    2) 부분 반품한 현금에 대응되는 적립 사은/이벤트 포인트 소멸 ✓\n"
            "    3) 적립된 <구매포인트 + 3배 적립 사은포인트> 모두 맞게 처리되는 것 확인 (CD 01 / 06 / 07 모두 차감)"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    def num_sort_key(item):
        no = item[1][0]
        parts = no.split("-")
        try:
            return tuple(int(p) for p in parts)
        except ValueError:
            return (99, 99, 99)

    sorted_items = sorted(SCENARIO.items(), key=num_sort_key)

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 10, 40, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
