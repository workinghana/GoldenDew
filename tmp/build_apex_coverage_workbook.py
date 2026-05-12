import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
RUN_DIR = ROOT / "tmp" / "coverage_all_20260504"
RESULT_JSON = RUN_DIR / "test-result-707JO00000ESofA.json"
CLASS_COVERAGE_JSON = RUN_DIR / "test-result-codecoverage.json"


def load_json(path):
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def percent(covered, total):
    if not total:
        return None
    return round((covered / total) * 100, 2)


def class_category(name):
    lower = name.lower()
    if lower.endswith("test") or lower.startswith("test") or "test" in lower:
        return "Test Class"
    return "Main Class"


def shorten_lines(lines, limit=300):
    values = sorted(set(int(x) for x in lines))
    text = ", ".join(str(x) for x in values)
    if len(text) <= limit:
        return text
    return text[:limit].rsplit(",", 1)[0] + ", ..."


def add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(value, str) and len(value) > 80)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        sample_limit = min(ws.max_row, 250)
        for row_idx in range(2, sample_limit + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 70))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 72)
    return ws


def main(output_path):
    result_payload = load_json(RESULT_JSON)
    result = result_payload.get("result", result_payload)
    summary = result.get("summary", {})
    coverage_payload = result.get("coverage", {})
    detailed_records = coverage_payload.get("records", [])
    class_coverage = load_json(CLASS_COVERAGE_JSON)

    local_files = {}
    for path in (ROOT / "force-app" / "main" / "default" / "classes").glob("*.cls"):
        local_files[path.stem] = str(path.relative_to(ROOT))

    class_totals = {}
    for item in class_coverage:
        name = item.get("name")
        if not name:
            continue
        total = item.get("totalLines") or 0
        covered = item.get("totalCovered") or 0
        class_totals[name] = {
            "total": total,
            "covered": covered,
            "uncovered": max(total - covered, 0),
            "percent": item.get("coveredPercent"),
        }

    tests_by_fullname = {}
    test_class_stats = defaultdict(lambda: {"Pass": 0, "Fail": 0, "Skip": 0, "Other": 0})
    for test in result.get("tests", []):
        test_class = (test.get("ApexClass") or {}).get("Name") or ""
        method = test.get("MethodName") or ""
        full_name = test.get("FullName") or f"{test_class}.{method}"
        outcome = test.get("Outcome") or "Other"
        tests_by_fullname[full_name] = test
        bucket = outcome if outcome in ("Pass", "Fail", "Skip") else "Other"
        test_class_stats[test_class][bucket] += 1

    by_class_test = defaultdict(lambda: defaultdict(lambda: {
        "methods": set(),
        "covered_lines": set(),
        "uncovered_lines": set(),
        "covered_count_sum": 0,
        "uncovered_count_sum": 0,
    }))
    detail_rows = []
    class_record_methods = defaultdict(set)
    class_record_test_classes = defaultdict(set)

    for rec in detailed_records:
        main_class = (rec.get("ApexClassOrTrigger") or {}).get("Name") or ""
        test_class = (rec.get("ApexTestClass") or {}).get("Name") or ""
        method = rec.get("TestMethodName") or ""
        full_name = f"{test_class}.{method}" if test_class and method else ""
        covered_lines = rec.get("Coverage", {}).get("coveredLines") or []
        uncovered_lines = rec.get("Coverage", {}).get("uncoveredLines") or []
        covered_count = rec.get("NumLinesCovered") or len(covered_lines)
        uncovered_count = rec.get("NumLinesUncovered") or len(uncovered_lines)
        total_for_method = covered_count + uncovered_count
        outcome = (tests_by_fullname.get(full_name) or {}).get("Outcome")

        agg = by_class_test[main_class][test_class]
        agg["methods"].add(method)
        agg["covered_lines"].update(covered_lines)
        agg["uncovered_lines"].update(uncovered_lines)
        agg["covered_count_sum"] += covered_count
        agg["uncovered_count_sum"] += uncovered_count
        class_record_methods[main_class].add(full_name)
        class_record_test_classes[main_class].add(test_class)

        detail_rows.append([
            main_class,
            test_class,
            method,
            outcome,
            covered_count,
            uncovered_count,
            percent(covered_count, total_for_method),
            shorten_lines(covered_lines),
            shorten_lines(uncovered_lines),
        ])

    class_rows = []
    all_class_names = sorted(set(local_files) | set(class_totals))
    for name in all_class_names:
        totals = class_totals.get(name, {})
        category = class_category(name)
        note = ""
        if name in local_files and name not in class_totals and category == "Main Class":
            note = "No coverage data returned in this test run"
        class_rows.append([
            name,
            category,
            local_files.get(name, ""),
            totals.get("total"),
            totals.get("covered"),
            totals.get("uncovered"),
            totals.get("percent"),
            len([x for x in class_record_test_classes.get(name, set()) if x]),
            len([x for x in class_record_methods.get(name, set()) if x]),
            "Y" if name in by_class_test else "N",
            note,
        ])

    by_test_rows = []
    for main_class in sorted(by_class_test):
        total_lines = (class_totals.get(main_class) or {}).get("total") or 0
        for test_class in sorted(by_class_test[main_class]):
            agg = by_class_test[main_class][test_class]
            unique_covered = len(agg["covered_lines"])
            unique_uncovered = len(agg["uncovered_lines"])
            stats = test_class_stats.get(test_class, {})
            by_test_rows.append([
                main_class,
                test_class,
                len([m for m in agg["methods"] if m]),
                unique_covered,
                unique_uncovered,
                percent(unique_covered, total_lines),
                agg["covered_count_sum"],
                agg["uncovered_count_sum"],
                stats.get("Pass", 0),
                stats.get("Fail", 0),
                shorten_lines(agg["covered_lines"]),
            ])

    failure_rows = []
    for test in result.get("tests", []):
        if test.get("Outcome") == "Fail":
            failure_rows.append([
                (test.get("ApexClass") or {}).get("Name"),
                test.get("MethodName"),
                test.get("Message"),
                test.get("StackTrace"),
                test.get("RunTime"),
            ])

    summary_rows = [
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Source Result JSON", str(RESULT_JSON)],
        ["Source Coverage JSON", str(CLASS_COVERAGE_JSON)],
        ["Test Run ID", summary.get("testRunId")],
        ["Username", summary.get("username")],
        ["Hostname", summary.get("hostname")],
        ["Outcome", summary.get("outcome")],
        ["Tests Ran", summary.get("testsRan")],
        ["Passing", summary.get("passing")],
        ["Failing", summary.get("failing")],
        ["Skipped", summary.get("skipped")],
        ["Pass Rate", summary.get("passRate")],
        ["Fail Rate", summary.get("failRate")],
        ["Org Wide Coverage", summary.get("orgWideCoverage")],
        ["Test Run Coverage", summary.get("testRunCoverage")],
        ["Test Start Time", summary.get("testStartTime")],
    ]

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "Summary", ["Item", "Value"], summary_rows)
    add_sheet(wb, "Class_Coverage", [
        "Apex Class/Trigger",
        "Local Category",
        "Local File",
        "Total Lines",
        "Covered Lines",
        "Uncovered Lines",
        "Coverage %",
        "Covered By Test Classes",
        "Covered By Test Methods",
        "Has Detailed Records",
        "Notes",
    ], class_rows)
    add_sheet(wb, "TestClass_By_Class", [
        "Apex Class/Trigger",
        "Test Class",
        "Test Methods",
        "Unique Covered Lines",
        "Unique Uncovered Lines Seen",
        "Coverage % By This Test Class",
        "Covered Count Sum",
        "Uncovered Count Sum",
        "Test Class Pass Methods",
        "Test Class Fail Methods",
        "Covered Line Numbers",
    ], by_test_rows)
    add_sheet(wb, "TestMethod_Detail", [
        "Apex Class/Trigger",
        "Test Class",
        "Test Method",
        "Outcome",
        "Covered Lines",
        "Uncovered Lines",
        "Method Coverage %",
        "Covered Line Numbers",
        "Uncovered Line Numbers",
    ], detail_rows)
    add_sheet(wb, "Test_Failures", [
        "Test Class",
        "Test Method",
        "Message",
        "Stack Trace",
        "Run Time",
    ], failure_rows)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python build_apex_coverage_workbook.py <output.xlsx>")
    main(sys.argv[1])
