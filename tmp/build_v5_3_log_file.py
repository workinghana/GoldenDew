"""V5-3 Log.xlsx — 회원 133005100137, 시나리오 V5-3.

「3. 주문에서 입력한 프로모션 번호의 품목코드를 수정 → 타 매장 인도건 판매처리 → 판매 삭제 테스트」
- [TEST] 통합시나리오V5_사은2배 (PRO202605041080)
- 시간: 2026-05-08 15:51 ~ 16:33 KST
- 주문: SOR202605080162 (StoreCode 99998 / 인도매장 99995)
- 판매(1차): SAL202605080231  →  삭제
- 판매(2차): SAL202605080250 (인도매장 99995)  →  삭제
"""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V5-3 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v5_3_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V5-3"
MEMBER_NO = "133005100137"
SHEET_NAME = "V5-3"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac.replace("_", ""):
        return "포인트 (적용 조회)"
    if "saledeposit" in ac.replace("_", ""):
        return "판매 입금"
    if "returndeposit" in ac.replace("_", ""):
        return "반품 입금"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    # ───── 준비 (1)(2)(3) : 주문 등록 (선수금 100%, 사은 2배 프로모션, 인도매장 99995) ─────
    "a12JO000000MHgaYAG": (
        "준비",
        "주문 등록 — 선수금 100% (현금 50% + 카드 50%) / 사은2배 / 인도매장 99995",
        (
            "주문 번호 : SOR202605080162  (시간 : 2026-05-08 15:51:07 KST)\n"
            "- 매장 코드 : 99998 (주문매장 / 백화점, storeType 09)\n"
            "- 인도매장 : 99995 (시나리오 텍스트 — audit body 의 StoreCode 와 별도 필드)\n"
            "- 실결제 금액 : 900,000원 (기획제품 2개)\n"
            "- 프로모션 번호 : PRO202605041080  ([TEST] 통합시나리오V5_사은2배)\n"
            "- 주문 항목 (기획제품 2개) :\n"
            "    1) 상품 코드 212500046  /  수량 1  /  순매가 단가 400,000\n"
            "    2) 상품 코드 2A2600001  /  수량 1  /  순매가 단가 500,000\n"
            "- 거래 원장 (입금 번호 DEP202605080246, 선수금 100%) :\n"
            "    1) 결제 수단 01 (현금)  /  입금 구분 01 (계약금)  /  거래 금액 450,000원 (50%)\n"
            "    2) 결제 수단 03 (카드)  /  입금 구분 01  /  거래 금액 450,000원 (50%)\n"
            "- 응답 : 「주문 등록 완료」\n"
            "- 검증 : 준비 (1) 선수금 100% 입력 / (2) 사은2배 프로모션 / (3) 주문매장 99998·인도매장 99995"
        ),
    ),
    # ───── 진행 (1)(1-1) : SAL231 — 품목 변경 후 첫 판매 (주문매장 99998) ─────
    "a12JO000000MFEzYAO": (
        "진행(1)(1-1)",
        "품목 변경 후 판매 처리 — 변경된 품목으로 로열티 바인딩 ✓",
        (
            "판매 번호 : SAL202605080231  /  원거래 주문 번호 : SOR202605080162\n"
            "- 시간 : 2026-05-08 16:10:15 KST\n"
            "- 매장 코드 : 99998 (주문매장 = 판매매장)  /  storeType 09\n"
            "- 실결제 금액 : 900,000원\n"
            "- 판매 항목 (테스트 환경 강제 UPDATE 로 품목 변경된 후 판매처리) :\n"
            "    1) 상품 코드 212500047  ← 원 주문 212500046 에서 변경\n"
            "    2) 상품 코드 3A2600005  ← 원 주문 2A2600001 에서 변경\n"
            "- 응답 적립 포인트 :\n"
            "    · 구매포인트 × 1.00 → 9,000P  (CD 01)\n"
            "    · 사은포인트 × 3.00 → 27,000P  (CD 01)  ← 사은 2배 + 등급 보너스 추정\n"
            "    · 이벤트포인트 × 3.00 → 27,000P  (CD 07)  ← 이벤트도 함께 적립\n"
            "- 응답 메시지 : 「판매 등록 완료」 / success: true\n"
            "- 시나리오 검증 : 진행 (1-1) 「판매처리 되어야함 (변경 품목으로 로열티 바인딩)」 ✓\n"
            "  (시나리오 텍스트는 「판매처리 불가」 였으나 골든듀 전산/IMC팀 협의 후 판매처리 가능으로 변경)"
        ),
    ),
    # ───── (SAL231 판매 삭제) DEP246-1 + DEP246-2 ─────
    "a12JO000000MCqYYAW": (
        "(SAL231 삭제)",
        "SAL231 판매 입금 삭제 — DEP246-1, DEP246-2 (적립 포인트 일괄 소멸)",
        (
            "Apex 클래스 : GdSaleDepositApiControllerV1 (DELETE × 2건)\n"
            "- 16:12:00 : DEP202605080246-1 삭제 (현금 450K 분)\n"
            "- 16:12:03 : DEP202605080246-2 삭제 (카드 450K 분)  ※ 별도 행으로 audit a12JO000000MI6LYAW\n"
            "- 응답 메시지 : 「판매 입금 삭제 삭제 완료」  /  success: true\n"
            "- 응답 적립 포인트 소멸 (각 DELETE 마다) :\n"
            "    · CD 01 / 4,500P  (구매포인트)\n"
            "    · CD 06 / 13,500P  (사은포인트)\n"
            "    · CD 07 / 13,500P  (이벤트포인트)\n"
            "  → 두 DELETE 합산 : 9,000 + 27,000 + 27,000 = SAL231 적립 전액 소멸 ✓\n"
            "- 검증 : SAL231 정리 후 진행 (2) 인도매장 판매처리 단계로 진행"
        ),
    ),
    # ───── 진행 (2)(2-1) : SAL250 — 인도매장 99995 로 판매처리 ─────
    "a12JO000000MHC0YAO": (
        "진행(2)(2-1)",
        "인도매장 99995 로 판매 처리 — 변경 품목 그대로 / 사은2배 정상 적용",
        (
            "판매 번호 : SAL202605080250  /  원거래 주문 번호 : SOR202605080162\n"
            "- 시간 : 2026-05-08 16:26:45 KST\n"
            "- 매장 코드 : 99995 (인도매장 / storeType 08)  ← 시나리오상 인도매장\n"
            "- 실결제 금액 : 900,000원\n"
            "- 판매 항목 (변경 품목 동일) :\n"
            "    1) 상품 코드 212500047\n"
            "    2) 상품 코드 3A2600005\n"
            "- 거래 원장 : DEP202605080246 (준비 입금 그대로 — 현금 450K + 카드 450K)\n"
            "- 응답 적립 포인트 :\n"
            "    · 구매포인트 × 1.00 → 9,000P\n"
            "    · 사은포인트 × 3.00 → 27,000P  ← 인도매장 변경되어도 사은2배 정상 적용 ✓\n"
            "    · 이벤트포인트 × 3.00 → 27,000P\n"
            "- 응답 메시지 : 「판매 등록 완료」  /  success: true\n"
            "- 시나리오 검증 :\n"
            "    · 진행 (2-1) 「변경된 내용으로 로열티 바인딩」 ✓\n"
            "    · 예상 결과 「인도매장이 달라도 사은포인트 2배수 정상 동작」 ✓\n"
            "- ※ audit StoreCode__c 는 99995 (인도매장) 로 기록 — 시나리오 「판매매장은 주문매장으로 SAL 생성」 의도와 차이 (시스템이 인도매장 그대로 등록한 것으로 확인 필요)"
        ),
    ),
    # ───── 진행 (3) : SAL250 판매 삭제 — 사은포인트 2배 정상 소멸 검증 ─────
    "a12JO000000MFDLYA4": (
        "진행(3)",
        "SAL250 판매 삭제 — 사은포인트 2배수 정상 소멸 적용",
        (
            "Apex 클래스 : GdSaleDepositApiControllerV1 (DELETE × 2건)\n"
            "- 16:33:10 : DEP202605080246-1 삭제 (현금 450K 분)\n"
            "- 16:33:14 : DEP202605080246-2 삭제 (카드 450K 분)  ※ 별도 행으로 audit a12JO000000MIZNYA4\n"
            "- 응답 메시지 : 「판매 입금 삭제 삭제 완료」  /  success: true\n"
            "- 응답 적립 포인트 소멸 (각 DELETE 마다) :\n"
            "    · CD 01 / 4,500P  (구매포인트)\n"
            "    · CD 06 / 13,500P  (사은포인트 — 2배수 적립분 정상 소멸)\n"
            "    · CD 07 / 13,500P  (이벤트포인트)\n"
            "  → 두 DELETE 합산 : 9,000 + 27,000 + 27,000 = SAL250 적립 전액 소멸 ✓\n"
            "- 시나리오 검증 : 진행 (3) 「판매 삭제 진행 시 사은포인트 2배수 정상 적용」 ✓"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 40, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
