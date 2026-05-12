"""V4 Log.xlsx — 회원 103006200124 한상희 VVIP.

시나리오 V4 / 3-1 ~ 3-12 : 백화점 여러 매장 다양하게 이용
- 주문1 (R군행사) / 주문2 (정상가, 사은3배) / 주문3 (계열별_특별_행사)
- 시간창 2026-05-08 11:00 ~ 14:00 KST (UTC 02:00 ~ 05:00)
"""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V4 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v4_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V4"
MEMBER_NO = "103006200124"
SHEET_NAME = "V4"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac.replace("_", ""):
        return "포인트 (적용 조회)"
    if "saledeposit" in ac.replace("_", ""):
        return "판매 입금"
    if "returndeposit" in ac.replace("_", ""):
        return "반품 입금"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if "saledeposit" in (apex_class or "").lower() or "returndeposit" in (apex_class or "").lower():
        return "DELETE"
    return "POST"


# (Salesforce log Id) -> (번호, 확인내용, 확인 기준)
SCENARIO = {
    # 3-1 (주문1 R군행사 저장) 은 이번 시간창(02:00~05:00 UTC)에 audit 가 없음 — SAL202605070287(2026-05-07 등록 주문)의
    #     판매처리(3-8)가 존재하므로, 주문1 저장은 5월 7일 등 사전 작업으로 처리된 것으로 추정.

    # ───── 3-2 : 주문2 저장 (프로모션 PRO041, 5만 쿠폰 + 53,000P) ─────
    "a12JO000000M6wXYAS": (
        "3-2",
        "주문2 저장 (프로모션 PRO202604161041, 5만원 쿠폰 + 포인트 53,000원)",
        (
            "주문 번호 : SOR202605080083  (시간 : 2026-05-08 11:01:49 KST)\n"
            "- 매장 코드 : 99998 (백화점)  /  유형 : 01 (계약)\n"
            "- 실결제 금액 : 3,000,000원\n"
            "- 프로모션 번호 : PRO202604161041\n"
            "- 주문 항목 (3건) : 212500047 + 2A2600001 + 3A2600005 (각 1,000,000)\n"
            "- 거래 원장 (입금 번호 DEP202605080033) :\n"
            "    1) 결제 수단 01 (현금)  /  입금 구분 01 (계약금)  /  거래 금액 850,000원\n"
            "    2) 결제 수단 90 (쿠폰)  /  입금 구분 01  /  거래 금액 50,000원  /  쿠폰 COP2026JA0029;-;169\n"
            "    3) 결제 수단 81 (포인트)  /  입금 구분 01  /  거래 금액 53,000원\n"
            "- 응답 차감 포인트 목록 : 4종 (CD 04: 28,622P / CD 01: 9,378P / CD 06: 15,000P)\n"
            "- 검증 : 5만원 쿠폰 + 포인트 53,000원 적용된 주문 정상 등록"
        ),
    ),
    # ───── 3-3 : 주문2 취소 ─────
    "a12JO000000M8yFYAS": (
        "3-3",
        "주문2 취소 (5만원 쿠폰 복원 + 포인트 53,000원 복원)",
        (
            "주문 번호 : SOR202605080083  /  주문 상태 : 09 (취소)\n"
            "- 시간 : 2026-05-08 11:03:09 KST\n"
            "- 추가 거래 원장 (입금 번호 DEP202605080035 — 취소 분개) :\n"
            "    1) 결제 수단 90 (쿠폰)  /  입금 구분 03 (취소)  /  거래 금액 50,000원  /  쿠폰 COP2026JA0029  ← 쿠폰 복원\n"
            "    2) 결제 수단 81 (포인트)  /  입금 구분 03  /  거래 금액 53,000원  ← 포인트 사용 취소\n"
            "    3) 결제 수단 01 (현금)  /  입금 구분 03  /  거래 금액 850,000원  ← 현금 환불\n"
            "- 응답 메시지 : 「주문 수정 완료」\n"
            "- 검증 : 5만원 쿠폰 복원 ✓  /  포인트 사용 53,000원 복원 ✓"
        ),
    ),
    # ───── 3-4 : 주문2 재판매를 위한 포인트 조회 ─────
    "a12JO000000M8zqYAC": (
        "3-4",
        "주문2 재판매를 위한 포인트 조회 (사용 가능 포인트 조회)",
        (
            "Apex 클래스 : GdPointHelpApiControllerV1 — /gd/v1/point/help\n"
            "- 시간 : 2026-05-08 11:06:26 KST\n"
            "- 회원 : 103006200124  /  매장 코드 : 99998\n"
            "- 프로모션 번호 : PRO202604201055  ← 새 프로모션(사은3배)으로 변경\n"
            "- 응답 메시지 : 「포인트 적용 조회 성공」\n"
            "- 검증 : 사용 가능 포인트 조회 시 직전 취소(3-3)로 복원된 포인트 정상 반영"
        ),
    ),
    # ───── 3-5 : 주문2 재주문 저장(3) - 사은3배 ─────
    "a12JO000000M6QGYA0": (
        "3-5",
        "주문2 재주문 저장(3) (프로모션 PRO055 — 사은포인트 3배 / 5만 쿠폰 + 53,000P)",
        (
            "주문 번호 : SOR202605080084  (시간 : 2026-05-08 11:07:10 KST)\n"
            "- 실결제 금액 : 3,000,000원\n"
            "- 프로모션 번호 : PRO202604201055  ([TEST] 사은3배 행사)\n"
            "- 주문 항목 (3건) : 212500047 + 2A2600001 + 3A2600005 (각 1,000,000)\n"
            "- 거래 원장 (입금 번호 DEP202605080039) :\n"
            "    1) 현금 01  /  계약금 01  /  850,000원\n"
            "    2) 포인트 81  /  계약금 01  /  53,000원\n"
            "    3) 쿠폰 90  /  계약금 01  /  50,000원  /  쿠폰 COP2026JA0029\n"
            "- 검증 : 동일 결제 패턴 — 사은3배 프로모션으로 재등록"
        ),
    ),
    # ───── 3-6 : 주문2 판매 저장(3) — 최종 성공 SAL05050297 ─────
    "a12JO000000MAFIYA4": (
        "3-6",
        "주문2 판매 저장(3) — 구매/사은(3배)/이벤트(3배 ×2회) 포인트 모두 적립",
        (
            "판매 번호 : SAL202605050297  (시간 : 2026-05-08 11:33:27 KST)\n"
            "  ※ 판매처리는 SAL202605080076 → SAL05050295 → SAL05050296 → SAL05050297 로 3회 입금 수정 후 최종 등록\n"
            "- 실결제 금액 : 3,000,000원\n"
            "- 프로모션 번호 : PRO202604201055\n"
            "- 거래 원장 (계약금 DEP039 + 잔금 DEP053) :\n"
            "    · DEP202605080039-1 : 현금 01 / 계약금 01 / 850,000원\n"
            "    · DEP202605080039-2 : 포인트 81 / 계약금 01 / 53,000원\n"
            "    · DEP202605080039-3 : 쿠폰 90 / 계약금 01 / 50,000원 (COP2026JA0029)\n"
            "    · DEP202605080053-1 : 현금 01 / 잔금 02 / 2,047,000원\n"
            "- 응답 적립 포인트 목록 :\n"
            "    · 구매포인트 × 2.00  →  57,940P  (CD 01)\n"
            "    · 사은포인트 × 3.00  →  86,910P  (CD 01)  ← 사은 3배\n"
            "    · 이벤트포인트 × 3.00  →  86,910P  (CD 07)\n"
            "    · 이벤트포인트 × 3.00  →  86,910P  (CD 07)  ← 사은3배 프로모션이므로 이벤트포인트 두 배\n"
            "- 시나리오 검증 : 「구매·사은·이벤트포인트 모두 쌓임 / 사은3배라서 이벤트포인트가 사은의 두 배」 ✓"
        ),
    ),
    # ───── 3-7 : 주문3 저장(4) — 계열별_특별_행사 / 보유 포인트 전체 사용 ─────
    "a12JO000000MAP3YAO": (
        "3-7",
        "주문3 저장(4) (프로모션 PRO053 — 계열별_특별_행사 / 보유 포인트 전체 사용)",
        (
            "주문 번호 : SOR202605080095  (시간 : 2026-05-08 11:37:37 KST)\n"
            "- 실결제 금액 : 1,500,000원\n"
            "- 프로모션 번호 : PRO202604171053  ([TEST] 계열별_특별_행사)\n"
            "- 주문 항목 (3건) : 212500046 + 2A2600001 + 3A2600005 (각 500,000)\n"
            "- 거래 원장 (입금 번호 DEP202605080058) :\n"
            "    1) 현금 01 / 계약금 01 / 150,000원\n"
            "    2) 포인트 81 / 계약금 01 / 280,000원  ← 보유 포인트 사용\n"
            "    3) 쿠폰 90 / 계약금 01 / 75,000원 (쿠폰 COP2026JA0020)\n"
            "- 응답 차감 포인트 목록 : 4종 (이벤트포인트 86,910 + 86,910 + CD01 19,270 + CD06 86,910)\n"
            "- 시나리오 검증 : 보유 포인트 전체 사용 (직전 3-6 적립분 활용)"
        ),
    ),
    # ───── 3-8 : 주문1 판매처리 — R군행사 (구매포인트만 적립) ─────
    "a12JO000000MAjwYAG": (
        "3-8",
        "주문1 판매처리 — R군행사(사은포인트 적립 불가) → 구매포인트만 적립",
        (
            "판매 번호 : SAL202605070287  (시간 : 2026-05-08 11:42:31 KST)\n"
            "  ※ SaleNo prefix '202605070287' = 5월 7일 작성된 주문1의 당일 판매처리\n"
            "- 실결제 금액 : 2,500,000원\n"
            "- 프로모션 번호 : PRO202605021071  ([TEST] R군행사 추정)\n"
            "- 시나리오 검증 : R군행사는 사은포인트 적립 불가 → 구매포인트만 적립되는지 확인\n"
            "- ※ 회원정보 도움창의 적립 내역에 SAL202605070287 / 판매금액 2,500,000 / 구매포인트 50,000P (사은포인트 없음) 단일 항목으로 표시 → 검증 통과"
        ),
    ),
    # ───── 3-9 : 주문2 반품 처리(3) — 반품 첫 시도 ─────
    "a12JO000000MBuWYAW": (
        "3-9",
        "주문2 반품 처리(3) — 첫 시도 (반품 사유 주문 번호 바인딩 잘못됨 / 후속 단계 보정)",
        (
            "반품 번호 : SAL202605080098  /  원거래 판매 번호 : SAL202605050297\n"
            "- 시간 : 2026-05-08 13:20:55 KST\n"
            "- 실결제 금액 : 1,000,000원  (1건만 부분 반품)\n"
            "- 판매 항목 : 상품 코드 212500047 / 수량 -1\n"
            "- 검증 : 적립된 포인트가 다른 주문(주문3)에서 사용되어 있어 반품 불가\n"
            "    → 후속 3-10 단계에서 주문3 입금/포인트 사용 정리 후 3-11 에서 재시도\n"
            "- ※ 응답 메시지는 「반품 등록 완료」 로 보이나, 시나리오 의도는 「반품 사유 주문 번호 바인딩 잘못됨」 → 실제 운영 영향 별도 확인 필요"
        ),
    ),
    # ───── 3-10 : 주문3 입금 수정(4) — 포인트 사용 취소 (정리 단계) ─────
    "a12JO000000MCAdYAO": (
        "3-10",
        "주문3 입금 수정(4) — 사용된 포인트 사용 취소 (반품 가능 상태로 복구)",
        (
            "Apex 클래스 : GdPointDebitApiControllerV1\n"
            "- 시간 : 2026-05-08 13:52:56 KST\n"
            "- 대상 판매 : SAL202605050297  (주문2 판매 — 적립 포인트가 주문3에서 사용된 상태)\n"
            "- 직전 04:50:20 / 04:50:28 : DEP202605080063 의 ReturnDeposit 분개 삭제 (DeposiSeqNo 3, 2)\n"
            "- 작업 의미 : 주문3 의 포인트 사용분을 차감 취소하여 주문2 적립 포인트 복원 → 3-11 부분 반품 진행 가능\n"
            "- 시나리오 검증 : 「사용했던 포인트 모두 사용 취소」 ✓"
        ),
    ),
    # ───── 3-11 : 주문2 반품 처리(3) — 재시도 (성공) ─────
    "a12JO000000MD3UYAW": (
        "3-11",
        "주문2 반품 처리(3) — 부분 반품 재시도 (포인트 정리 후 진행 가능)",
        (
            "반품 번호 : SAL202605080134  /  원거래 판매 번호 : SAL202605050297\n"
            "- 시간 : 2026-05-08 13:54:00 KST\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 판매 항목 : 상품 코드 212500047 / 수량 -1 (1건 부분 반품)\n"
            "- 시나리오 검증 :\n"
            "    1) 부분 반품 진행 가능 ✓ (3-10 에서 포인트 사용 취소 후 처리)\n"
            "    2) 부분 반품 시 구매포인트(1%) + 사은포인트(3배) 적립된 분 모두 맞게 소멸 처리되는지 확인\n"
            "    3) 기존 판매 건에 사용되었던 생일쿠폰 5만원 + 포인트 53,000원은 남은 2개 제품 판매에 남김\n"
            "- 응답 메시지 : 「반품 등록 완료」"
        ),
    ),
    # 3-12 (주문3 판매 처리 — 잔여 포인트 모두 사용) 는 이번 시간창(02:00~05:00 UTC) 내에 audit 없음 — 후속 작업으로 추정.
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    def num_sort_key(item):
        no = item[1][0]
        parts = no.split("-")
        try:
            return tuple(int(p) for p in parts)
        except ValueError:
            return (99, 99, 99)

    sorted_items = sorted(SCENARIO.items(), key=num_sort_key)

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 10, 40, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
