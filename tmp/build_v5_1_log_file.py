"""V5-1 Log.xlsx — 회원 163001700337, 시나리오 V5-1.

「원 판매(주문)에 포인트와 쿠폰을 사용했는데, 최종 통 반품 시점에 포인트와 쿠폰을 100% 입력하지 않고 저장하는 경우」
- [TEST] 통합시나리오V5_구매2배 (PRO202605041079)
- 시간: 2026-05-08 15:39:58 ~ 16:09:10 KST (UTC 06:39:58 ~ 07:09:10)
- 주문: SOR202605080160  /  판매: SAL202605080213
- 반품: SAL202605080222 (부분), SAL202605080227 (시도→삭제), SAL202605080229 (최종 PASS)
"""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V5-1 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v5_1_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V5-1"
MEMBER_NO = "163001700337"
SHEET_NAME = "V5-1"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac.replace("_", ""):
        return "포인트 (적용 조회)"
    if "saledeposit" in ac.replace("_", ""):
        return "판매 입금"
    if "returndeposit" in ac.replace("_", ""):
        return "반품 입금"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith('{"SaleNo'):
        return "DELETE"
    return "POST"


SCENARIO = {
    # ───── 사전 셋팅 ─────
    "a12JO000000MHBwYAO": (
        "사전(1)(2)",
        "주문 등록 (포인트 10K + 정액 쿠폰 30K + 선수금 카드 10% / 구매적립 2배 프로모션)",
        (
            "주문 번호 : SOR202605080160  (시간 : 2026-05-08 15:42:07 KST)\n"
            "- 매장 코드 : 99998 (백화점)  /  유형 : 01 (계약)\n"
            "- 실결제 금액 : 1,000,000원  (2A2600001 × 2건, 각 500,000)\n"
            "- 프로모션 번호 : PRO202605041079  ([TEST] 통합시나리오V5_구매2배)\n"
            "- 거래 원장 (입금 번호 DEP202605080235) :\n"
            "    1) 결제 수단 03 (카드)  /  입금 구분 01 (계약금)  /  거래 금액 100,000원  ← 선수금 카드 10%\n"
            "    2) 결제 수단 90 (쿠폰)  /  입금 구분 01  /  거래 금액 30,000원  /  쿠폰 COP2026MY0004 (정액쿠폰)\n"
            "    3) 결제 수단 81 (포인트)  /  입금 구분 01  /  거래 금액 10,000원\n"
            "- 응답 : 「주문 등록 완료」  /  차감 포인트 10,000P (CD 04)\n"
            "- 검증 : 사전 (1) 포인트/쿠폰/선수금 입력 정상 / (2) 구매적립 2배 프로모션 적용"
        ),
    ),
    "a12JO000000MFGbYAO": (
        "사전(3)",
        "판매 처리 (주문 끌어와 원판매번호 생성)",
        (
            "판매 번호 : SAL202605080213  /  원거래 주문 번호 : SOR202605080160\n"
            "- 시간 : 2026-05-08 15:46:42 KST\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 거래 원장 :\n"
            "    · DEP202605080235-1 : 카드 03 / 계약금 01 / 100,000원\n"
            "    · DEP202605080235-2 : 쿠폰 90 / 계약금 01 / 30,000원 (COP2026MY0004)\n"
            "    · DEP202605080235-3 : 포인트 81 / 계약금 01 / 10,000원\n"
            "    · DEP202605080241-1 : 현금 01 / 잔금 02 / 860,000원\n"
            "- 응답 적립 포인트 :\n"
            "    · 구매포인트 × 2.00 → 19,200P  (CD 01)  ← 구매 2배 프로모션 적용\n"
            "    · 사은포인트 × 3.00 → 28,800P  (CD 01)\n"
            "    · 이벤트포인트 × 2.00 → 19,200P  (CD 07)  ← 구매 2배라 이벤트도 2배\n"
            "- 검증 : 사전 (3) 원판매번호 SAL213 정상 생성 / 구매적립 2배 적용 ✓"
        ),
    ),
    "a12JO000000MFZuYAO": (
        "사전(4)",
        "에코포인트 6,000원 지급",
        (
            "Apex 클래스 : GdPointCreditApiControllerV1 — /gd/v1/point/credit\n"
            "- 시간 : 2026-05-08 15:46:58 KST  (사전(3) 판매 직후 16초)\n"
            "- 회원 : 163001700337  /  대상 판매 번호 : SAL202605080213\n"
            "- PointsList : [ point 6,000P / cd_type_point 05 (에코포인트) ]\n"
            "- 응답 : 「포인트 지급 성공」 / 적립 6,000P (TX_REMARK 「이벤트포인트」)\n"
            "- 검증 : 사전 (4) 에코포인트 6,000원 정상 지급 ✓"
        ),
    ),
    # ───── 진행 ─────
    "a12JO000000MHjoYAG": (
        "진행(1)",
        "원판매 부분반품 (1개 품목 / 포인트만 5,000원 입력)",
        (
            "반품 번호 : SAL202605080222  /  원거래 판매 번호 : SAL202605080213\n"
            "- 시간 : 2026-05-08 15:57:48 KST\n"
            "- 실결제 금액 : 500,000원 (1개 품목 부분반품)\n"
            "- 판매 항목 : 상품 코드 2A2600001 / 수량 -1 / 순매가 단가 500,000\n"
            "- 거래 원장 (입금 번호 DEP202605080253) :\n"
            "    1) 결제 수단 81 (포인트)  /  입금 구분 03 (반품)  /  거래 금액 5,000원  ← 포인트만 5,000원 입력\n"
            "    2) 결제 수단 01 (현금)  /  입금 구분 03  /  거래 금액 495,000원\n"
            "  ※ 쿠폰 분개(결제수단 90) 없음 — 시나리오 진행(1) 의도와 일치\n"
            "- 응답 메시지 : 「반품 등록 완료」  /  success: true\n"
            "- 응답 적립 포인트 소멸 :\n"
            "    · CD 01 / 9,900P  (구매포인트 부분 소멸)\n"
            "    · CD 06 / 14,850P  (사은포인트 부분 소멸)\n"
            "    · CD 07 / 9,900P  (이벤트포인트 부분 소멸)\n"
            "- 검증 : 진행 (1) 저장 성공 ✓"
        ),
    ),
    "a12JO000000MHoeYAG": (
        "진행(2-시도)",
        "통반품 시도 — 저장 후 삭제 (포인트 5K + 쿠폰 30K)",
        (
            "반품 번호 : SAL202605080227  /  원거래 판매 번호 : SAL202605080213\n"
            "- 시간 : 2026-05-08 16:02:02 KST\n"
            "- 실결제 금액 : 500,000원 (나머지 1개 품목)\n"
            "- 판매 항목 : 상품 코드 2A2600001 / 수량 -1 / 순매가 단가 500,000\n"
            "- 거래 원장 (입금 번호 DEP202605080258) :\n"
            "    1) 결제 수단 90 (쿠폰)  /  입금 구분 03  /  거래 금액 30,000원  /  쿠폰 COP2026MY0004\n"
            "    2) 결제 수단 81 (포인트)  /  입금 구분 03  /  거래 금액 5,000원\n"
            "    3) 결제 수단 01 (현금)  /  입금 구분 03  /  거래 금액 465,000원\n"
            "- 응답 메시지 : 「반품 등록 완료」  /  success: true  (audit 상 정상 등록)\n"
            "- 응답 적립 포인트 소멸 : CD 01 / 9,300P + CD 06 / 13,950P + CD 07 / 9,300P\n"
            "- ※ 진행 (2-1) 「쿠폰만 20K」, (2-2) 「쿠폰만 30K」 의 저장 시도는 ERP 화면 단계에서 차단된 것으로 추정 (별도 audit 없음)\n"
            "- ※ SAL227 은 이후 16:06:59 에 DELETE 처리됨 — 진행(2-3) 정식 저장 직전 정리 단계로 추정"
        ),
    ),
    "a12JO000000MH78YAG": (
        "진행(2-시도-삭제)",
        "통반품 시도(SAL227) 삭제",
        (
            "Apex 클래스 : GdReturnApiControllerV1 (DELETE)\n"
            "- 시간 : 2026-05-08 16:06:59 KST\n"
            "- 대상 반품 번호 : SAL202605080227\n"
            "- URL : {\"SaleNo__c\":\"SAL202605080227\"} (RequestBody 없음)\n"
            "- 응답 메시지 : 「[판매반품, 교환반품] 전체 삭제 완료」  /  success: true\n"
            "- 검증 : 직전 SAL227 시도 분 정리 → 진행(2-3) 정식 저장으로 진행"
        ),
    ),
    "a12JO000000MDd4YAG": (
        "진행(2-3)",
        "통반품 PASS (포인트 5,000 + 쿠폰 30,000 + 현금 465,000)",
        (
            "반품 번호 : SAL202605080229  /  원거래 판매 번호 : SAL202605080213\n"
            "- 시간 : 2026-05-08 16:07:42 KST  (진행(2-시도-삭제) 후 43초)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 판매 항목 : 상품 코드 2A2600001 / 수량 -1\n"
            "- 거래 원장 (입금 번호 DEP202605080264) :\n"
            "    1) 결제 수단 90 (쿠폰)  /  입금 구분 03  /  거래 금액 30,000원  /  쿠폰 COP2026MY0004 (쿠폰 환불·복원)\n"
            "    2) 결제 수단 81 (포인트)  /  입금 구분 03  /  거래 금액 5,000원  (포인트 사용 복원)\n"
            "    3) 결제 수단 01 (현금)  /  입금 구분 03  /  거래 금액 465,000원\n"
            "- 응답 메시지 : 「반품 등록 완료」  /  success: true\n"
            "- 응답 적립 포인트 소멸 :\n"
            "    · CD 01 / 9,300P (구매포인트 잔여 소멸)\n"
            "    · CD 06 / 13,950P\n"
            "    · CD 07 / 9,300P (이벤트포인트 잔여 소멸)\n"
            "- 시나리오 검증 : 「진행 (2-3) 저장 성공 — 포인트 : 5,000원 / 쿠폰 30,000원 입력 후 저장」 PASS ✓\n"
            "- 예상 결과 :\n"
            "    1) 진행 (1) 저장 성공 ✓ (SAL222)\n"
            "    2) 진행 (2-3) 저장 성공 ✓ (SAL229)\n"
            "    3) 프로모션 구매적립 2배수 정상 적용 ✓ (구매포인트 19,200P + 이벤트 19,200P)\n"
            "    4) 에코포인트 6,000P 정상 적용 ✓ (사전(4))"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 40, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
