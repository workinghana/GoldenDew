"""V5-6 Log.xlsx — 회원 269999800214, 시나리오 V5-6.

「6. 포인트 분할로 사용된 판매 건 부분 반품 진행 시」
- 시간: 2026-05-08 17:16 ~ 17:23 KST (UTC 08:16 ~ 08:23)
- 5건 사전 판매 (A 판매에서 적립된 100만P 를 B·C·D·E 판매에 분할 사용)
- 진행 (6)(7) 부분 반품 시도 — UI 단계 팝업 동작 확인 (audit 없음)
"""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V5-6 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v5_6_v2_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V5-6"
MEMBER_NO = "269999800214"
SHEET_NAME = "V5-6"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac.replace("_", ""):
        return "포인트 (적용 조회)"
    if "saledeposit" in ac.replace("_", ""):
        return "판매 입금"
    if "returndeposit" in ac.replace("_", ""):
        return "반품 입금"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    if isinstance(url, str) and url.startswith("{"):
        return "DELETE"
    return "POST"


SCENARIO = {
    # 1) 1차 판매 - A 판매 (25,000,000원, 적립 100만P)
    "a12JO000000MK08YAG": (
        "1) A 판매",
        "1차 판매 등록 — 구매 금액 25,000,000원 / 포인트 100만원 적립",
        (
            "판매 번호 : SAL202605070290  (시간 : 2026-05-08 17:16:52 KST, EndDate 20260507)\n"
            "- 매장 코드 : 99998 (백화점)\n"
            "- 실결제 금액 : 25,000,000원\n"
            "- 프로모션 번호 : PRO202604161041\n"
            "- 거래 원장 (입금 번호 DEP202605080315) :\n"
            "    1) 결제 수단 01 (현금)  /  입금 구분 02 (잔금)  /  거래 금액 25,000,000원\n"
            "- 응답 메시지 : 「판매 등록 완료」\n"
            "- 응답 적립 포인트 : 1,000,000P (Accrual 「구매 시 포인트 지급」)  ← 100만원 포인트 적립\n"
            "- 검증 : 1) 「1차 판매 건으로 100만원 포인트 발생」 ✓\n"
            "- ※ 직전에 동일 회원으로 SAL202605070289 (40M, 17:14:16) / SAL202605080277 (250M, 17:15:56) 두 건이 등록되었다가 모두 삭제되고 본 SAL290 으로 정리됨"
        ),
    ),
    # 2) B 판매 - 30만 포인트 사용
    "a12JO000000MGkYYAW": (
        "2) B 판매",
        "2차 판매 — 1차 적립 포인트 30만원 사용 (SAL202605080280)",
        (
            "판매 번호 : SAL202605080280  (시간 : 2026-05-08 17:18:10 KST)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 거래 원장 (입금 번호 DEP202605080317) :\n"
            "    1) 결제 수단 01 (현금)  /  입금 구분 02  /  거래 금액 200,000원\n"
            "    2) 결제 수단 81 (포인트)  /  입금 구분 02  /  거래 금액 300,000원  ← A 적립 100만P 중 30만P 사용\n"
            "- 응답 적립 포인트 : 8,000P (현금 200K × 4%)\n"
            "- 응답 차감 포인트 : 300,000P (CD 01)\n"
            "- 검증 : 2) 「B 판매 건에 1차 적립 포인트 30만 사용」 ✓"
        ),
    ),
    # 3) C 판매 - 30만 포인트 사용
    "a12JO000000MK9oYAG": (
        "3) C 판매",
        "3차 판매 — 1차 적립 포인트 30만원 사용 (SAL202605080282)",
        (
            "판매 번호 : SAL202605080282  (시간 : 2026-05-08 17:19:46 KST)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 거래 원장 (입금 번호 DEP202605080320) :\n"
            "    1) 결제 수단 81 (포인트)  /  입금 구분 02  /  거래 금액 300,000원  ← A 적립 잔여 70만P 중 30만P 사용\n"
            "    2) 결제 수단 01 (현금)  /  입금 구분 02  /  거래 금액 200,000원\n"
            "- 응답 적립 포인트 : 8,000P\n"
            "- 응답 차감 포인트 : 300,000P\n"
            "- 검증 : 3) 「C 판매 건에 1차 적립 포인트 30만 사용」 ✓\n"
            "- 누적 사용 : 60만P (A 적립 100만 중)"
        ),
    ),
    # 4) D 판매 - 10만 포인트 사용
    "a12JO000000MHC6YAO": (
        "4) D 판매",
        "4차 판매 — 1차 적립 포인트 10만원 사용 (SAL202605080284)",
        (
            "판매 번호 : SAL202605080284  (시간 : 2026-05-08 17:21:41 KST)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 거래 원장 (입금 번호 DEP202605080322) :\n"
            "    1) 결제 수단 81 (포인트)  /  입금 구분 02  /  거래 금액 100,000원  ← A 적립 잔여 40만P 중 10만P 사용\n"
            "    2) 결제 수단 01 (현금)  /  입금 구분 02  /  거래 금액 400,000원\n"
            "- 응답 적립 포인트 : 16,000P\n"
            "- 응답 차감 포인트 : 100,000P\n"
            "- 검증 : 4) 「D 판매 건에 1차 적립 포인트 10만 사용」 ✓\n"
            "- 누적 사용 : 70만P"
        ),
    ),
    # 5) E 판매 - 10만 포인트 사용
    "a12JO000000MKPwYAO": (
        "5) E 판매",
        "5차 판매 — 1차 적립 포인트 10만원 사용 (SAL202605080286)",
        (
            "판매 번호 : SAL202605080286  (시간 : 2026-05-08 17:23:25 KST)\n"
            "- 실결제 금액 : 500,000원\n"
            "- 거래 원장 (입금 번호 DEP202605080327) :\n"
            "    1) 결제 수단 81 (포인트)  /  입금 구분 02  /  거래 금액 100,000원  ← A 적립 잔여 30만P 중 10만P 사용\n"
            "    2) 결제 수단 01 (현금)  /  입금 구분 02  /  거래 금액 400,000원\n"
            "- 응답 적립 포인트 : 16,000P\n"
            "- 응답 차감 포인트 : 100,000P\n"
            "- 검증 : 5) 「E 판매 건에 1차 적립 포인트 10만 사용」 ✓\n"
            "- 누적 사용 : 80만P / A 적립 잔여 : 20만P"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if r["Id"] in SCENARIO:
                by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Top note row
    note_text = (
        "※ 진행 (6) 「1차 판매 건 부분 반품 — 5만원 포인트만 소멸처리」 시도 + 진행 (7) 「팝업 후 일부 판매 건 포인트 정리 후 부분 반품」 "
        "audit 가 본 캐시(2026-05-08T08:25 ~ 14:59 UTC)에서 발견되지 않음 — UI 단계 팝업/조작 후 부분 반품 미실행 또는 후속 시간대로 추정. "
        "시나리오 검증 「A 판매건에서 적립된 포인트를 사용한 판매번호 목록 모두 뜸」 (확인 완료/정상) — 팝업 동작 정상."
    )

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    note_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.cell(row=1, column=1, value=note_text)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out_headers))
    ws.cell(row=1, column=1).fill = note_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=1, column=1).font = Font(color="9C5700", italic=True)
    ws.row_dimensions[1].height = 60

    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    body_align = Alignment(vertical="top", wrap_text=True)

    sorted_items = sorted(
        SCENARIO.items(),
        key=lambda kv: by_id.get(kv[0], {}).get("CreatedDate", "")
    )

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=3):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 14, 40, 90, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C3"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
