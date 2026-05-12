"""V2-4 Log.xlsx — 회원 103006200153, 시나리오 V2-4 ([TEST] 기획행사_프로모션 / 4-1).

시간 범위: 2026-05-08 13:23:30 ~ 13:33:46 KST  (UTC 04:23 ~ 04:34)
사용 프로모션: PRO202604161046  ([TEST] 기획행사_프로모션)
검증 기준: 구매 포인트 적립 : 10,000P
"""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(os.path.expanduser("~")) / "Downloads"
OUT = DOWNLOADS / "V2-4 Log.xlsx"
MATCHED_CSV = ROOT / "tmp" / "v2_4_matched.csv"

EXCEL_CELL_LIMIT = 32767
SCENARIO_VERSION = "V2-4"
MEMBER_NO = "103006200153"
SHEET_NAME = "V2-4"


def classify_domain(apex_class, url):
    ac = (apex_class or "").lower()
    if "memberhelpinfo" in ac.replace("_", ""):
        return "회원 도움창"
    if "memberapi" in ac.replace("_", ""):
        return "회원"
    if "voucherhelp" in ac.replace("_", ""):
        return "쿠폰 (사용 가능 조회)"
    if "pointhelp" in ac.replace("_", ""):
        return "포인트 (적용 조회)"
    if "orderapi" in ac.replace("_", ""):
        return "주문"
    if "saleapi" in ac.replace("_", ""):
        return "판매"
    if "returnapi" in ac.replace("_", ""):
        return "반품"
    if "pointcredit" in ac.replace("_", ""):
        return "포인트 (지급)"
    if "pointdebit" in ac.replace("_", ""):
        return "포인트 (사용/차감)"
    return ""


def http_method(apex_class, url):
    if (url or "").startswith("callout:"):
        return "CALLOUT"
    return "POST"


# (Salesforce log Id) -> (번호, 확인내용, 확인 기준)
SCENARIO = {
    "a12JO000000MC0xYAG": (
        "4-1-1",
        "포인트 적용 조회 (선행)",
        (
            "Apex 클래스 : GdPointHelpApiControllerV1 — /gd/v1/point/help\n"
            "- 시간 : 2026-05-08 13:24:54 KST\n"
            "- 회원 : 103006200153\n"
            "- 매장 코드 : 99998\n"
            "- 프로모션 번호 : PRO202604161046  ([TEST] 기획행사_프로모션)\n"
            "- 주문 일자 : 20260508\n"
            "- 주문/판매 등록 직전, 사용 가능 포인트 조회를 호출하여 적용 가능 포인트 종류와 잔액을 확인.\n"
            "- 응답 : 적용 가능 포인트 5종 (포인트 유형 코드 01·04·05·06·07) 모두 정상 조회"
        ),
    ),
    "a12JO000000MBxnYAG": (
        "4-1-2",
        "주문 등록 (프로모션 적용 가능 제품 2건)",
        (
            "주문 번호 : SOR202605080124  (Salesforce Id : 801JO00000IAC8nYAH)\n"
            "- 시간 : 2026-05-08 13:25:03 KST\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 프로모션 번호 : PRO202604161046  ([TEST] 기획행사_프로모션)\n"
            "- 주문 항목 : 2건  (각 순매가 단가 500,000)\n"
            "    1) 상품 코드 2A2600001  수량 1  제외 여부 false\n"
            "    2) 상품 코드 3A2600005  수량 1  제외 여부 false\n"
            "- 입금 (입금 번호 DEP202605080106) : 100,000원 (결제 수단 01 현금) + 50,000원 (결제 수단 81 포인트)\n"
            "- 응답 차감 포인트 목록 : 포인트 유형 코드 04 / 사용 포인트 50,000 (포인트 사용)\n"
            "- ※ 이 시점에는 두 제품 모두 프로모션 마스터에서 적용 가능 (구매포인트 적립 대상)"
        ),
    ),
    "a12JO000000MCKKYA4": (
        "4-1-3",
        "판매 저장 (프로모션 마스터에서 적용 가능 제품 1건 삭제 후) — 검증",
        (
            "판매 번호 : SAL202605080117  /  원거래 주문 번호 : SOR202605080124\n"
            "(Salesforce Id : 801JO00000IACOvYAP)\n"
            "- 시간 : 2026-05-08 13:31:33 KST  (주문 등록 후 약 6분 경과 — 그 사이 ERP 프로모션 마스터에서 적용 가능 제품 1건 삭제)\n"
            "- 실결제 금액 : 1,000,000원\n"
            "- 프로모션 번호 : PRO202604161046  ([TEST] 기획행사_프로모션)\n"
            "- 판매 항목 : 2건\n"
            "    1) 상품 코드 212500046  (주문 시 2A2600001 → 판매 단계에서 변경된 코드, 순매가 단가 500,000)\n"
            "    2) 상품 코드 3A2600005  (순매가 단가 500,000)\n"
            "- 거래 원장 (TransactionJournal) :\n"
            "    · 입금 번호 DEP202605080106-1 : 100,000원 (결제 수단 01 현금 / 입금 구분 01)\n"
            "    · 입금 번호 DEP202605080106-2 :  50,000원 (결제 수단 81 포인트 / 입금 구분 01)\n"
            "    · 입금 번호 DEP202605080109-1 : 850,000원 (결제 수단 01 현금 / 입금 구분 02)\n"
            "- 응답 적립 포인트 목록 :\n"
            "    · 적립 사유 「구매포인트 × 2.00」  /  포인트 유형 코드 01  /  적립·차감 구분 Credit\n"
            "    · 적립 포인트 19,000P  /  매출 금액 1,000,000원  /  회원 등급 04\n"
            "- 시나리오 검증 기준 : 「구매 포인트 적립 : 10,000P」\n"
            "- ※ 실제 적립 포인트는 응답의 적립 포인트 값으로 확인 — 프로모션 마스터에서 제품 1건이 삭제되었으므로 적립 대상 금액과 결과 포인트 비교 필요"
        ),
    ),
}


def trim(value):
    if value is None:
        return ""
    s = str(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 50] + "\n...[truncated by export]"
    return s


def main():
    by_id = {}
    with MATCHED_CSV.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            by_id[r["Id"]] = r

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    out_headers = [
        "시나리오 버전",
        "번호",
        "확인내용",
        "확인 기준 (정답지 - 오류 내용 제외)",
        "회원번호",
        "도메인",
        "URL",
        "METHOD",
        "Apex 클래스",
        "로그 ID",
        "생성 일시",
        "요청 본문",
        "응답 본문",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    def num_sort_key(item):
        no = item[1][0]
        parts = no.split("-")
        try:
            return tuple(int(p) for p in parts)
        except ValueError:
            return (99, 99, 99)

    sorted_items = sorted(SCENARIO.items(), key=num_sort_key)

    matched = 0
    for r_idx, (sf_id, (no, label, criterion)) in enumerate(sorted_items, start=2):
        rec = by_id.get(sf_id)
        if not rec:
            print(f"[warn] {no}: log Id {sf_id} not in matched csv")
            continue
        matched += 1
        url = rec.get("RequestUrl__c") or ""
        apex = rec.get("ApexClass__c") or ""
        method = http_method(apex, url)
        domain = classify_domain(apex, url)

        values = [
            SCENARIO_VERSION, no, label, criterion,
            MEMBER_NO,
            domain, trim(url), method, trim(apex),
            trim(rec.get("Id") or ""), trim(rec.get("CreatedDate") or ""),
            trim(rec.get("RequestBody__c") or ""),
            trim(rec.get("ResponseBody__c") or ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.alignment = body_align

    widths = [12, 10, 38, 80, 16, 22, 38, 9, 32, 22, 26, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    wb.save(OUT)
    print(f"[saved] {OUT}")
    print(f"[summary] matched={matched}  total={len(SCENARIO)}")


if __name__ == "__main__":
    main()
